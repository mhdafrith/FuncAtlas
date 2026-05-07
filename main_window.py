"""
main_window.py
──────────────
ReuseAnalysisWindow — the central QMainWindow.
All page-builder logic lives in pages/*.py.
All heavy-lifting lives in services/*.py.
This file only wires navigation, theme, and business logic slots.
"""

import os
import time
import re
import sys
from collections import OrderedDict

from PySide6.QtCore import Qt, QSize, QTimer, QPropertyAnimation, QEasingCurve, QThread
from PySide6.QtGui import QColor, QFont, QPixmap, QDesktopServices, QIcon, QPainter, QBrush, QPalette
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QFrame, QLabel, QPushButton,
    QVBoxLayout, QHBoxLayout, QStackedWidget, QScrollArea, QSizePolicy,
    QGraphicsOpacityEffect, QColorDialog, QFontDialog, QMessageBox,
    QTreeWidgetItem, QTableWidget, QTableWidgetItem, QHeaderView, QMenu,
    QDialog, QTextEdit, QFileDialog, QStyledItemDelegate, QStyleOptionViewItem
)

from core.theme import ThemeManager, VectorIconFactory
from core.utils import (
    normalize_path, normalize_name, clean_text,
    extract_function_body, SCAN_CACHE, SCAN_CACHE_LOCK,
    is_valid_excel_reference
)
from ui.widgets import (
    NavButton, IconTextButton, SectionTitle, PremiumCard,
    StatChip, CollapsiblePanel, add_shadow
)
from ui.dialogs import HelpOverlayDialog, CompletionPopupDialog
from services.analysis import (
    scan_source_for_all_functions, match_target_with_function_list,
    match_target_with_reference_bases, merge_record_sets,
    parse_function_list_files, ConsolidatedWorker,
    extract_functions_from_folder_to_excel,
    detect_best_column_in_workbook
)
from services.report_worker import ReportCompareWorker
from services.analysis import BuiltinExtractionWorker
from services.upfront_worker import UpfrontExtractionWorker
from core.function_cache import FUNCTION_CACHE

# Page builders
from pages.home import create_home_page
from pages.input_page import create_input_page
from pages.reference_page import create_reference_page
from pages.consolidated_page import create_consolidated_page
from pages.view_page import create_view_page
from pages.diff_page import create_diff_page
from pages.report_page import create_report_page
from pages.help_page import create_help_page
from pages.complexity_page import create_complexity_page
from core.logger import get_logger, log_user_action, log_file_upload, log_output_file

_log = get_logger(__name__)


def _load_home_hero_pixmap() -> QPixmap:
    candidates = [
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "image.png"),
        os.path.join(os.path.dirname(os.path.abspath(__file__)),
                     "Gemini_Generated_Image_ps5x2aps5x2aps5x (1).png"),
    ]
    for path in candidates:
        if os.path.exists(path):
            pm = QPixmap(path)
            if not pm.isNull():
                return pm
    return QPixmap()


def _find_function_in_folder(folder: str, function_name: str) -> list:
    matches = []
    for root, _, files in os.walk(folder):
        for file in files:
            if not file.endswith((".c", ".cpp", ".h", ".txt")):
                continue
            full_path = os.path.join(root, file)
            try:
                from core.utils import read_source_file as _read_src
                content = _read_src(full_path)
            except Exception:
                continue
            if re.compile(rf"\b{function_name}\s*\([^;]*\)\s*\{{").search(content):
                matches.append(full_path)
    return matches


# ─────────────────────────────────────────────────────────────────────────────
class ReuseAnalysisWindow(QMainWindow):

    def __init__(self):
        super().__init__()
        self._log = get_logger(self.__class__.__name__)
        self._log.info("ReuseAnalysisWindow initialising")
        self.setWindowTitle("FuncAtlas")
        self.resize(1500, 900)
        self.setMinimumSize(1240, 740)

        # -- Window icon --
        _icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "app_icon.ico")
        if os.path.isfile(_icon_path):
            self.setWindowIcon(QIcon(_icon_path))
            self._log.debug("Window icon set from %s", _icon_path)

        self.default_accent   = QColor("#3BA8FF")
        self.accent_color     = QColor(self.default_accent)
        self.base_font_family = "Segoe UI"
        self.base_font_size   = 12   # point 7: bigger font size throughout
        self.current_theme    = "light"
        self.theme            = ThemeManager.THEMES[self.current_theme]

        self.icons_white = VectorIconFactory(QColor("#FFFFFF"))
        self.icons       = self.icons_white
        self.nav_buttons: dict  = {}
        self.pages: dict        = {}
        self.accent_cards: list = []
        self.home_stat_chips    = []

        self.function_records           = OrderedDict()
        self.current_function_name      = ""
        self.current_function_file      = ""
        self.current_function_body      = ""
        self.current_function_list_paths = []
        self.available_sources          = []
        self.active_source_root         = ""
        self.current_reference_folders  = []
        self._ref_function_filter       = []
        self.con_thread = None
        self.con_worker = None
        self._active_threads: list[QThread] = []  # strong refs to all running threads

        self._report_ext_thread  = None
        self._report_ext_worker  = None
        self._report_cmp_thread  = None
        self._report_cmp_worker  = None
        self._report_output_file = ""
        self._last_report_excel  = ""
        self._last_complexity_html = ""
        self._report_bases_list  = []
        self._report_total_steps = 0
        self._report_done_steps  = 0

        # ── upfront extraction worker (runs at Submit) ────────────────────────
        self._upfront_thread  = None
        self._upfront_worker  = None
        # Folder paths that have already been cached in this session
        # {folder_path: 'target'|'reference'}
        self._cached_folders: dict = {}

        self._setup_ui()
        self.apply_styles()
        self.rebuild_icons()
        self.show_page("home")

    # ── accent card registry ──────────────────────────────────────────────────
    def register_accent_card(self, card: PremiumCard) -> PremiumCard:
        self.accent_cards.append(card)
        return card

    # ── navigation helpers ────────────────────────────────────────────────────
    def wire_animated_navigation(self, button: QPushButton, destination: str):
        if not button:
            return
        try:
            button.clicked.disconnect()
        except Exception:
            pass
        button.clicked.connect(
            lambda checked=False, btn=button, dest=destination:
                self.animate_button_and_navigate(btn, dest)
        )

    def animate_button_and_navigate(self, button: QPushButton, destination: str):
        self.animate_button_click(button, lambda dest=destination: self.show_page(dest))

    def animate_button_click(self, button: QPushButton, callback=None):
        if callback:
            callback()

    def pulse_button(self, button: QPushButton):
        self.animate_button_click(button, None)

    # ── hero image ────────────────────────────────────────────────────────────
    def refresh_home_hero_image(self):
        if not hasattr(self, "home_hero_image"):
            return
        pm = _load_home_hero_pixmap()
        if pm.isNull():
            self.home_hero_image.clear()
            self.home_hero_image.setText("FuncAtlas")
            return
        sz = self.home_hero_image.size()
        w  = sz.width()  if sz.width()  > 10 else 700
        h  = sz.height() if sz.height() > 10 else 360
        scaled = pm.scaled(w, h, Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation)
        self.home_hero_image.setPixmap(scaled)
        self.home_hero_image.setAlignment(Qt.AlignCenter)
        self.home_hero_image.setScaledContents(False)

    def animate_card_entrance(self, cards: list):
        for card in cards:
            card.show()
            card.update()

    def lighten_color(self, color: QColor, factor: int = 112) -> str:
        return color.lighter(factor).name()

    # ── styles ────────────────────────────────────────────────────────────────
    def apply_styles(self):
        t           = self.theme
        accent      = self.accent_color.name()
        accent_hover = QColor(accent).lighter(114).name()
        accent_dark  = QColor(accent).darker(108).name()

        # Auto-contrast: use dark text on light accents, white on dark accents
        r, g, b = self.accent_color.red(), self.accent_color.green(), self.accent_color.blue()
        luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255
        btn_text = "#1a1a1a" if luminance > 0.55 else "white"
        if self.current_theme == "dark":
            glossy_primary = (f"qlineargradient(x1:0,y1:0,x2:0,y2:1,"
                              f" stop:0 {QColor(accent).lighter(114).name()},"
                              f" stop:0.48 {accent}, stop:1 {accent_dark})")
            glossy_surface = (f"qlineargradient(x1:0,y1:0,x2:0,y2:1,"
                              f" stop:0 {t['bg_card']}, stop:1 {t['bg_soft']})")
        else:
            glossy_primary = (f"qlineargradient(x1:0,y1:0,x2:0,y2:1,"
                              f" stop:0 {QColor(accent).lighter(128).name()},"
                              f" stop:0.55 {accent}, stop:1 {accent_dark})")
            glossy_surface = "qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #FFFFFF, stop:1 #EEF4FB)"

        self.setStyleSheet(f"""
            QWidget {{
                background: {t['bg_main']};
                color: {t['text_primary']};
                font-family: "{self.base_font_family}";
                font-size: {self.base_font_size + 1}px;
            }}
            QLabel {{ background: transparent; color: {t['text_primary']}; }}
            QFrame#sidebar {{
                background: {t['bg_sidebar']};
                border-right: 1px solid {t['border']};
            }}
            QFrame#topHeader {{
                background: transparent;
                border: none;
                border-radius: 0px;
            }}
            QFrame#pageCard {{
                background: {t['bg_card']};
                border: 1px solid {t['border']};
                border-radius: 16px;
            }}
            QFrame#heroCard {{
                background: {t['bg_card']};
                border: 1px solid {t['border_strong']};
                border-radius: 20px;
            }}
            QFrame#softPanel {{
                background: {t['bg_soft']};
                border: 1px solid {t['border']};
                border-radius: 12px;
            }}
            QFrame#premiumCard {{
                background: {t['bg_card']};
                border: 1px solid {t['border']};
                border-radius: 18px;
            }}
            QFrame#quickActionsFrame {{
                background: {t['bg_card']};
                border: 1px solid {t['border_strong']};
                border-radius: 20px;
            }}
            QLabel#complexityInfoLabel {{
                color: {t['text_muted']}; font-size: {self.base_font_size}px;
            }}
            QLineEdit#complexityReportDisplay {{
                background: {t['bg_input']}; border: 1px solid {t['border']};
                border-radius: 10px; padding: 8px 12px; color: {t['text_primary']};
            }}
            QLabel#brandTitle {{ color: {t['text_primary']}; font-size: {self.base_font_size+11}px; font-weight: 900; }}
            QLabel#brandSubtitle {{ color: {t['text_secondary']}; font-size: {self.base_font_size-1}px; }}
            QLabel#headerTitle {{ color: {t['text_primary']}; font-size: {self.base_font_size+4}px; font-weight: 900; }}
            QLabel#headerSubtitle {{ color: {t['text_secondary']}; font-size: {self.base_font_size-1}px; }}
            QLabel#sectionTitle {{ color: {t['text_primary']}; font-size: {self.base_font_size+2}px; font-weight: 900; }}
            QLabel#sectionSubtitle {{ color: {t['text_muted']}; font-size: {self.base_font_size-1}px; }}
            QLabel#fieldLabel {{ color: {t['text_secondary']}; font-weight: 800; font-size: {self.base_font_size}px; }}
            QLabel#panelTitle {{ color: {t['text_primary']}; font-weight: 900; font-size: {self.base_font_size+1}px; }}
            QLabel#panelSubtitle {{ color: {t['text_muted']}; font-size: {self.base_font_size-1}px; }}
            QLabel#helpStepTitle {{ color: {t['text_primary']}; font-size: 15px; font-weight: 900; }}
            QLabel#cardTitle {{ color: {t['text_primary']}; font-weight: 900; font-size: {self.base_font_size+1}px; }}
            QLabel#cardSubtitle {{ color: {t['text_secondary']}; font-size: {self.base_font_size-1}px; }}
            QLabel#heroTitle {{ color: {t['text_primary']}; font-size: {self.base_font_size+18}px; font-weight: 900; }}
            QLabel#heroKicker {{ color: {t['text_primary']}; font-size: {self.base_font_size-1}px; font-weight: 900; letter-spacing: 2px; }}
            QLabel#heroSubtitle {{ color: {t['text_secondary']}; font-size: {self.base_font_size+1}px; }}
            QLabel#heroBadge {{
                background: {accent}; color: {btn_text};
                border-radius: 14px; font-size: {self.base_font_size+1}px; font-weight: 900;
            }}
            QLabel#statChipValue {{ color: {accent}; font-size: {self.base_font_size+3}px; font-weight: 900; }}
            QLabel#statChipLabel {{ color: {t['text_muted']}; font-size: {self.base_font_size-2}px; font-weight: 700; }}
            QFrame#statChip {{
                background: {t['bg_soft']}; border: 1px solid {t['border']}; border-radius: 14px;
            }}
            QPushButton {{
                background: {glossy_surface}; color: {t['text_primary']};
                border: 1px solid {t['border']};
                border-radius: 16px; padding: 5px 12px; text-align: left; font-weight: 800;
            }}
            QPushButton:hover {{ border: 1px solid {accent}; }}
            QPushButton:pressed {{ padding-top: 7px; padding-bottom: 3px; }}
            QPushButton:checked {{ background: {glossy_primary}; color: {btn_text}; border: 1px solid {accent_dark}; }}
            QPushButton#smallPrimaryButton {{
                background: {glossy_primary}; color: {btn_text}; border: 1px solid {accent_dark};
                border-radius: 14px; min-height: 32px; min-width: 70px;
                padding: 6px 14px; font-size: {self.base_font_size}px; font-weight: 900; text-align: center;
            }}
            QPushButton#smallPrimaryButton:hover {{ border: 1px solid {accent_hover}; }}
            QPushButton#ghostButton {{
                background: {glossy_surface}; color: {t['text_primary']}; border: 1px solid {t['border_strong']};
                border-radius: 14px; min-height: 32px; min-width: 54px;
                padding: 6px 12px; font-weight: 800; font-size: {self.base_font_size}px;
            }}
            QPushButton#ghostButton:hover {{ border: 1px solid {accent}; }}
            QPushButton#pickerButton {{
                background: {glossy_primary}; color: {btn_text}; border: 1px solid {accent_dark};
                border-radius: 13px; min-height: 34px; min-width: 108px;
                padding: 6px 14px; font-weight: 900; font-size: {self.base_font_size}px; text-align: center;
            }}
            QPushButton#pickerButton:hover {{ border: 1px solid {accent_hover}; background: {glossy_primary}; }}
            QPushButton#pickerButton:disabled {{
                background: {t['bg_soft']}; color: {t['text_muted']};
                border: 1px solid {t['border']};
            }}
            QPushButton#clearButton {{
                background: {glossy_primary}; color: {btn_text}; border: 1px solid {accent_dark};
                border-radius: 13px; min-height: 34px; min-width: 80px;
                padding: 6px 14px; font-weight: 900; font-size: {self.base_font_size}px; text-align: center;
            }}
            QPushButton#clearButton:hover {{ border: 1px solid {accent_hover}; background: {glossy_primary}; }}
            QPushButton#clearButton:pressed {{ background: {accent_dark}; }}
            QPushButton#clearButtonRect {{
                background: {glossy_primary}; color: {btn_text}; border: 1px solid {accent_dark};
                border-radius: 13px; min-height: 34px; min-width: 80px;
                padding: 6px 14px; font-weight: 900; font-size: {self.base_font_size}px; text-align: center;
            }}
            QPushButton#clearButtonRect:hover {{ border: 1px solid {accent_hover}; background: {glossy_primary}; }}
            QPushButton#clearButtonRect:pressed {{ background: {accent_dark}; }}
            QPushButton#toggleLeft {{
                background: {t['bg_soft']}; color: {t['text_muted']};
                border: 1px solid {accent_dark};
                border-top-left-radius: 10px; border-bottom-left-radius: 10px;
                border-top-right-radius: 0px; border-bottom-right-radius: 0px;
                min-width: 160px; padding: 4px 14px; font-weight: 800;
                font-size: {self.base_font_size}px;
            }}
            QPushButton#toggleLeft:checked {{
                background: {accent}; color: {btn_text};border: 2px solid {accent_hover};
            }}
            QPushButton#toggleRight {{
                background: {t['bg_soft']}; color: {t['text_muted']};
                border: 1px solid {accent_dark};
                border-top-right-radius: 10px; border-bottom-right-radius: 10px;
                border-top-left-radius: 0px; border-bottom-left-radius: 0px;
                min-width: 140px; padding: 4px 14px; font-weight: 800;
                font-size: {self.base_font_size}px;
            }}
            QPushButton#toggleRight:checked {{
                background: {accent}; color: {btn_text}; border: 2px solid {accent_hover};
            }}
            QWidget#funcColField QLineEdit:disabled {{
                background: {t['bg_soft']}; color: {t['text_muted']};
                border: 1px solid {t['border']};
            }}
            QWidget#funcColField QPushButton:disabled {{
                background: {t['bg_soft']}; color: {t['text_muted']};
                border: 1px solid {t['border']};
            }}
            /* ── 2-colour scheme: Blue = Function inputs, Green = Consolidated DB inputs ── */
            /* Function List Files box */
            QWidget#funcListField QTextEdit {{
                border: 2px solid #2196F3; border-radius: 10px;
            }}
            QWidget#funcListField QLabel#fieldLabel {{ color: #1565C0; font-weight: 900; }}
            /* Function Name Column box */
            QWidget#funcColField QLineEdit {{
                border: 2px solid #2196F3; border-radius: 10px;
            }}
            QWidget#funcColField QLabel#fieldLabel {{ color: #1565C0; font-weight: 900; }}
            /* Consolidated DB Excel File box */
            QWidget#dbExcelField QTextEdit, QWidget#dbExcelField QLineEdit {{
                border: 2px solid #43A047; border-radius: 10px;
            }}
            QWidget#dbExcelField QLabel#fieldLabel {{ color: #2E7D32; font-weight: 900; }}
            /* Base File Name Column box */
            QWidget#baseColField QLineEdit {{
                border: 2px solid #43A047; border-radius: 10px;
            }}
            QWidget#baseColField QLabel#fieldLabel {{ color: #2E7D32; font-weight: 900; }}
            QPushButton#collapseToggle {{
                background: {t['bg_soft']}; color: {t['text_primary']}; border: 1px solid {t['border']};
                border-radius: 12px; padding: 8px 14px; text-align: left; font-weight: 800;
            }}
            QLineEdit, QTextEdit {{
                background: {t['bg_input']}; color: {t['text_primary']};
                border: 1px solid {t['border']}; border-radius: 10px; padding: 6px 10px;
            }}
            QLineEdit:focus, QTextEdit:focus {{ border: 1px solid {accent}; }}
            QComboBox {{
                background: {t['bg_input']}; color: {t['text_primary']};
                border: 1px solid {t['border']}; border-radius: 10px; padding: 5px 10px; font-weight: 700;
            }}
            QComboBox:focus {{ border: 1px solid {accent}; }}
            QTreeWidget {{
                background: {t['bg_input']}; color: {t['text_primary']};
                border: 1px solid {t['border']}; border-radius: 10px;
            }}
            QTreeWidget::item:selected {{ background: {accent}55; color: {t['text_primary']}; }}
            QTableWidget {{
                background: {t['bg_card']}; color: {t['text_primary']};
                border: 1px solid {t['border']}; border-radius: 8px;
                gridline-color: {t['border']};
            }}
            QHeaderView::section {{
                background: {t['bg_soft']}; color: {t['text_primary']};
                border: 1px solid {t['border']}; padding: 4px 8px;
                font-weight: 700;
            }}
            QTabWidget::pane {{
                border: 1px solid {t['border']}; border-radius: 8px;
                background: {t['bg_card']};
            }}
            QTabBar::tab {{
                background: {t['bg_soft']}; color: {t['text_secondary']};
                border: 1px solid {t['border']}; border-bottom: none;
                border-top-left-radius: 8px; border-top-right-radius: 8px;
                padding: 6px 14px; font-weight: 700; min-width: 120px;
            }}
            QTabBar::tab:selected {{
                background: {t['bg_card']}; color: {t['text_primary']};
                border-bottom: 2px solid {accent};
            }}
            QTabBar::tab:hover:!selected {{
                background: {t['bg_input']}; color: {t['text_primary']};
            }}
            QScrollBar:vertical {{ background: transparent; width: 10px; margin: 2px; }}
            QScrollBar::handle:vertical {{ background: {t['border_strong']}; min-height: 28px; border-radius: 5px; }}
            QScrollBar::handle:vertical:hover {{ background: {accent}; }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0px; background: transparent; border: none; }}
            QProgressBar {{
                background: {t['bg_input']}; border: 1px solid {t['border']}; border-radius: 4px;
                text-align: center; height: 8px; font-weight: 800; color: {t['text_primary']};
            }}
            QProgressBar::chunk {{ background: {glossy_primary}; border-radius: 3px; margin: 1px; }}
            QProgressBar#reportOverallProgress {{
                background: {t['bg_input']}; border: 1px solid {t['border']}; border-radius: 4px; height: 8px;
            }}
            QProgressBar#reportOverallProgress::chunk {{ background: {glossy_primary}; border-radius: 3px; margin: 1px; }}
            QProgressBar#reportStepProgress {{
                background: {t['bg_input']}; border: none; border-radius: 2px; height: 5px;
            }}
            QProgressBar#reportStepProgress::chunk {{ background: {accent}88; border-radius: 2px; }}
            QSplitter::handle {{ background: {t['bg_soft']}; border-radius: 4px; }}
        """)

        if hasattr(self, "header_status"):
            self.header_status.setStyleSheet(
                f"background: transparent; color: {accent}; border: 1px solid {t['border_strong']};"
                f" border-radius: 12px; padding: 6px 12px; font-weight: 900;"
            )
        if hasattr(self, "theme_btn"):
            self.theme_btn.setText("☀ Light" if self.current_theme == "dark" else "Dark")
        if hasattr(self, "settings_theme_combo"):
            self.settings_theme_combo.blockSignals(True)
            self.settings_theme_combo.setCurrentIndex(0 if self.current_theme == "dark" else 1)
            self.settings_theme_combo.blockSignals(False)
        if hasattr(self, "theme_note_label"):
            self.theme_note_label.setText(f"Current theme: {t['name']} • Accent: {accent}")
        if hasattr(self, "report_pct_label"):
            self.report_pct_label.setStyleSheet(
                f"color: {accent}; font-weight: 900; font-size: {self.base_font_size+1}px; background: transparent;"
            )
        if hasattr(self, "report_phase_label"):
            self.report_phase_label.setStyleSheet(
                f"color: {accent}; font-weight: 900; font-size: {self.base_font_size+1}px; background: transparent;"
            )
        for card in self.accent_cards:
            try:
                card.update_accent(accent)
                card._shadow = add_shadow(card, blur=28, y_offset=9, alpha=80)
            except Exception:
                pass

        # Refresh active nav button label color for new accent
        if hasattr(self, "nav_buttons"):
            for btn in self.nav_buttons.values():
                if btn.isChecked():
                    btn._on_toggled(True)

    def _update_progress_btn_colors(self):
        """Push current accent color into every ProgressButton palette so
        their custom paintEvent picks up the right color."""
        from PySide6.QtGui import QPalette
        from ui.widgets import ProgressButton
        accent = self.accent_color
        # Walk all child widgets and update any ProgressButton
        for btn in self.findChildren(ProgressButton):
            pal = btn.palette()
            pal.setColor(QPalette.Button, accent)
            btn.setPalette(pal)
            btn.update()

    def rebuild_icons(self):
        icon_color = QColor("#FFFFFF") if self.current_theme == "dark" else QColor("#102134")
        # For white icons on nav buttons, use contrast vs accent color
        r, g, b = self.accent_color.red(), self.accent_color.green(), self.accent_color.blue()
        lum = (0.299 * r + 0.587 * g + 0.114 * b) / 255
        nav_icon_color = QColor("#1a1a1a") if lum > 0.55 else QColor("#FFFFFF")
        self.icons = VectorIconFactory(icon_color)
        self.icons_white = VectorIconFactory(nav_icon_color)
        icon_map = {
            "home": "home", "input": "input", "view": "view",
            "diff": "diff", "report": "report", "complexity": "settings",
            "help": "help",
        }
        for key, btn in self.nav_buttons.items():
            btn.setIcon(self.icons.icon(icon_map[key], 18))

        for field_attr, icon_name in [
            ("ref_target_field",      "folder"),
            ("ref_bases_field",       "folder"),
            ("ref_function_field",    "document"),
            ("con_function_field",    "document"),
            ("con_folder_field",      "folder"),
            ("con_db_excel_field",    "excel"),
            ("con_func_col_field",    "column"),
            ("con_base_col_field",    "column"),
            ("con_output_link_field", "link"),
        ]:
            if hasattr(self, field_attr) and hasattr(getattr(self, field_attr), "button"):
                getattr(self, field_attr).button.setIcon(self.icons.icon(icon_name, 15))

        if hasattr(self, "_report_browse_btn"):
            self._report_browse_btn.setIcon(self.icons.icon("folder", 15))

        for attr, icon_name in [
            ("settings_accent_btn", "palette"),
            ("settings_font_btn",   "font"),
            ("settings_reset_btn",  "reset"),
            ("con_open_output_btn", "link"),
        ]:
            if hasattr(self, attr):
                getattr(self, attr).setIcon(self.icons.icon(icon_name, 15 if attr != "theme_btn" else 18))

        for attr, icon_name in [
            ("ref_submit_btn", "submit"), ("ref_back_btn", "back"),   ("ref_clear_btn", "clear"),
            ("con_submit_btn", "submit"), ("con_back_btn", "back"),   ("con_clear_btn", "clear"),
            ("report_generate_btn", "submit"), ("report_clear_btn", "clear"),
            ("report_open_btn", "link"),
        ]:
            if hasattr(self, attr):
                getattr(self, attr).setIcon(self.icons.icon(icon_name, 15))

    # ── theme actions ─────────────────────────────────────────────────────────
    def set_theme_mode(self, mode: str):
        mode = (mode or "dark").lower()
        if mode not in ThemeManager.THEMES:
            mode = "dark"
        log_user_action("toggle", f"Theme → {mode}", page="settings")
        self.current_theme = mode
        self.theme = ThemeManager.THEMES[mode]
        self.apply_styles()
        self.rebuild_icons()
        self._update_progress_btn_colors()
        self._update_color_btn_swatch()
        self._update_help_badges()
        # Refresh palette + header on all live diff tables so they reflect new theme
        for tbl in getattr(self, "_diff_tables", []):
            try:
                if hasattr(tbl, "_apply_palette"):
                    tbl._apply_palette(tbl, self.theme)
                    tbl.viewport().update()
                    tbl.update()
            except RuntimeError:
                pass  # table was deleted

    def toggle_theme(self):
        self.set_theme_mode("light" if self.current_theme == "dark" else "dark")

    def on_theme_combo_changed(self, idx: int):
        self.set_theme_mode("dark" if idx == 0 else "light")

    def pick_accent_color(self):
        """Open a QColorDialog with a proper standalone Reset Color button."""
        from PySide6.QtWidgets import (
            QDialogButtonBox, QDialog, QVBoxLayout, QHBoxLayout as _QHBox
        )

        dlg = QColorDialog(self.accent_color, self)
        dlg.setWindowTitle("Choose Accent Color")
        dlg.setOption(QColorDialog.DontUseNativeDialog, True)

        # ── Find the dialog's top-level layout and button box ─────────────────
        btn_box = dlg.findChild(QDialogButtonBox)

        # Build a standalone Reset button with fixed dimensions
        reset_btn = QPushButton("↺  Reset Color")
        reset_btn.setFixedHeight(36)
        reset_btn.setMinimumWidth(130)
        reset_btn.setStyleSheet(
            "QPushButton {"
            "  border: 1px solid #888; border-radius: 6px;"
            "  padding: 4px 14px; font-weight: 700; font-size: 13px;"
            "  background: #2a2a2a; color: #ffffff;"
            "}"
            "QPushButton:hover { background: #3a3a3a; border-color: #aaa; }"
            "QPushButton:pressed { background: #1a1a1a; }"
        )
        reset_btn.setToolTip("Reset accent to default blue (#3BA8FF) and apply immediately")

        def _do_reset():
            self.accent_color = QColor(self.default_accent)
            self.apply_styles()
            self.rebuild_icons()
            self._update_progress_btn_colors()
            self._update_color_btn_swatch()
            self._update_help_badges()
            dlg.reject()

        reset_btn.clicked.connect(_do_reset)

        # ── Inject reset button into a new row ABOVE the OK/Cancel row ────────
        if btn_box is not None:
            parent_layout = btn_box.parent().layout() if btn_box.parent() else dlg.layout()
            if parent_layout is not None:
                # Wrap reset in its own left-aligned row
                row = _QHBox()
                row.setContentsMargins(0, 4, 0, 0)
                row.addWidget(reset_btn)
                row.addStretch()
                idx = parent_layout.indexOf(btn_box)
                parent_layout.insertLayout(idx, row)

        if dlg.exec() == QColorDialog.Accepted:
            color = dlg.selectedColor()
            if color.isValid():
                self.accent_color = color
                self.apply_styles()
                self.rebuild_icons()
                self._update_progress_btn_colors()
                self._update_color_btn_swatch()
                self._update_help_badges()

    def reset_accent_color(self):
        """Reset only the accent color to the default, keeping theme and data intact."""
        self.accent_color = QColor(self.default_accent)
        self.apply_styles()
        self.rebuild_icons()
        self._update_progress_btn_colors()
        self._update_color_btn_swatch()
        self._update_help_badges()

    def _update_color_btn_swatch(self):
        """Update the Theme Color button to show the current accent color as a swatch."""
        if not hasattr(self, "color_btn"):
            return
        c = self.accent_color.name()
        self.color_btn.setText(f"Theme Color")
        self.color_btn.setStyleSheet(
            f"border-left: 5px solid {c};"
        )

    def _update_help_badges(self):
        """Repaint all help page step badges with the current accent color."""
        if not hasattr(self, "_help_badges"):
            return
        c = self.accent_color.name()
        r, g, b = self.accent_color.red(), self.accent_color.green(), self.accent_color.blue()
        lum = (0.299 * r + 0.587 * g + 0.114 * b) / 255
        txt = "#1a1a1a" if lum > 0.55 else "white"
        for badge in self._help_badges:
            badge.setStyleSheet(
                f"background: {c}; color: {txt}; "
                f"border-radius: 10px; font-size: 16px; font-weight: 900;"
            )

    def pick_font(self):
        font, ok = QFontDialog.getFont(QFont(self.base_font_family, self.base_font_size), self, "Choose Font")
        if ok:
            self.base_font_family = font.family()
            pt = font.pointSize() if font.pointSize() > 0 else 10
            self.base_font_size = max(9, min(pt, 14))
            self.apply_styles()
            self.rebuild_icons()

    def reset_theme(self):
        self.accent_color     = QColor(self.default_accent)
        self.base_font_family = "Segoe UI"
        self.base_font_size   = 10
        self.set_theme_mode("dark")
        self._update_color_btn_swatch()

    def reset_all(self):
        """Hard reset: clear all cache, progress, and form state across every page."""
        from PySide6.QtWidgets import QMessageBox
        reply = QMessageBox.question(
            self, "Reset Everything",
            "This will clear all loaded data, cache, and form inputs.\nAre you sure?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        log_user_action("click", "Reset All", extra="user confirmed full reset")

        # ── Clear scan cache ──────────────────────────────────────────────────
        with SCAN_CACHE_LOCK:
            SCAN_CACHE.clear()

        # ── Clear function body disk cache ────────────────────────────────────
        FUNCTION_CACHE.clear()
        self._cached_folders = {}

        # ── Reset all core state ──────────────────────────────────────────────
        self.function_records            = OrderedDict()
        self.current_function_name       = ""
        self.current_function_file       = ""
        self.current_function_body       = ""
        self.current_function_list_paths = []
        self.available_sources           = []
        self.active_source_root          = ""
        self.current_reference_folders   = []
        self._ref_function_filter        = []

        # ── Clear reference form ──────────────────────────────────────────────
        if hasattr(self, "ref_target_field"):
            self.ref_target_field.clear_selection()
        if hasattr(self, "ref_bases_field"):
            self.ref_bases_field.clear_selection()
        if hasattr(self, "ref_function_field"):
            self.ref_function_field.clear_selection()

        # ── Clear consolidated form ───────────────────────────────────────────
        if hasattr(self, "con_function_field"):
            self.con_function_field.clear_selection()
        if hasattr(self, "con_db_excel_field"):
            self.con_db_excel_field.clear_selection()
        if hasattr(self, "con_func_col_field"):
            self.con_func_col_field.clear_selection()
        if hasattr(self, "con_base_col_field"):
            self.con_base_col_field.clear_selection()
        if hasattr(self, "con_output_link_field"):
            self.con_output_link_field.clear_selection()
        if hasattr(self, "con_folder_field"):
            self.con_folder_field.clear_selection()

        # ── Reset view page ───────────────────────────────────────────────────
        if hasattr(self, "source_combo"):
            self.source_combo.blockSignals(True)
            self.source_combo.clear()
            self.source_combo.blockSignals(False)
        if hasattr(self, "tree"):
            self.tree.clear()
        if hasattr(self, "view_text"):
            self.view_text.setText("")
        if hasattr(self, "view_title"):
            self.view_title.setText("")
        if hasattr(self, "view_meta"):
            self.view_meta.setText("")
        if hasattr(self, "view_mode_chip"):
            self.view_mode_chip.setVisible(False)

        # ── Reset diff page (including raw data so filters stay empty) ────────
        self._diff_raw_data = []
        self._clear_diff()

        # ── Reset report fields + all progress UI ────────────────────────────────
        if hasattr(self, "report_output_field"):
            self.report_output_field.clear_selection()
        if hasattr(self, "report_target_field"):
            self.report_target_field.clear_selection()
        # Delegate to _on_report_clear which resets all progress widgets,
        # step list, stat chips, labels, log box and Open Report button.
        if hasattr(self, "_on_report_clear"):
            self._on_report_clear()
        # Explicitly clear the output folder display and path list
        if hasattr(self, "_report_folder_display"):
            self._report_folder_display.clear()
        if hasattr(self, "_report_folder_path"):
            self._report_folder_path.clear()
        # Clear cached output paths
        self._report_output_file = ""
        self._last_report_excel  = ""

        # ── Reset Complexity & Compatibility page ─────────────────────────────
        if hasattr(self, "complexity_report_display"):
            self.complexity_report_display.clear()
        if hasattr(self, "complexity_progress_bar"):
            self.complexity_progress_bar.setValue(0)
            self.complexity_progress_bar.setVisible(False)
        if hasattr(self, "complexity_status_lbl"):
            self.complexity_status_lbl.setText("")
            self.complexity_status_lbl.setVisible(False)
        if hasattr(self, "complexity_log"):
            self.complexity_log.clear()
            self.complexity_log.setVisible(False)
        if hasattr(self, "complexity_open_report_btn"):
            self.complexity_open_report_btn.setEnabled(False)
        if hasattr(self, "complexity_timer_lbl"):
            self.complexity_timer_lbl.setText("")
            self.complexity_timer_lbl.setVisible(False)
        if hasattr(self, "_cx_timer"):
            self._cx_timer.stop()
        self._cx_thread = None
        self._cx_worker = None

        QMessageBox.information(self, "Reset Complete", "All data and cache have been cleared.")

    # ── form clear ────────────────────────────────────────────────────────────
    def clear_reference_form(self):
        self.ref_target_field.clear_selection()
        self.ref_bases_field.clear_selection()
        self.ref_function_field.clear_selection()
        self._ref_function_filter = []
        # Point 6: auto-clear diff when inputs are cleared
        self._clear_diff()
        self.available_sources = []
        self.current_function_list_paths = []
        self.current_reference_folders = []
        self.function_records = OrderedDict()
        # Clear disk cache for this session
        FUNCTION_CACHE.clear()
        self._cached_folders = {}
        if hasattr(self, "source_combo"):
            self.source_combo.blockSignals(True)
            self.source_combo.clear()
            self.source_combo.blockSignals(False)
        if hasattr(self, "view_mode_chip"):
            self.view_mode_chip.setVisible(False)

    def clear_consolidated_form(self):
        self.con_function_field.clear_selection()
        self.con_db_excel_field.clear_selection()
        self.con_func_col_field.clear_selection()
        self.con_base_col_field.clear_selection()
        self.con_output_link_field.clear_selection()
        if hasattr(self, "con_folder_field"):
            self.con_folder_field.clear_selection()
        # Reset toggle back to Excel mode
        if hasattr(self, "con_toggle_excel_btn"):
            self.con_toggle_excel_btn.setChecked(True)
            self.con_source_stack.setCurrentIndex(0)
            self.con_func_col_field.setEnabled(True)

    # ── UI scaffold ───────────────────────────────────────────────────────────
    def _setup_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QHBoxLayout(central)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # Sidebar
        self.sidebar = QFrame()
        self.sidebar.setObjectName("sidebar")
        self.sidebar.setFixedWidth(230)
        sidebar_layout = QVBoxLayout(self.sidebar)
        sidebar_layout.setContentsMargins(12, 16, 12, 16)
        sidebar_layout.setSpacing(8)

        brand_wrap   = QWidget()
        brand_wrap.setStyleSheet("background: transparent;")
        brand_layout = QHBoxLayout(brand_wrap)
        brand_layout.setContentsMargins(4, 0, 0, 0)
        brand_layout.setSpacing(10)

        # ── Logo icon ────────────────────────────────────────────────────────
        _icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "app_icon.ico")
        brand_logo_lbl = QLabel()
        brand_logo_lbl.setFixedSize(38, 38)
        brand_logo_lbl.setStyleSheet("background: transparent;")
        if os.path.isfile(_icon_path):
            _pm = QPixmap(_icon_path).scaled(38, 38, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            brand_logo_lbl.setPixmap(_pm)
        brand_layout.addWidget(brand_logo_lbl)

        # ── Text block ───────────────────────────────────────────────────────
        brand_text_wrap = QWidget()
        brand_text_wrap.setStyleSheet("background: transparent;")
        brand_text_layout = QVBoxLayout(brand_text_wrap)
        brand_text_layout.setContentsMargins(0, 0, 0, 0)
        brand_text_layout.setSpacing(1)
        brand_title    = QLabel("FuncAtlas")
        brand_title.setObjectName("brandTitle")
        # brand_subtitle = QLabel("Enterprise Desktop Suite")
        # brand_subtitle.setObjectName("brandSubtitle")
        brand_text_layout.addWidget(brand_title)
        # brand_text_layout.addWidget(brand_subtitle)
        brand_layout.addWidget(brand_text_wrap, 1)

        sidebar_layout.addWidget(brand_wrap)

        nav_items = [
            ("home",       "Home",                    "home"),
            ("input",      "Input",                   "input"),
            ("view",       "View",                    "view"),
            ("diff",       "Diff",                    "diff"),
            ("report",     "Report",                  "report"),
            ("complexity", "Complexity & Compatibility", "settings"),
            ("help",       "Help",                    "help"),
        ]
        for key, text, icon_name in nav_items:
            btn = NavButton(text, self.icons.icon(icon_name, 18))
            btn.clicked.connect(lambda checked=False, k=key: self.show_page(k))
            sidebar_layout.addWidget(btn, alignment=Qt.AlignLeft)
            self.nav_buttons[key] = btn
        sidebar_layout.addStretch()
        root.addWidget(self.sidebar)

        # Right content
        right_wrap   = QWidget()
        right_layout = QVBoxLayout(right_wrap)
        right_layout.setContentsMargins(12, 12, 12, 12)
        right_layout.setSpacing(10)

        self.header = QFrame()
        self.header.setObjectName("topHeader")
        self.header.setFixedHeight(82)
        header_layout = QHBoxLayout(self.header)
        header_layout.setContentsMargins(18, 12, 18, 12)

        header_left = QWidget()
        hl_layout   = QVBoxLayout(header_left)
        hl_layout.setContentsMargins(0, 0, 0, 0)
        hl_layout.setSpacing(1)
        self.header_title    = QLabel("Dashboard")
        self.header_title.setObjectName("headerTitle")
        self.header_subtitle = QLabel("Professional reusable function analysis workspace")
        self.header_subtitle.setObjectName("headerSubtitle")
        hl_layout.addWidget(self.header_title)
        hl_layout.addWidget(self.header_subtitle)
        header_layout.addWidget(header_left)
        header_layout.addStretch()

        self.theme_btn = IconTextButton("Dark", QIcon())
        self.theme_btn.setObjectName("pickerButton")
        self.theme_btn.setMinimumWidth(118)
        self.theme_btn.setFixedHeight(38)
        self.theme_btn.clicked.connect(self.toggle_theme)

        self.color_btn = IconTextButton("Theme Color", QIcon())
        self.color_btn.setObjectName("pickerButton")
        self.color_btn.setMinimumWidth(138)
        self.color_btn.setFixedHeight(38)
        self.color_btn.clicked.connect(self.pick_accent_color)

        self.reset_all_btn = IconTextButton("↺ Reset", QIcon())
        self.reset_all_btn.setObjectName("pickerButton")
        self.reset_all_btn.setMinimumWidth(118)
        self.reset_all_btn.setFixedHeight(38)
        self.reset_all_btn.clicked.connect(self.reset_all)

        header_layout.addWidget(self.theme_btn)
        header_layout.addSpacing(8)
        header_layout.addWidget(self.color_btn)
        header_layout.addSpacing(8)
        header_layout.addWidget(self.reset_all_btn)
        right_layout.addWidget(self.header)

        self.stack = QStackedWidget()
        right_layout.addWidget(self.stack)
        root.addWidget(right_wrap)

        # Build all pages
        create_home_page(self)
        create_input_page(self)
        create_reference_page(self)
        create_consolidated_page(self)
        create_view_page(self)
        create_diff_page(self)
        create_report_page(self)
        create_complexity_page(self)
        create_help_page(self)

        # ── Global loading overlay (covers entire window) ─────────────────────
        self._setup_global_overlay()

    def _setup_global_overlay(self):
        """Create a full-window dim + spinner overlay parented to centralWidget."""
        central = self.centralWidget()

        self._global_overlay = QWidget(central)
        self._global_overlay.setObjectName("globalOverlay")
        self._global_overlay.setAttribute(Qt.WA_TransparentForMouseEvents, False)
        self._global_overlay.setVisible(False)

        # Semi-transparent black dim
        self._global_overlay_dim = QLabel(self._global_overlay)
        self._global_overlay_dim.setStyleSheet(
            "background: rgba(0,0,0,150); border-radius: 0px;"
        )

        # Centered spinner card
        spinner_card = QFrame(self._global_overlay)
        spinner_card.setObjectName("globalSpinnerCard")
        spinner_card.setFixedSize(160, 160)
        spinner_card.setStyleSheet(
            "QFrame#globalSpinnerCard {"
            "  background: rgba(255,255,255,235);"
            "  border-radius: 18px;"
            "}"
        )
        sc_layout = QVBoxLayout(spinner_card)
        sc_layout.setContentsMargins(16, 16, 16, 16)
        sc_layout.setSpacing(10)
        sc_layout.setAlignment(Qt.AlignCenter)

        self._global_spinner_lbl = QLabel("⠋")
        self._global_spinner_lbl.setAlignment(Qt.AlignCenter)
        self._global_spinner_lbl.setStyleSheet(
            "font-size: 42px; color: #1565C0; background: transparent;"
        )
        self._global_loading_txt = QLabel("Please wait…")
        self._global_loading_txt.setAlignment(Qt.AlignCenter)
        self._global_loading_txt.setStyleSheet(
            "font-size: 13px; font-weight: 700; color: #333; background: transparent;"
        )
        sc_layout.addWidget(self._global_spinner_lbl)
        sc_layout.addWidget(self._global_loading_txt)
        self._global_spinner_card = spinner_card

        _FRAMES = ["⠋", "⠙", "⠹", "⠸", "⠼", "⠴", "⠦", "⠧", "⠇", "⠏"]
        self._global_spinner_idx = 0
        self._global_spinner_timer = QTimer(self)
        self._global_spinner_timer.setInterval(80)

        def _tick():
            self._global_spinner_idx = (self._global_spinner_idx + 1) % len(_FRAMES)
            self._global_spinner_lbl.setText(_FRAMES[self._global_spinner_idx])

        self._global_spinner_timer.timeout.connect(_tick)

        def _resize_global_overlay():
            w, h = central.width(), central.height()
            self._global_overlay.setGeometry(0, 0, w, h)
            self._global_overlay_dim.setGeometry(0, 0, w, h)
            cx = (w - spinner_card.width()) // 2
            cy = (h - spinner_card.height()) // 2
            spinner_card.move(cx, cy)

        self._global_overlay.resizeEvent = lambda e: _resize_global_overlay()
        _orig_resize = central.resizeEvent if hasattr(central, "resizeEvent") else None

        def _central_resize(e):
            _resize_global_overlay()
            if _orig_resize:
                _orig_resize(e)
            else:
                e.accept()

        central.resizeEvent = _central_resize

    def show_loading(self, message="Please wait…"):
        """Show the global full-window loading overlay."""
        central = self.centralWidget()
        w, h = central.width(), central.height()
        self._global_overlay.setGeometry(0, 0, w, h)
        self._global_overlay_dim.setGeometry(0, 0, w, h)
        cx = (w - self._global_spinner_card.width()) // 2
        cy = (h - self._global_spinner_card.height()) // 2
        self._global_spinner_card.move(cx, cy)
        self._global_loading_txt.setText(message)
        self._global_overlay.raise_()
        self._global_overlay.setVisible(True)
        self._global_spinner_timer.start()
        from PySide6.QtWidgets import QApplication
        QApplication.processEvents()

    def hide_loading(self):
        """Hide the global loading overlay."""
        self._global_spinner_timer.stop()
        self._global_overlay.setVisible(False)

    # ── nav lock helpers (called while any worker is running) ─────────────────
    def _lock_nav(self):
        """Disable all sidebar nav buttons while a background job is running."""
        for btn in self.nav_buttons.values():
            btn.setEnabled(False)

    def _unlock_nav(self):
        """Re-enable all sidebar nav buttons after a background job finishes."""
        for btn in self.nav_buttons.values():
            btn.setEnabled(True)

    # ── page navigation ───────────────────────────────────────────────────────
    def make_scroll_page(self, content_widget: QWidget) -> QScrollArea:
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setContentsMargins(0, 0, 0, 0)
        scroll.setWidget(content_widget)
        return scroll

    def set_header(self, title: str, subtitle: str):
        self.header_title.setText(title)
        self.header_subtitle.setText(subtitle)

    def show_page(self, page_name: str):
        headers = {
            "home":         ("Dashboard",            "Professional overview of the FuncAtlas workspace"),
            "input":        ("Input Workspace",       "Choose the correct data input flow"),
            "reference":    ("Reference Bases",       "Upload target base folder, reference folders, and function list file"),
            "consolidated": ("Consolidated Database", "Upload function list files, consolidated Excel, detect columns, and auto-generate output"),
            "view":         ("Function Explorer",     "Browse selected source files and click functions to preview body"),
            "diff":         ("Diff Workspace",        "Compare logic and reusable code structures"),
            "report":       ("Report Generator",           "Extracting Functions, Comparing and Generating Reports"),
            "complexity":   ("Complexity & compatibility",  "Review the generated report and manage complexity settings"),
            "help":         ("Help & Guidance",             "Instructions, workflow rules, and user support"),
        }
        active_nav = {
            "home": "home", "input": "input", "reference": "input", "consolidated": "input",
            "view": "view", "diff": "diff", "report": "report",
            "complexity": "complexity",
            "help": "help",
        }
        for btn in self.nav_buttons.values():
            btn.setChecked(False)
        self.nav_buttons[active_nav[page_name]].setChecked(True)

        widget = self.pages[page_name]
        self.stack.setCurrentWidget(widget)
        self.stack.updateGeometry()
        widget.updateGeometry()
        widget.adjustSize()
        widget.update()
        QApplication.processEvents()

        effect = QGraphicsOpacityEffect(widget)
        widget.setGraphicsEffect(effect)
        fade = QPropertyAnimation(effect, b"opacity", self)
        fade.setDuration(190)
        fade.setStartValue(0.0)
        fade.setEndValue(1.0)
        fade.setEasingCurve(QEasingCurve.OutCubic)
        fade.finished.connect(lambda: widget.setGraphicsEffect(None))
        fade.start()
        self._page_fade = fade

        title, subtitle = headers[page_name]
        self.set_header(title, subtitle)
        log_user_action("navigate", f"{page_name} page", page=page_name)
        if hasattr(self, "header_status"):
            self.header_status.setText(title)

        if page_name == "home" and hasattr(self, "home_cards"):
            QTimer.singleShot(220, lambda: self.animate_card_entrance(self.home_cards))
        elif page_name == "input" and hasattr(self, "input_cards"):
            QTimer.singleShot(220, lambda: self.animate_card_entrance(self.input_cards))

    # ── source helpers ────────────────────────────────────────────────────────
    def get_source_display_name(self, path: str) -> str:
        normalized = normalize_path(path)
        parts = [p for p in normalized.replace("\\", "/").split("/") if p]
        if not parts:
            return normalized
        if parts[-1].lower() == "src" and len(parts) >= 2:
            return parts[-2]
        return parts[-1]

    def build_source_entry(self, source_type: str, path: str) -> dict:
        base_name = self.get_source_display_name(path)
        prefix = {"target": "Target Base", "reference": "Reference Base",
                  "consolidated": "Consolidated DB"}.get(source_type, "Source")
        return {"type": source_type, "path": normalize_path(path), "label": f"{prefix} - {base_name}"}

    def register_sources(self, entries: list, function_list_paths: list, reference_folders: list = None):
        self.available_sources           = entries
        self.current_function_list_paths = function_list_paths or []
        self.current_reference_folders   = reference_folders or []
        self.source_combo.blockSignals(True)
        self.source_combo.clear()
        for entry in self.available_sources:
            self.source_combo.addItem(entry["label"], entry["path"])
        self.source_combo.blockSignals(False)
        if self.available_sources:
            self.source_combo.setCurrentIndex(0)
            self.load_selected_source()
            self.show_page("view")

    def load_selected_source(self):
        if not self.available_sources:
            return
        idx = self.source_combo.currentIndex()
        if idx < 0 or idx >= len(self.available_sources):
            return
        source = self.available_sources[idx]
        self.active_source_root = source["path"]

        # Points 1,2,3: filtering only applies to TARGET base; refs always show all.
        is_target = (source.get("type") == "target")

        try:
            # ── Try disk cache first (populated at Submit time) ───────────────
            role = "target" if is_target else "reference"
            cached_meta = FUNCTION_CACHE.get_meta(source["path"], role)
            if cached_meta is not None:
                all_records = OrderedDict(
                    (fp, info) for fp, info in cached_meta.items()
                )
            else:
                # Fallback: live scan (should rarely happen after Submit)
                all_records = scan_source_for_all_functions(self.active_source_root)

            matched_by_list = OrderedDict()
            function_filter = getattr(self, "_ref_function_filter", [])
            has_xlsx        = bool(self.current_function_list_paths)  # point 4

            if is_target:
                if has_xlsx:
                    # Point 2: xlsx uploaded -> only show functions that match the list
                    if cached_meta is not None:
                        # Filter from cache without re-scanning
                        filter_set = {normalize_name(fn) for fn in function_filter if fn}
                        if filter_set:
                            matched_by_list = OrderedDict()
                            for fp, info in all_records.items():
                                matched_fns = [fn for fn in info.get("functions", [])
                                               if normalize_name(fn) in filter_set]
                                if matched_fns:
                                    matched_by_list[fp] = dict(info, functions=matched_fns)
                            self.function_records = matched_by_list if matched_by_list else all_records
                        else:
                            self.function_records = all_records
                    else:
                        matched_by_list = match_target_with_function_list(
                            self.active_source_root, self.current_function_list_paths)
                        self.function_records = matched_by_list if matched_by_list else all_records
                else:
                    # Point 1: no xlsx -> show ALL target functions
                    self.function_records = all_records
            else:
                # Point 3: reference bases -> always all functions, no filtering
                self.function_records = all_records

            # Point 4: mode chip visible ONLY when xlsx function list is loaded for target
            if hasattr(self, "view_mode_chip"):
                if has_xlsx and is_target:
                    fn_count = sum(len(v["functions"]) for v in self.function_records.values())
                    self.view_mode_chip.setText(
                        "  Function List Filter  ·  {} functions  ".format(fn_count)
                    )
                    self.view_mode_chip.setVisible(True)
                else:
                    self.view_mode_chip.setVisible(False)

            self.populate_view_tree()
            self.populate_diff_tree()
            self.view_title.setText(source["label"])
            self.view_meta.setText(self.active_source_root)
            self.view_text.setText(
                "Selected source:\\n{}\\n{}\\n\\nLoaded Files: {}\\n\\n"
                "Drag the center divider left or right to resize the panes.\\n"
                "Click a function on the left to preview the real body.".format(
                    source["label"], self.active_source_root, len(self.function_records))
            )
            self.current_function_name = ""
            self.current_function_file = ""
            self.current_function_body = ""
        except Exception as e:
            QMessageBox.critical(self, "Load Error", str(e))

    def on_source_changed(self, _index):
        self.load_selected_source()

    # ── tree population ───────────────────────────────────────────────────────
    def populate_view_tree(self):
        self.tree.clear()
        if not self.function_records:
            self.view_text.setText("No matching files/functions found for the selected source.")
            return
        for file_path, info in self.function_records.items():
            parent_text = f"{info['display_name']} ({len(info['functions'])})"
            parent = QTreeWidgetItem([parent_text])
            parent.setData(0, Qt.UserRole, {"type": "file", "file_path": file_path})
            for func_name in info["functions"]:
                child = QTreeWidgetItem([func_name])
                child.setData(0, Qt.UserRole, {"type": "function", "file_path": file_path, "function_name": func_name})
                parent.addChild(child)
            self.tree.addTopLevelItem(parent)
        self.tree.expandAll()

    def populate_diff_tree(self):
        """Build diff tree:
        - Fix 1: always use the TARGET source on the left; ignore which source
          is currently selected in the View dropdown.
        - Fix 2: support multiple reference folders (one tab per ref)
        - Tag each function added/deleted/modified for filter buttons
        """
        import difflib as _difflib
        self.diff_tree.clear()
        self._diff_raw_data = []   # reset before repopulating
        ref_folders = getattr(self, "current_reference_folders", [])

        # --- Fix 1: resolve target records independently of the current view selection ---
        target_entry = next(
            (s for s in getattr(self, "available_sources", []) if s.get("type") == "target"),
            None,
        )
        if target_entry:
            try:
                # Use cached meta when available
                cached = FUNCTION_CACHE.get_meta(target_entry["path"], "target")
                target_records = OrderedDict(cached) if cached else scan_source_for_all_functions(target_entry["path"])
            except Exception:
                target_records = {}
        else:
            # Fallback: if no explicit target, use whatever is in function_records
            target_records = self.function_records

        # Build reference index per FOLDER:
        # {display_name -> {func_name -> [(folder_path, file_path), ...]}}
        ref_index: dict = {}
        for folder in ref_folders:
            # Use cached meta when available
            cached_ref = FUNCTION_CACHE.get_meta(folder, "reference")
            ref_records = OrderedDict(cached_ref) if cached_ref else scan_source_for_all_functions(folder)
            for fp, rinfo in ref_records.items():
                dn = rinfo["display_name"]
                ref_index.setdefault(dn, {})
                for fn in rinfo["functions"]:
                    ref_index[dn].setdefault(fn, []).append((folder, fp))

        for file_path, info in target_records.items():
            dn = info["display_name"]
            ref_fns = ref_index.get(dn, {})  # {func_name -> [(folder, fp)]}

            # If this file does not exist in ANY reference folder,
            # skip it entirely — it should not appear in the diff tree.
            if not ref_fns and ref_folders:
                continue

            tagged_children = []
            for func_name in info["functions"]:
                if func_name in ref_fns:
                    # Check across ALL reference copies — use worst-case tag
                    # (if any ref differs → modified; if all identical → equal)
                    overall_tag = "equal"
                    best_tgt_body = ""
                    best_ref_body = ""
                    for folder, ref_fp in ref_fns[func_name]:
                        tgt_body = (FUNCTION_CACHE.get_body(target_entry["path"] if target_entry else file_path,
                                                             "target", file_path, func_name)
                                    or extract_function_body(file_path, func_name))
                        ref_body = (FUNCTION_CACHE.get_body(folder, "reference", ref_fp, func_name)
                                    or extract_function_body(ref_fp, func_name))
                        t_lines  = [l for l in tgt_body.splitlines() if l.strip()]
                        r_lines  = [l for l in ref_body.splitlines() if l.strip()]
                        ratio    = _difflib.SequenceMatcher(None, t_lines, r_lines).ratio()
                        if ratio < 1.0:
                            overall_tag = "modified"
                            best_tgt_body = tgt_body
                            best_ref_body = ref_body
                            break
                    tag = overall_tag

                    # Compute line-level diff flags for filter button logic.
                    has_added    = False
                    has_deleted  = False
                    has_modified = False
                    if tag == "modified":
                        tgt_lines = best_tgt_body.splitlines()
                        ref_lines = best_ref_body.splitlines()
                        matcher   = _difflib.SequenceMatcher(None, tgt_lines, ref_lines)
                        for op, i1, i2, j1, j2 in matcher.get_opcodes():
                            if op == "equal":
                                continue
                            nc = tgt_lines[i1:i2]
                            rc = ref_lines[j1:j2]
                            if op == "replace":
                                if len(nc) == 1 and len(rc) == 1:
                                    sim = _difflib.SequenceMatcher(
                                        None, nc[0].strip(), rc[0].strip()
                                    ).ratio()
                                    if sim >= 0.5:
                                        has_modified = True
                                    else:
                                        if any(l.strip() for l in nc): has_added   = True
                                        if any(l.strip() for l in rc): has_deleted = True
                                else:
                                    if any(l.strip() for l in nc): has_added   = True
                                    if any(l.strip() for l in rc): has_deleted = True
                            elif op == "delete":
                                if any(l.strip() for l in nc): has_added = True
                            elif op == "insert":
                                if any(l.strip() for l in rc): has_deleted = True

                    tagged_children.append((
                        func_name, tag, file_path,
                        ref_fns.get(func_name, []),  # list of (folder, fp)
                        has_added, has_deleted, has_modified,
                    ))
                else:
                    # Function only in target and NOT present in any reference
                    # file — skip it; there is nothing to diff against.
                    continue

            # Deleted: in reference but NOT in target
            target_fn_set = set(info["functions"])
            for ref_fn, ref_copies in ref_fns.items():
                if ref_fn not in target_fn_set:
                    tagged_children.append((
                        ref_fn, "deleted", file_path, ref_copies,
                        False, True, False,  # has_added=F, has_deleted=T, has_modified=F
                    ))

            if not tagged_children:
                continue

            # Store raw entry so _apply_diff_filter can rebuild the tree cleanly
            self._diff_raw_data.append({
                "display_name": dn,
                "file_path":    file_path,
                "functions":    tagged_children,
            })

        # Delegate all tree-building to _apply_diff_filter so the active
        # filter/search is always respected from a single code path.
        self._apply_diff_filter()

    def filter_tree_items(self, text: str, tree_widget=None):
        query = text.strip().lower()
        tree  = tree_widget if tree_widget is not None else self.tree
        # If this is the diff tree, delegate to _apply_diff_filter so tag
        # filter and search filter are always combined correctly (Bug 3 fix).
        if tree is getattr(self, "diff_tree", None):
            self._apply_diff_filter()
            return
        for i in range(tree.topLevelItemCount()):
            parent = tree.topLevelItem(i)
            parent_visible = False
            for j in range(parent.childCount()):
                child = parent.child(j)
                child_match = (query in child.text(0).lower()) if query else True
                child.setHidden(not child_match)
                if child_match:
                    parent_visible = True
            parent_match = (query in parent.text(0).lower()) if query else True
            parent.setHidden(not (parent_match or parent_visible))

    # ── tree click handlers ───────────────────────────────────────────────────
    def on_tree_item_clicked(self, item, column):
        payload = item.data(0, Qt.UserRole)
        if not payload:
            return
        if payload["type"] == "file":
            fp = payload["file_path"]
            self.current_function_name = ""; self.current_function_file = fp; self.current_function_body = ""
            self.view_title.setText(os.path.basename(fp))
            self.view_meta.setText(fp)
            self.view_text.setText(f"File selected:\n{fp}\n\nFunctions are listed under this file.")
        elif payload["type"] == "function":
            fp   = payload["file_path"]
            name = payload["function_name"]
            # Try cached .txt first
            source = next(
                (s for s in getattr(self, "available_sources", [])
                 if s["path"] == self.active_source_root),
                None,
            )
            role = "target" if (source and source.get("type") == "target") else "reference"
            body = FUNCTION_CACHE.get_body(self.active_source_root, role, fp, name)
            if body is None:
                body = extract_function_body(fp, name)
            self.current_function_name = name
            self.current_function_file = fp
            self.current_function_body = body
            self.view_title.setText(name)
            self.view_meta.setText(fp)
            self.view_text.setText(body)
            log_user_action("click", f"Function: {name}", page="view",
                            extra=f"file={os.path.basename(fp)}")

    def on_diff_item_clicked(self, item, column):
        """Show diff for clicked function.
        Issue 2: one tab per reference folder.
        Issue 3: colour on BOTH left and right columns correctly.
        """
        import difflib
        payload = item.data(0, Qt.UserRole)
        if not payload or payload["type"] != "function":
            return

        function_name = payload["function_name"]
        target_file   = payload["file_path"]
        ref_copies    = payload.get("ref_copies", [])   # [(folder_path, file_path), ...]
        diff_tag      = payload.get("diff_tag", "")

        # Update fullscreen info bar if active
        if getattr(self, "_diff_fullscreen", False):
            self._diff_update_fs_info()
            self._diff_update_nav_state()

        # Always clear old diff first
        self.diff_tabs.clear()

        # Fix 1: always derive target_label from the target source entry, not the
        # currently-selected source in the view dropdown
        target_entry_for_label = next(
            (s for s in getattr(self, "available_sources", []) if s.get("type") == "target"),
            None,
        )
        if target_entry_for_label:
            base_path    = os.path.normpath(target_entry_for_label["path"])
        else:
            base_path    = os.path.normpath(self.active_source_root)
        parts        = base_path.split(os.sep)
        target_label = (parts[parts.index("src") - 1]
                        if "src" in parts and parts.index("src") > 0
                        else os.path.basename(base_path))

        target_entry_obj = next(
            (s for s in getattr(self, "available_sources", []) if s.get("type") == "target"),
            None,
        )
        target_code = ""
        if target_file and os.path.isfile(target_file):
            tgt_folder = target_entry_obj["path"] if target_entry_obj else self.active_source_root
            target_code = (FUNCTION_CACHE.get_body(tgt_folder, "target", target_file, function_name)
                           or extract_function_body(target_file, function_name))

        def _make_ref_label(folder_path):
            norm = os.path.normpath(folder_path).split(os.sep)
            return (norm[norm.index("src") - 1]
                    if "src" in norm and norm.index("src") > 0
                    else os.path.basename(folder_path))

        def _build_table(tgt_code, ref_code, tgt_lbl, ref_lbl, fn_tag):
            table = QTableWidget()
            table.setColumnCount(2)
            table.setHorizontalHeaderLabels([tgt_lbl, ref_lbl])
            table.setEditTriggers(QTableWidget.NoEditTriggers)
            table.setSelectionMode(QTableWidget.NoSelection)
            table.verticalHeader().setVisible(False)
            table.setWordWrap(False)
            table.setShowGrid(True)
            fnt = QFont("Consolas")
            fnt.setStyleHint(QFont.Monospace)
            fnt.setFixedPitch(True)
            table.setFont(fnt)
            table.setHorizontalScrollMode(QTableWidget.ScrollPerPixel)
            table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
            hdr = table.horizontalHeader()
            hdr.setSectionResizeMode(0, QHeaderView.Stretch)
            hdr.setSectionResizeMode(1, QHeaderView.Stretch)

            new_lines = tgt_code.splitlines() if tgt_code else []
            ref_lines = ref_code.splitlines() if ref_code else []

            rows_diff = []
            if fn_tag == "added":
                for ln in new_lines:
                    rows_diff.append(("only_target", ln, ""))
            elif fn_tag == "deleted":
                for ln in ref_lines:
                    rows_diff.append(("only_ref", "", ln))
            else:
                matcher = difflib.SequenceMatcher(None, new_lines, ref_lines)
                for op, i1, i2, j1, j2 in matcher.get_opcodes():
                    nc = new_lines[i1:i2]
                    rc = ref_lines[j1:j2]
                    if op == "equal":
                        for n, r in zip(nc, rc):
                            rows_diff.append(("equal", n, r))
                    elif op == "replace":
                        if len(nc) == 1 and len(rc) == 1:
                            similarity = difflib.SequenceMatcher(
                                None, nc[0].strip(), rc[0].strip()
                            ).ratio()
                            if similarity >= 0.5:
                                rows_diff.append(("modified", nc[0], rc[0]))
                            else:
                                rows_diff.append(("only_target", nc[0], ""))
                                rows_diff.append(("only_ref", "", rc[0]))
                        else:
                            for n in nc:
                                rows_diff.append(("only_target", n, ""))
                            for r in rc:
                                rows_diff.append(("only_ref", "", r))
                    elif op == "delete":
                        for n in nc:
                            rows_diff.append(("only_target", n, ""))
                    elif op == "insert":
                        for r in rc:
                            rows_diff.append(("only_ref", "", r))

            table.setRowCount(len(rows_diff))

            # Resolve live theme ref
            t = self.theme

            # Mark equal rows with a sentinel so the delegate knows to use theme colors
            DIFF_BG  = Qt.BackgroundRole
            DIFF_FG  = Qt.ForegroundRole
            IS_EQUAL = Qt.UserRole + 10   # sentinel: True = use live theme colors

            # Delegate reads self.theme live on every paint — works after theme switch
            win_ref = self
            class _DiffDelegate(QStyledItemDelegate):
                def paint(self, painter, option, index):
                    t_live   = win_ref.theme
                    is_equal = index.data(IS_EQUAL)
                    painter.save()
                    if is_equal:
                        # Use live theme colors for equal rows
                        bg = QColor(t_live["bg_soft"] if index.row() % 2 == 1 else t_live["bg_card"])
                        fg = QColor(t_live["text_primary"])
                    else:
                        bg = index.data(DIFF_BG)
                        fg = index.data(DIFF_FG)
                        if not bg or not bg.isValid():
                            bg = QColor(t_live["bg_card"])
                        if not fg or not fg.isValid():
                            fg = QColor(t_live["text_primary"])
                    painter.fillRect(option.rect, QBrush(bg))
                    painter.setPen(fg)
                    text = index.data(Qt.DisplayRole) or ""
                    painter.drawText(option.rect.adjusted(6, 0, -4, 0),
                                     Qt.AlignLeft | Qt.AlignVCenter, text)
                    painter.restore()

            delegate = _DiffDelegate(table)
            table.setItemDelegate(delegate)

            # Apply palette so viewport bg matches theme (no stylesheet on table)
            def _apply_table_palette(tbl, theme):
                pal = tbl.palette()
                pal.setColor(QPalette.Base,          QColor(theme["bg_card"]))
                pal.setColor(QPalette.AlternateBase, QColor(theme["bg_soft"]))
                pal.setColor(QPalette.Text,          QColor(theme["text_primary"]))
                pal.setColor(QPalette.Window,        QColor(theme["bg_card"]))
                tbl.setPalette(pal)
                tbl.setAutoFillBackground(True)
                vpal = tbl.viewport().palette()
                vpal.setColor(QPalette.Base,          QColor(theme["bg_card"]))
                vpal.setColor(QPalette.AlternateBase, QColor(theme["bg_soft"]))
                vpal.setColor(QPalette.Text,          QColor(theme["text_primary"]))
                tbl.viewport().setPalette(vpal)
                tbl.viewport().setAutoFillBackground(True)
                tbl.horizontalHeader().setStyleSheet(
                    f"QHeaderView::section {{ background: {theme['bg_soft']}; color: {theme['text_primary']};"
                    f"  border: 1px solid {theme['border']}; font-weight: 700; padding: 4px 8px; }}"
                )
                tbl.viewport().update()

            _apply_table_palette(table, t)

            # Register table so set_theme_mode can refresh it on theme switch
            if not hasattr(self, "_diff_tables"):
                self._diff_tables = []
            self._diff_tables = [tbl for tbl in self._diff_tables if tbl.parent() is not None]
            self._diff_tables.append(table)
            table._apply_palette = _apply_table_palette

            for row_idx, (rtag, nl, rl) in enumerate(rows_diff):
                ni = QTableWidgetItem(nl)
                ri = QTableWidgetItem(rl)
                ni.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                ri.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)

                if rtag == "only_target" and nl.strip():
                    for item in (ni, ri):
                        item.setData(DIFF_BG, QColor("#2E7D32"))
                        item.setData(DIFF_FG, QColor("#FFFFFF"))
                elif rtag == "only_ref" and rl.strip():
                    for item in (ni, ri):
                        item.setData(DIFF_BG, QColor("#F9A825"))
                        item.setData(DIFF_FG, QColor("#000000"))
                elif rtag == "modified":
                    for item in (ni, ri):
                        item.setData(DIFF_BG, QColor("#1565C0"))
                        item.setData(DIFF_FG, QColor("#FFFFFF"))
                else:
                    # equal rows — sentinel so delegate uses live theme colors
                    for item in (ni, ri):
                        item.setData(IS_EQUAL, True)

                table.setItem(row_idx, 0, ni)
                table.setItem(row_idx, 1, ri)

            table.resizeRowsToContents()
            return table

        if not ref_copies:
            table = _build_table(target_code, "", target_label, "Reference", diff_tag)
            self.diff_tabs.addTab(table, "{} vs Reference".format(target_label))
        else:
            for folder_path, ref_fp in ref_copies:
                ref_label = _make_ref_label(folder_path)
                ref_code  = ""
                if ref_fp and os.path.isfile(ref_fp):
                    ref_code = (FUNCTION_CACHE.get_body(folder_path, "reference", ref_fp, function_name)
                                or extract_function_body(ref_fp, function_name))
                table = _build_table(target_code, ref_code, target_label, ref_label, diff_tag)
                tab_title = "{} vs {}".format(target_label, ref_label)
                self.diff_tabs.addTab(table, tab_title)

    # ── diff controls ─────────────────────────────────────────────────────────
    def _toggle_diff_fullscreen(self):
        self._diff_fullscreen = not self._diff_fullscreen
        if self._diff_fullscreen:
            self.sidebar.hide()
            self.header.hide()
            self._diff_left_panel.hide()
            self._diff_fs_btn.setText("Exit Fullscreen")
            if hasattr(self, "_diff_filter_icon_lbl"):
                self._diff_filter_icon_lbl.setVisible(False)
            for btn in getattr(self, "_diff_filter_btns", {}).values():
                btn.setVisible(False)
            if hasattr(self, "_diff_fs_info_bar"):
                self._diff_fs_info_bar.setVisible(True)
            self._diff_update_fs_info()
            self._diff_connect_nav()
        else:
            self.sidebar.show()
            self.header.show()
            self._diff_left_panel.show()
            self._diff_fs_btn.setText("⛶  Fullscreen")
            if hasattr(self, "_diff_filter_icon_lbl"):
                self._diff_filter_icon_lbl.setVisible(True)
            for btn in getattr(self, "_diff_filter_btns", {}).values():
                btn.setVisible(True)
            if hasattr(self, "_diff_fs_info_bar"):
                self._diff_fs_info_bar.setVisible(False)

    def _diff_update_fs_info(self):
        """Update the fullscreen info bar labels with current file/function."""
        item = self.diff_tree.currentItem()
        if item and item.data(0, Qt.UserRole):
            payload = item.data(0, Qt.UserRole)
            if payload.get("type") == "function":
                fn  = payload.get("function_name", "")
                fp  = payload.get("file_path", "")
                fname = os.path.basename(fp) if fp else ""
                if hasattr(self, "_diff_fs_file_label"):
                    self._diff_fs_file_label.setText(f"📄 {fname}")
                if hasattr(self, "_diff_fs_func_label"):
                    self._diff_fs_func_label.setText(f"ƒ  {fn}")
                return
        if hasattr(self, "_diff_fs_file_label"):
            self._diff_fs_file_label.setText("")
        if hasattr(self, "_diff_fs_func_label"):
            self._diff_fs_func_label.setText("")

    def _diff_connect_nav(self):
        """Wire Prev/Next buttons for keyboard-free navigation in fullscreen."""
        try:
            self._diff_prev_btn.clicked.disconnect()
            self._diff_next_btn.clicked.disconnect()
        except Exception:
            pass
        self._diff_prev_btn.clicked.connect(lambda: self._diff_navigate(-1))
        self._diff_next_btn.clicked.connect(lambda: self._diff_navigate(1))
        self._diff_update_nav_state()

    def _diff_get_all_function_items(self):
        """Return flat list of all visible function-level QTreeWidgetItems."""
        items = []
        root = self.diff_tree.invisibleRootItem()
        for i in range(root.childCount()):
            file_item = root.child(i)
            for j in range(file_item.childCount()):
                fn_item = file_item.child(j)
                payload = fn_item.data(0, Qt.UserRole)
                if payload and payload.get("type") == "function":
                    items.append(fn_item)
        return items

    def _diff_navigate(self, direction: int):
        """Navigate to the next (+1) or previous (-1) function item."""
        items = self._diff_get_all_function_items()
        if not items:
            return
        current = self.diff_tree.currentItem()
        try:
            idx = items.index(current)
        except ValueError:
            idx = -1
        new_idx = idx + direction
        if new_idx < 0 or new_idx >= len(items):
            return
        new_item = items[new_idx]
        self.diff_tree.setCurrentItem(new_item)
        self.diff_tree.scrollToItem(new_item)
        self.on_diff_item_clicked(new_item, 0)
        self._diff_update_fs_info()
        self._diff_update_nav_state()

    def _diff_update_nav_state(self):
        """Enable/disable Prev/Next based on current position."""
        if not hasattr(self, "_diff_prev_btn"):
            return
        items = self._diff_get_all_function_items()
        current = self.diff_tree.currentItem()
        try:
            idx = items.index(current)
        except ValueError:
            idx = -1
        self._diff_prev_btn.setEnabled(idx > 0)
        self._diff_next_btn.setEnabled(0 <= idx < len(items) - 1)

    def _clear_diff(self):
        self.diff_tabs.clear()
        self.diff_tree.clear()
        self.diff_search_box.clear()
        # Point 6: reset filter buttons when diff is cleared
        self._diff_active_filter = None
        if hasattr(self, "_diff_filter_btns") and hasattr(self, "_diff_style_btn"):
            for k, btn in self._diff_filter_btns.items():
                color = btn.property("activeColor")
                fg    = btn.property("activeFg")
                btn.setChecked(False)
                self._diff_style_btn(btn, active=False, greyed=False, color=color, fg=fg)
        with SCAN_CACHE_LOCK:
            SCAN_CACHE.clear()

    def _apply_diff_filter(self):
        """Rebuild the diff tree from raw data, applying the active tag filter
        and search text.
        """
        active      = getattr(self, "_diff_active_filter", None)
        search_text = self.diff_search_box.text().strip().lower()

        self.diff_tree.clear()

        for entry in getattr(self, "_diff_raw_data", []):
            dn        = entry["display_name"]
            file_path = entry["file_path"]
            all_fns   = entry["functions"]

            visible_fns = []
            for fn_tuple in all_fns:
                func_name, tag, tgt_fp, ref_copies = fn_tuple[0], fn_tuple[1], fn_tuple[2], fn_tuple[3]
                has_added    = fn_tuple[4] if len(fn_tuple) > 4 else (tag == "added")
                has_deleted  = fn_tuple[5] if len(fn_tuple) > 5 else (tag == "deleted")
                has_modified = fn_tuple[6] if len(fn_tuple) > 6 else (tag == "modified")

                if active is None:
                    tag_ok = True
                elif tag == "equal":
                    tag_ok = False
                elif active == "added":
                    tag_ok = has_added
                elif active == "deleted":
                    tag_ok = has_deleted
                elif active == "modified":
                    tag_ok = has_modified
                else:
                    tag_ok = (tag == active)

                search_ok = (not search_text) or (
                    search_text in func_name.lower() or
                    search_text in dn.lower()
                )
                if tag_ok and search_ok:
                    visible_fns.append((func_name, tag, tgt_fp, ref_copies))

            if not visible_fns:
                continue

            parent = QTreeWidgetItem([dn])
            parent.setData(0, Qt.UserRole, {
                "type":      "file",
                "file_path": file_path,
            })

            for func_name, tag, tgt_fp, ref_copies in visible_fns:
                child = QTreeWidgetItem([func_name])
                child.setData(0, Qt.UserRole, {
                    "type":          "function",
                    "file_path":     tgt_fp,
                    "ref_copies":    ref_copies,
                    "function_name": func_name,
                    "diff_tag":      tag,
                })
                parent.addChild(child)

            self.diff_tree.addTopLevelItem(parent)

        self.diff_tree.expandAll()

    # ── submit reset helper ───────────────────────────────────────────────────
    def _reset_view_and_diff(self):
        """Clear view page and diff page state fully before a new submit."""
        with SCAN_CACHE_LOCK:
            SCAN_CACHE.clear()

        self.function_records = OrderedDict()
        self.available_sources = []
        self.current_function_list_paths = []
        self.current_reference_folders = []
        self._ref_function_filter = []
        self.current_function_name = ""
        self.current_function_file = ""
        self.current_function_body = ""

        if hasattr(self, "source_combo"):
            self.source_combo.blockSignals(True)
            self.source_combo.clear()
            self.source_combo.blockSignals(False)
        if hasattr(self, "tree"):
            self.tree.clear()
        if hasattr(self, "view_text"):
            self.view_text.setText("")
        if hasattr(self, "view_title"):
            self.view_title.setText("")
        if hasattr(self, "view_meta"):
            self.view_meta.setText("")
        if hasattr(self, "view_mode_chip"):
            self.view_mode_chip.setVisible(False)

        self._clear_diff()

    # ── submit reference ──────────────────────────────────────────────────────
    def submit_reference(self):
        target_folders = self.ref_target_field.value()
        ref_folders    = self.ref_bases_field.value()
        function_list  = self.ref_function_field.value()
        log_user_action("click", "Submit Reference button", page="reference",
                        extra=f"target_count={len(target_folders)}, ref_count={len(ref_folders)}, fn_files={len(function_list)}")

        if not target_folders:
            QMessageBox.warning(self, "Missing Input", "Please select Target Base Folder.")
            return

        target_root = target_folders[0]
        norm_target = os.path.normpath(target_root).lower()

        # ── Duplicate: same folder added twice in Reference Bases ────────────
        seen_refs = set()
        dup_refs  = []
        for f in ref_folders:
            key = os.path.normpath(f).lower()
            if key in seen_refs:
                dup_refs.append(f)
            else:
                seen_refs.add(key)
        if dup_refs:
            QMessageBox.warning(
                self, "Duplicate Reference Folder",
                "The following folder(s) appear more than once in Reference Bases — "
                "please remove duplicates before submitting:\n\n" +
                "\n".join(dup_refs)
            )
            return

        # ── Target folder also appears in Reference Bases ────────────────────
        clashing_refs = [f for f in ref_folders
                         if os.path.normpath(f).lower() == norm_target]
        if clashing_refs:
            QMessageBox.warning(
                self, "Conflicting Folders",
                f"The Target Base Folder is also listed as a Reference Base Folder:\n\n"
                f"{target_root}\n\n"
                "A folder cannot be both the target and a reference. "
                "Please choose a different folder for one of them."
            )
            return

        self.ref_submit_btn.start_progress("Extracting …")
        from PySide6.QtWidgets import QApplication
        QApplication.processEvents()

        try:
            self._reset_view_and_diff()

            # ── Clear old cache before a fresh extraction ─────────────────
            FUNCTION_CACHE.clear()
            self._cached_folders = {}

            self.current_target_folders = ref_folders
            self._ref_function_filter = parse_function_list_files(function_list) if function_list else []

            entries = [self.build_source_entry("target", target_root)]
            for folder in ref_folders:
                entries.append(self.build_source_entry("reference", folder))

            for f in function_list:
                log_file_upload("file", f, field="Function List")

            # ── Build bases list for upfront extraction worker ────────────
            bases = [{"folder_path": target_root, "role": "target",
                      "label": entries[0]["label"]}]
            for i, folder in enumerate(ref_folders, 1):
                bases.append({"folder_path": folder, "role": "reference",
                               "label": entries[i]["label"]})

            fn_filter_set = {normalize_name(fn) for fn in self._ref_function_filter if fn}

            # Store these for use after extraction completes
            self._pending_entries        = entries
            self._pending_function_list  = function_list
            self._pending_ref_folders    = ref_folders

            # ── Launch upfront extraction worker ──────────────────────────
            self._upfront_worker = UpfrontExtractionWorker(bases, fn_filter_set or None)
            self._upfront_thread = QThread()
            self._upfront_worker.moveToThread(self._upfront_thread)
            self._upfront_thread.started.connect(self._upfront_worker.run)
            self._upfront_worker.started.connect(
                lambda lbl: self.ref_submit_btn.start_progress("Running …")
            )
            self._upfront_worker.finished.connect(self._on_upfront_done)
            self._upfront_worker.error.connect(self._on_upfront_error)
            self._upfront_worker.finished.connect(self._upfront_thread.quit)
            self._upfront_worker.error.connect(self._upfront_thread.quit)
            self._upfront_worker.finished.connect(self._upfront_worker.deleteLater)
            self._upfront_thread.finished.connect(self._upfront_thread.deleteLater)
            self._upfront_thread.start()

        except Exception as e:
            self.ref_submit_btn.stop_progress("Submit")
            QMessageBox.critical(self, "Submit Error", str(e))

    def _on_upfront_done(self, results: dict):
        """Called when the upfront extraction worker finishes successfully."""
        try:
            self.ref_submit_btn.stop_progress("Submit")

            entries       = self._pending_entries
            function_list = self._pending_function_list
            ref_folders   = self._pending_ref_folders

            # Record which folders are now cached (use entries list for correct role)
            for entry in entries:
                fp   = normalize_path(entry["path"])
                role = "target" if entry.get("type") == "target" else "reference"
                self._cached_folders[fp] = role

            target_root = entries[0]["path"] if entries else ""
            self.register_sources(entries, function_list, ref_folders)

            _log.info("submit_reference success: target=%s, refs=%d, fn_files=%d, sources=%d",
                      target_root, len(ref_folders), len(function_list), len(entries))

            ref_text = "\n".join(ref_folders) if ref_folders else "No reference folders selected"
            fn_text  = "\n".join(function_list) if function_list else "No function list selected"
            total_fns = sum(
                sum(len(v.get("functions", [])) for v in data.get("meta", {}).values())
                for data in results.values()
            )
            QMessageBox.information(self, "Success",
                f"Reference data loaded successfully.\n\nTarget Base Folder:\n{target_root}\n\n"
                f"Reference Base Folders:\n{ref_text}\n\nFunction List:\n{fn_text}\n\n"
                f"Dropdown Sources: {len(entries)}\n"
                f"Functions cached to disk: {total_fns}")

        except Exception as e:
            QMessageBox.critical(self, "Submit Error", str(e))
        finally:
            self._upfront_thread = None
            self._upfront_worker = None

    def _on_upfront_error(self, message: str):
        """Called when the upfront extraction worker fails."""
        self.ref_submit_btn.stop_progress("Submit")
        self._upfront_thread = None
        self._upfront_worker = None
        if message != "__CANCELLED__":
            QMessageBox.critical(self, "Extraction Error", message)

    # ── submit consolidated ───────────────────────────────────────────────────
    def submit_consolidated(self):
        use_folder_mode = hasattr(self, "con_toggle_folder_btn") and \
                          self.con_toggle_folder_btn.isChecked()
        log_user_action("click", "Submit Consolidated button", page="consolidated",
                        extra=f"folder_mode={use_folder_mode}")

        self._reset_view_and_diff()

        if use_folder_mode:
            target_folder = self.con_folder_field.value()
            if not target_folder or not os.path.isdir(target_folder):
                QMessageBox.warning(self, "Missing Input",
                    "Please select a Target Folder to extract functions from."); return

            consolidated_excel = self.con_db_excel_field.value()
            if not consolidated_excel:
                QMessageBox.warning(self, "Missing Input",
                    "Please select Consolidated DB Excel file."); return

            base_col_ref = self.con_base_col_field.value()
            if not base_col_ref or not is_valid_excel_reference(base_col_ref):
                QMessageBox.warning(self, "Invalid Input",
                    "Base Col must be detected or entered like B2, C1, G3, etc."); return

            try:
                temp_excel_path, fn_count = extract_functions_from_folder_to_excel(target_folder)
            except Exception as exc:
                QMessageBox.critical(self, "Extraction Error",
                    f"Failed to extract functions from folder:\n{exc}"); return

            if fn_count == 0:
                QMessageBox.warning(self, "No Functions Found",
                    "No functions were detected in the selected folder."); return

            func_col_ref = "C1"
            self.con_func_col_field.input.setText(func_col_ref)
            self.con_func_col_field.preview.setText("Preview: auto-detected (folder mode)")

            function_list_files = [temp_excel_path]

        else:
            function_list_files = self.con_function_field.value()
            consolidated_excel  = self.con_db_excel_field.value()
            func_col_ref        = self.con_func_col_field.value()
            base_col_ref        = self.con_base_col_field.value()

            if not function_list_files:
                QMessageBox.warning(self, "Missing Input",
                    "Please select one or more Function List files."); return
            if not consolidated_excel:
                QMessageBox.warning(self, "Missing Input",
                    "Please select Consolidated DB Excel file."); return
            if not func_col_ref or not is_valid_excel_reference(func_col_ref):
                QMessageBox.warning(self, "Invalid Input",
                    "Func Col must be detected or entered like A2, B1, F3, etc."); return
            if not base_col_ref or not is_valid_excel_reference(base_col_ref):
                QMessageBox.warning(self, "Invalid Input",
                    "Base Col must be detected or entered like B2, C1, G3, etc."); return

        preferred_sheet = None
        if getattr(self.con_func_col_field, "detected_info", None):
            preferred_sheet = self.con_func_col_field.detected_info.get("sheet")
        elif getattr(self.con_base_col_field, "detected_info", None):
            preferred_sheet = self.con_base_col_field.detected_info.get("sheet")

        self.con_submit_btn.setEnabled(False)
        self.con_submit_btn.setText("Submitting...")
        self._lock_nav()
        from PySide6.QtWidgets import QApplication
        QApplication.processEvents()
        self.con_output_link_field.clear_selection()

        self.con_thread = QThread(parent=self)
        self.con_worker = ConsolidatedWorker(
            function_list_files, consolidated_excel, func_col_ref, base_col_ref,
            preferred_sheet=preferred_sheet)
        self.con_worker.moveToThread(self.con_thread)
        self.con_thread.started.connect(self.con_worker.run)
        self.con_worker.finished.connect(self.on_consolidated_finished)
        self.con_worker.error.connect(self.on_consolidated_error)
        self.con_worker.finished.connect(self.con_thread.quit)
        self.con_worker.error.connect(self.con_thread.quit)
        self.con_worker.finished.connect(self.con_worker.deleteLater)
        self.con_worker.error.connect(self.con_worker.deleteLater)
        self.con_thread.finished.connect(self.con_thread.deleteLater)
        self.con_thread.finished.connect(
            lambda t=self.con_thread: self._active_threads.remove(t)
            if t in self._active_threads else None)
        self._active_threads.append(self.con_thread)
        self.con_thread.start()

    def on_consolidated_finished(self, result: dict):
        self.con_submit_btn.setEnabled(True)
        self.con_submit_btn.setText("Submit")
        self._unlock_nav()
        self.con_output_link_field.set_output(result["output_file"])
        log_output_file(result["output_file"], kind="Consolidated Output Excel")
        _log.info("consolidated finished: fn_list_files=%d, functions_read=%d, matched=%d, unmatched=%d",
                  result['function_list_count'], result['functions_read'],
                  result['matched_count'], result.get('unmatched_count', 0))
        QMessageBox.information(self, "Success",
            f"Output Excel generated successfully.\n\n"
            f"Function List Files: {result['function_list_count']}\n"
            f"Functions Read: {result['functions_read']}\n"
            f"Matched Functions In Output: {result['matched_count']}\n"
            f"Unmatched Parsed Functions: {result.get('unmatched_count', 0)}\n"
            f"Output File:\n{result['output_file']}")
        self.con_worker = None; self.con_thread = None

    def on_consolidated_error(self, message: str):
        self.con_submit_btn.setEnabled(True)
        self.con_submit_btn.setText("Submit")
        self._unlock_nav()
        _log.error("consolidated error: %s", message)
        QMessageBox.critical(self, "Processing Error", message)
        self.con_worker = None; self.con_thread = None

    # ── report folder picker ──────────────────────────────────────────────────
    def _pick_report_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Output Folder", "",
                    QFileDialog.ShowDirsOnly | QFileDialog.DontResolveSymlinks)
        if not folder:
            return
        folder = normalize_path(folder)
        self._report_folder_path.clear()
        self._report_folder_path.append(folder)
        self._report_folder_display.setText(folder)
        log_file_upload("folder", folder, field="Report Output Folder")

    # ── report slots ──────────────────────────────────────────────────────────
    def _set_report_progress(self, pct: int, status_text: str = ""):
        """Set the OVERALL cumulative progress bar (0-100) directly."""
        pct = max(0, min(100, int(pct)))
        self.report_progress_bar.setValue(pct)
        self.report_pct_label.setText(f"{pct} %")
        if status_text:
            self.report_status_label.setText(status_text)
            if hasattr(self, "report_summary_chips"):
                self.report_summary_chips["phase"].set_value(status_text[:22])

    def _cumulative_progress(self, current_step_pct: int) -> int:
        total = getattr(self, "_report_total_steps", 1) or 1
        done  = getattr(self, "_report_done_steps",  0)
        overall = (done + current_step_pct / 100.0) / total * 100
        return max(0, min(100, int(overall)))

    def _update_report_step_chip(self):
        if hasattr(self, "report_summary_chips") and hasattr(self, "_report_total_steps"):
            self.report_summary_chips["steps"].set_value(str(self._report_total_steps or 0))

    def _on_base_extraction_started(self, label: str):
        from ui.widgets import StepStatusWidget
        self.report_phase_label.setText(f"Extracting — {label}")
        overall = self._cumulative_progress(0)
        self._set_report_progress(overall, f"{label}: starting …")
        self.report_step_widget.set_state(label, StepStatusWidget.STATE_RUNNING,
                                          "Scanning source files …", pct=0)

    def _on_base_extraction_progress(self, label: str, pct: int, detail: str):
        from ui.widgets import StepStatusWidget
        self.report_phase_label.setText(f"Extracting — {label}")
        overall = self._cumulative_progress(pct)
        self._set_report_progress(overall, f"{label}: {detail}")
        self.report_step_widget.set_state(label, StepStatusWidget.STATE_RUNNING,
                                          detail, pct=pct)

    def _on_compare_progress(self, pct: int, detail: str):
        from ui.widgets import StepStatusWidget
        self.report_phase_label.setText("Comparing & Writing Excel Report")
        overall = self._cumulative_progress(pct)
        self._set_report_progress(overall, detail)
        self.report_step_widget.set_state("📊 Comparing & Writing Excel Report",
                                          StepStatusWidget.STATE_RUNNING, detail, pct=pct)

    def _on_report_generate(self):
        from ui.widgets import StepStatusWidget
        output_root = self.report_output_field.value()
        log_user_action("click", "Generate Report button", page="report",
                        extra=f"output_root={output_root!r}")
        if not output_root:
            QMessageBox.warning(self, "Missing Folder", "Please select an output folder.")
            return
        if not self.available_sources:
            QMessageBox.warning(self, "No Data Loaded",
                "Please submit reference bases first (Input → Reference Bases).")
            return

        target_entry = next((s for s in self.available_sources if s["type"] == "target"), None)
        if not target_entry:
            QMessageBox.warning(self, "No Target", "No target base found in submitted sources.")
            return

        ref_entries = [s for s in self.available_sources if s["type"] == "reference"]
        bases = [{"label": target_entry["label"], "src_path": target_entry["path"], "is_target": True}]
        for ref in ref_entries:
            bases.append({"label": ref["label"], "src_path": ref["path"], "is_target": False})

        self._report_bases_list  = bases
        self._report_total_steps = len(bases) + 1
        self._report_done_steps  = 0
        self._report_total_funcs_extracted = 0
        self._update_report_step_chip()
        self._report_target_label = target_entry["label"]
        self._report_ref_labels   = [r["label"] for r in ref_entries]
        self._report_output_root  = output_root

        self._set_report_progress(0, "Preparing extraction …")
        self.report_log_box.clear()
        self.report_phase_label.setText("Extracting — waiting to start")
        self.report_generate_btn.setEnabled(False)
        self.report_generate_btn.setText("Running …")
        self._lock_nav()
        # ── Start elapsed timer ───────────────────────────────────────────
        self._report_start_time = time.time()
        if not hasattr(self, "_report_timer"):
            self._report_timer = QTimer(self)
            self._report_timer.setInterval(1000)
            self._report_timer.timeout.connect(self._update_report_elapsed)
        self._report_timer.start()
        self.report_cancel_btn.setVisible(True)
        self.report_cancel_btn.setEnabled(True)
        self.report_cancel_btn.setText("Cancel")
        self.report_open_btn.setEnabled(False)
        self._report_output_file = ""
        if hasattr(self, "report_summary_chips"):
            self.report_summary_chips["phase"].set_value("Starting")
            self.report_summary_chips["result"].set_value("Pending")

        import shutil as _shutil
        with SCAN_CACHE_LOCK:
            SCAN_CACHE.clear()
        import tempfile as _tf
        _extracted_root = os.path.join(_tf.gettempdir(), 'FuncAtlas_Extracted')
        if os.path.isdir(_extracted_root):
            try:
                _shutil.rmtree(_extracted_root)
                self.report_log_box.append("🗑 Cleared previous extraction cache.")
            except Exception as _ce:
                self.report_log_box.append(f"Warning: could not clear cache: {_ce}")

        self.report_step_widget.clear_steps()
        for b in bases:
            self.report_step_widget.add_step(b["label"])
        self.report_step_widget.add_step("📊 Comparing & Writing Excel Report")

        self._report_ext_thread = QThread(parent=self)
        self._report_ext_worker = BuiltinExtractionWorker(
            bases, output_root,
            function_filter=self._ref_function_filter if self._ref_function_filter else None
        )
        self._report_ext_worker.moveToThread(self._report_ext_thread)
        self._report_ext_thread.started.connect(self._report_ext_worker.run)
        self._report_ext_worker.base_started.connect(self._on_base_extraction_started)
        self._report_ext_worker.base_progress.connect(self._on_base_extraction_progress)
        self._report_ext_worker.step_done.connect(self._on_step_extracted)
        self._report_ext_worker.log.connect(self._report_append_log)
        self._report_ext_worker.finished.connect(self._on_extraction_done)
        self._report_ext_worker.error.connect(self._on_report_error)
        self._report_ext_worker.finished.connect(self._report_ext_thread.quit)
        self._report_ext_worker.error.connect(self._report_ext_thread.quit)
        self._report_ext_worker.finished.connect(self._report_ext_worker.deleteLater)
        self._report_ext_thread.finished.connect(self._report_ext_thread.deleteLater)
        self._report_ext_thread.finished.connect(
            lambda t=self._report_ext_thread: self._active_threads.remove(t)
            if t in self._active_threads else None)
        self._active_threads.append(self._report_ext_thread)
        self._report_ext_thread.start()

    def _on_step_extracted(self, label: str, func_count: int):
        from ui.widgets import StepStatusWidget
        self._report_done_steps = getattr(self, "_report_done_steps", 0) + 1
        # Accumulate total extracted functions for the success popup
        self._report_total_funcs_extracted = getattr(self, "_report_total_funcs_extracted", 0) + func_count
        self.report_step_widget.set_state(label, StepStatusWidget.STATE_DONE,
                                          f"{func_count} functions extracted")
        overall = self._cumulative_progress(100)
        self._set_report_progress(overall, f"{label}: completed")

    def _on_extraction_done(self, results: dict):
        from ui.widgets import StepStatusWidget
        self._report_append_log("─── All extractions complete. Starting comparison … ───")
        self.report_phase_label.setText("Comparing & Writing Excel Report")
        overall = self._cumulative_progress(0)
        self._set_report_progress(overall, "Preparing comparison …")
        self.report_step_widget.set_state("📊 Comparing & Writing Excel Report",
                                          StepStatusWidget.STATE_RUNNING, "Preparing comparison …")

        target_folder = results.get(self._report_target_label, "")
        ref_folders   = [results[lbl] for lbl in self._report_ref_labels if lbl in results]

        if not target_folder or not os.path.isdir(target_folder):
            self._on_report_error(f"Target extraction folder not found:\n{target_folder}")
            return

        _bases_list   = getattr(self, '_report_bases_list', [])
        _target_src   = next((b['src_path'] for b in _bases_list if b['label'] == self._report_target_label), '')
        _ref_srcs     = [next((b['src_path'] for b in _bases_list if b['label'] == lbl), '')
                         for lbl in self._report_ref_labels]

        self._report_cmp_thread = QThread(parent=self)
        self._report_cmp_worker = ReportCompareWorker(
            target_label    = self._report_target_label,
            target_folder   = target_folder,
            ref_labels      = self._report_ref_labels,
            ref_folders     = ref_folders,
            output_root     = self._report_output_root,
            target_src_path = _target_src,
            ref_src_paths   = _ref_srcs,
            weights         = getattr(self, '_complexity_weights', None),
            bands           = getattr(self, '_complexity_bands',   None),
        )
        self._report_cmp_worker.moveToThread(self._report_cmp_thread)
        self._report_cmp_thread.started.connect(self._report_cmp_worker.run)
        self._report_cmp_worker.progress.connect(self._on_compare_progress)
        self._report_cmp_worker.log.connect(self._report_append_log)
        self._report_cmp_worker.finished.connect(self._on_compare_done)
        self._report_cmp_worker.error.connect(self._on_report_error)
        self._report_cmp_worker.finished.connect(self._report_cmp_thread.quit)
        self._report_cmp_worker.error.connect(self._report_cmp_thread.quit)
        self._report_cmp_worker.finished.connect(self._report_cmp_worker.deleteLater)
        self._report_cmp_thread.finished.connect(self._report_cmp_thread.deleteLater)
        self._report_cmp_thread.finished.connect(
            lambda t=self._report_cmp_thread: self._active_threads.remove(t)
            if t in self._active_threads else None)
        self._active_threads.append(self._report_cmp_thread)
        self._report_cmp_thread.start()

    # ── HTML report generation ────────────────────────────────────────────────
    def _on_report_generate_html(self):
        """Same pipeline as _on_report_generate but produces an HTML report."""
        from ui.widgets import StepStatusWidget
        output_root = self.report_output_field.value()
        log_user_action("click", "Generate HTML Report button", page="report",
                        extra=f"output_root={output_root!r}")
        if not output_root:
            QMessageBox.warning(self, "Missing Folder", "Please select an output folder.")
            return
        if not self.available_sources:
            QMessageBox.warning(self, "No Data Loaded",
                "Please submit reference bases first (Input → Reference Bases).")
            return

        target_entry = next((s for s in self.available_sources if s["type"] == "target"), None)
        if not target_entry:
            QMessageBox.warning(self, "No Target", "No target base found in submitted sources.")
            return

        ref_entries = [s for s in self.available_sources if s["type"] == "reference"]
        bases = [{"label": target_entry["label"], "src_path": target_entry["path"], "is_target": True}]
        for ref in ref_entries:
            bases.append({"label": ref["label"], "src_path": ref["path"], "is_target": False})

        self._report_bases_list   = bases
        self._report_total_steps  = len(bases) + 1
        self._update_report_step_chip()
        self._report_target_label = target_entry["label"]
        self._report_ref_labels   = [r["label"] for r in ref_entries]
        self._report_output_root  = output_root
        self._report_html_mode    = True

        self._set_report_progress(0, "Preparing extraction …")
        self.report_log_box.clear()
        self.report_phase_label.setText("Extracting — waiting to start")
        self.report_generate_html_btn.setEnabled(False)
        self.report_generate_html_btn.setText("Running …")
        self._lock_nav()
        self.report_cancel_btn.setVisible(True)
        self.report_cancel_btn.setEnabled(True)
        self.report_cancel_btn.setText("Cancel")
        self.report_open_btn.setEnabled(False)
        self._report_output_file = ""
        if hasattr(self, "report_summary_chips"):
            self.report_summary_chips["phase"].set_value("Starting")
            self.report_summary_chips["result"].set_value("Pending")

        import shutil as _shutil
        with SCAN_CACHE_LOCK:
            SCAN_CACHE.clear()
        import tempfile as _tf
        _extracted_root = os.path.join(_tf.gettempdir(), 'FuncAtlas_Extracted')
        if os.path.isdir(_extracted_root):
            try:
                _shutil.rmtree(_extracted_root)
                self.report_log_box.append("🗑 Cleared previous extraction cache.")
            except Exception as _ce:
                self.report_log_box.append(f"Warning: could not clear cache: {_ce}")

        self.report_step_widget.clear_steps()
        for b in bases:
            self.report_step_widget.add_step(b["label"])
        self.report_step_widget.add_step("📊 Comparing & Writing HTML Report")

        self._report_ext_thread = QThread(parent=self)
        self._report_ext_worker = BuiltinExtractionWorker(
            bases, output_root,
            function_filter=self._ref_function_filter if self._ref_function_filter else None
        )
        self._report_ext_worker.moveToThread(self._report_ext_thread)
        self._report_ext_thread.started.connect(self._report_ext_worker.run)
        self._report_ext_worker.base_started.connect(self._on_base_extraction_started)
        self._report_ext_worker.base_progress.connect(self._on_base_extraction_progress)
        self._report_ext_worker.step_done.connect(self._on_step_extracted)
        self._report_ext_worker.log.connect(self._report_append_log)
        self._report_ext_worker.finished.connect(self._on_extraction_done_html)
        self._report_ext_worker.error.connect(self._on_report_error_html)
        self._report_ext_worker.finished.connect(self._report_ext_thread.quit)
        self._report_ext_worker.error.connect(self._report_ext_thread.quit)
        self._report_ext_worker.finished.connect(self._report_ext_worker.deleteLater)
        self._report_ext_thread.finished.connect(self._report_ext_thread.deleteLater)
        self._report_ext_thread.finished.connect(
            lambda t=self._report_ext_thread: self._active_threads.remove(t)
            if t in self._active_threads else None)
        self._active_threads.append(self._report_ext_thread)
        self._report_ext_thread.start()

    def _on_extraction_done_html(self, results: dict):
        from ui.widgets import StepStatusWidget
        self._report_append_log("─── All extractions complete. Starting HTML comparison … ───")
        self.report_phase_label.setText("Comparing & Writing HTML Report")
        self._set_report_progress(0, "Preparing comparison …")
        self.report_step_widget.set_state("📊 Comparing & Writing HTML Report",
                                          StepStatusWidget.STATE_RUNNING, "Preparing comparison …")

        target_folder = results.get(self._report_target_label, "")
        ref_folders   = [results[lbl] for lbl in self._report_ref_labels if lbl in results]

        if not target_folder or not os.path.isdir(target_folder):
            self._on_report_error_html(f"Target extraction folder not found:\n{target_folder}")
            return

        _bases_list  = getattr(self, '_report_bases_list', [])
        _target_src  = next((b['src_path'] for b in _bases_list if b['label'] == self._report_target_label), '')
        _ref_srcs    = [next((b['src_path'] for b in _bases_list if b['label'] == lbl), '')
                        for lbl in self._report_ref_labels]

        self._report_cmp_thread = QThread(parent=self)
        self._report_cmp_worker = ReportCompareWorker(
            target_label    = self._report_target_label,
            target_folder   = target_folder,
            ref_labels      = self._report_ref_labels,
            ref_folders     = ref_folders,
            output_root     = self._report_output_root,
            target_src_path = _target_src,
            ref_src_paths   = _ref_srcs,
            weights         = getattr(self, '_complexity_weights', None),
            bands           = getattr(self, '_complexity_bands',   None),
        )
        self._report_cmp_worker.moveToThread(self._report_cmp_thread)
        self._report_cmp_thread.started.connect(self._report_cmp_worker.run)
        self._report_cmp_worker.progress.connect(self._on_compare_progress)
        self._report_cmp_worker.log.connect(self._report_append_log)
        self._report_cmp_worker.finished.connect(self._on_compare_done_html)
        self._report_cmp_worker.error.connect(self._on_report_error_html)
        self._report_cmp_worker.finished.connect(self._report_cmp_thread.quit)
        self._report_cmp_worker.error.connect(self._report_cmp_thread.quit)
        self._report_cmp_worker.finished.connect(self._report_cmp_worker.deleteLater)
        self._report_cmp_thread.finished.connect(self._report_cmp_thread.deleteLater)
        self._report_cmp_thread.finished.connect(
            lambda t=self._report_cmp_thread: self._active_threads.remove(t)
            if t in self._active_threads else None)
        self._active_threads.append(self._report_cmp_thread)
        self._report_cmp_thread.start()

    def _on_compare_done_html(self, out_excel: str):
        """Convert the generated Excel to HTML — ask user where to save."""
        self._unlock_nav()
        from ui.widgets import StepStatusWidget
        log_output_file(out_excel, kind="Excel (pre-HTML conversion)")
        import re as _re
        _ts = __import__("datetime").datetime.now().strftime("%Y%m%d_%H%M%S")
        # Strip any trailing _FuncAtlas_Report_TIMESTAMP already in the Excel stem
        # so we never get double "FuncAtlas_Report_..._FuncAtlas_Report_..." names
        _excel_stem = os.path.splitext(os.path.basename(out_excel))[0]
        _excel_stem = _re.sub(r"_FuncAtlas_Report_\d{8}_\d{6}$", "", _excel_stem)
        default_name = f"{_excel_stem}_FuncAtlas_Report_{_ts}.html"
        default_dir  = os.path.dirname(out_excel)
        save_path, _ = QFileDialog.getSaveFileName(
            self, "Save HTML Report As", os.path.join(default_dir, default_name),
            "HTML Files (*.html);;All Files (*.*)"
        )
        if not save_path:
            self.report_step_widget.set_state("📊 Comparing & Writing HTML Report",
                                              StepStatusWidget.STATE_DONE, "Cancelled by user")
            self._set_report_progress(100, "HTML save cancelled")
            self.report_generate_html_btn.setEnabled(True)
            self.report_generate_html_btn.setText("Generate HTML Report")
            self.report_cancel_btn.setVisible(False)
            return
        try:
            html_path = self._write_html_from_excel(out_excel, save_path=save_path)
        except Exception as exc:
            self._on_report_error_html(f"HTML conversion failed: {exc}")
            return

        self._report_output_file = html_path
        self._last_report_excel  = out_excel
        self.report_step_widget.set_state("📊 Comparing & Writing HTML Report",
                                          StepStatusWidget.STATE_DONE, "HTML report saved")
        self._set_report_progress(100, "HTML report ready")
        self.report_phase_label.setText("✅ Complete")
        self.report_generate_html_btn.setEnabled(True)
        self.report_generate_html_btn.setText("Generate HTML Report")
        self.report_cancel_btn.setVisible(False)
        self.report_open_btn.setEnabled(True)
        if hasattr(self, "report_summary_chips"):
            self.report_summary_chips["phase"].set_value("Completed")
            self.report_summary_chips["result"].set_value("HTML ready")
        self._report_append_log(f"✓ HTML Report: {html_path}")
        self._unlock_nav()


    def _on_complexity_generate_html(self):
        """Generate HTML report from the Excel on the Complexity page — ask where to save."""
        report_path = getattr(self, 'complexity_report_display', None)
        if report_path is None:
            QMessageBox.warning(self, "Not Available", "Complexity report display not found.")
            return
        path = report_path.text().strip()
        if not path or not os.path.isfile(path):
            QMessageBox.warning(self, "No Report",
                "Please select or generate an Excel report first\n"
                "(use Browse Report or Generate Report on this page).")
            return
        _ts = __import__("datetime").datetime.now().strftime("%Y%m%d_%H%M%S")
        import re as _re
        _excel_stem = os.path.splitext(os.path.basename(path))[0]
        _excel_stem = _re.sub(r"_FuncAtlas_Report_\d{8}_\d{6}$", "", _excel_stem)
        default_name = f"{_excel_stem}_FuncAtlas_Report_{_ts}.html"
        save_path, _ = QFileDialog.getSaveFileName(
            self, "Save HTML Report As",
            os.path.join(os.path.dirname(path), default_name),
            "HTML Files (*.html);;All Files (*.*)"
        )
        if not save_path:
            return
        try:
            html_path = self._write_html_from_excel(path, save_path=save_path)
        except Exception as exc:
            QMessageBox.critical(self, "HTML Report Error", f"Failed to generate HTML:\n{exc}")
            return
        self._last_complexity_html = html_path
        QMessageBox.information(self, "HTML Report Ready",
            f"HTML report saved successfully.\n\nFile:\n{html_path}")

    def _restructure_excel_sheets(self, excel_path: str) -> None:
        """
        Post-process the Excel workbook:
        1. Move Complexity Level Summary + Compatibility Score Distribution
           from Construct_Summary sheet to Summary sheet.
        2. Delete Construct_Summary sheet.
        3. Remove Construct-by-Construct Totals section from Summary sheet.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from copy import copy

        wb = openpyxl.load_workbook(excel_path)
        changed = False

        # ── Step 1: Extract cx/cp sections from Construct_Summary ────────────
        cx_rows_data = []   # list of (level, count, pct, row_obj)
        cp_rows_data = []   # list of (range_label, count, pct, row_obj)

        if "Construct_Summary" in wb.sheetnames:
            ws_cs = wb["Construct_Summary"]
            section = None
            for row in ws_cs.iter_rows():
                vals = [c.value for c in row]
                if not any(v is not None for v in vals):
                    continue
                first = str(vals[0] or "").strip()
                if "Complexity Level Summary" in first:
                    section = "cx"; continue
                if "Compatibility Score Distribution" in first:
                    section = "cp"; continue
                if "Construct-by-Construct" in first:
                    section = "skip"; continue
                if first in ("Complexity Level", "Score Range", "Construct",
                             "Function Count", "% of Total"):
                    continue
                if first == "Total":
                    section = None; continue
                if section == "cx" and vals[1] is not None:
                    cx_rows_data.append(vals)
                elif section == "cp" and vals[1] is not None:
                    cp_rows_data.append(vals)

            # ── Step 2: Append cx/cp to Summary sheet ────────────────────────
            if "Summary" in wb.sheetnames and (cx_rows_data or cp_rows_data):
                ws_sum = wb["Summary"]

                # Find last used row
                last_row = ws_sum.max_row
                # Add blank separator
                last_row += 2

                header_fill   = PatternFill("solid", start_color="1F4E78")
                header_font   = Font(bold=True, color="FFFFFF", size=11)
                section_fill  = PatternFill("solid", start_color="2E75B6")
                section_font  = Font(bold=True, color="FFFFFF", size=11)
                center_align  = Alignment(horizontal="center", vertical="center")

                def _write_section(ws, start_row, title, col_headers, data_rows, row_colors):
                    # Section title
                    title_cell = ws.cell(row=start_row, column=1, value=title)
                    title_cell.fill = section_fill
                    title_cell.font = section_font
                    title_cell.alignment = center_align
                    ws.merge_cells(start_row=start_row, start_column=1,
                                   end_row=start_row, end_column=3)
                    r = start_row + 1
                    # Header row
                    for ci, hdr in enumerate(col_headers, 1):
                        c = ws.cell(row=r, column=ci, value=hdr)
                        c.fill = header_fill
                        c.font = header_font
                        c.alignment = center_align
                    r += 1
                    # Data rows
                    for vals in data_rows:
                        label = str(vals[0] or "").strip()
                        bg, fg = row_colors.get(label, ("FFFFFF", "000000"))
                        for ci, v in enumerate(vals[:3], 1):
                            cell = ws.cell(row=r, column=ci, value=v)
                            cell.fill = PatternFill("solid", start_color=bg)
                            cell.font = Font(bold=(ci == 1), color=fg)
                            if ci > 1:
                                cell.alignment = center_align
                        r += 1
                    # Total row
                    total_count = sum(int(v[1] or 0) for v in data_rows if v[1] is not None)
                    ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
                    ws.cell(row=r, column=2, value=total_count).alignment = center_align
                    ws.cell(row=r, column=2).font = Font(bold=True)
                    ws.cell(row=r, column=3, value="100%").alignment = center_align
                    ws.cell(row=r, column=3).font = Font(bold=True)
                    return r + 1

                CX_COLORS = {
                    "Low":       ("D6E4BC", "3A5A10"),
                    "Medium":    ("FFE699", "7A5500"),
                    "High":      ("F4B183", "7A3010"),
                    "Very High": ("FF7070", "5A0000"),
                    "Complex":   ("CC0000", "FFFFFF"),
                }
                CP_COLORS = {
                    "0% – 24%  (Poor)":        ("FFC7CE", "9C0006"),
                    "25% – 49%  (Low)":        ("FFEB9C", "9C5700"),
                    "50% – 74%  (Medium)":     ("FFEB9C", "9C5700"),
                    "75% – 89%  (Good)":       ("C6EFCE", "276221"),
                    "90% – 100%  (Excellent)": ("A9D18E", "1A3A00"),
                }

                next_row = last_row
                if cx_rows_data:
                    next_row = _write_section(
                        ws_sum, next_row,
                        "📊 Complexity Level Summary",
                        ["Complexity Level", "Function Count", "% of Total"],
                        cx_rows_data, CX_COLORS
                    )
                    next_row += 1

                if cp_rows_data:
                    _write_section(
                        ws_sum, next_row,
                        "🔗 Compatibility Score Distribution",
                        ["Score Range", "Function Count", "% of Total"],
                        cp_rows_data, CP_COLORS
                    )

                changed = True

            # ── Step 3: Delete Construct_Summary sheet ────────────────────────
            del wb["Construct_Summary"]
            changed = True

        # ── Step 4: Remove Construct-by-Construct Totals from Summary ────────
        if "Summary" in wb.sheetnames:
            ws_sum = wb["Summary"]
            rows_to_delete = []
            in_construct = False
            for row in ws_sum.iter_rows():
                first = str(row[0].value or "").strip()
                if "Construct-by-Construct" in first:
                    in_construct = True
                if in_construct:
                    rows_to_delete.append(row[0].row)
            # Delete from bottom up to preserve row indices
            for r in reversed(rows_to_delete):
                ws_sum.delete_rows(r)
            if rows_to_delete:
                changed = True

        if changed:
            wb.save(excel_path)
        wb.close()

    def _write_html_from_excel(self, excel_path: str, save_path: str = None) -> str:
        """
        Read all four sheets of FuncAtlas_Report.xlsx and produce a
        styled standalone HTML report.
        The original Excel file is never modified — restructuring is done
        on a temporary copy that is deleted after use.
        """
        import openpyxl
        import tempfile
        import shutil

        # Excel is already restructured after Generate Report.
        # For the report-page HTML flow (which doesn't go through Generate Report),
        # restructure on a temp copy so the original is never modified.
        import tempfile, shutil as _shutil
        if "Construct_Summary" in openpyxl.load_workbook(excel_path, read_only=True).sheetnames:
            tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
            os.close(tmp_fd)
            _shutil.copy2(excel_path, tmp_path)
            try:
                self._restructure_excel_sheets(tmp_path)
                wb = openpyxl.load_workbook(tmp_path, data_only=True)
            finally:
                try:
                    os.remove(tmp_path)
                except OSError:
                    pass
        else:
            wb = openpyxl.load_workbook(excel_path, data_only=True)

        # ── Sheet: Summary (contains key-value rows + complexity level + compat distribution) ──
        cx_summary  = {}   # level -> (count, pct)
        cp_summary  = {}   # range label -> (count, pct)
        summary_rows = []
        if "Summary" in wb.sheetnames:
            ws2 = wb["Summary"]
            all_vals = [r for r in ws2.iter_rows(min_row=1, values_only=True)]
            section = None
            for row in all_vals:
                if not row or all(v is None for v in row):
                    continue
                first = str(row[0] or "").strip()
                if "Complexity Level Summary" in first:
                    section = "cx"; continue
                if "Compatibility Score Distribution" in first:
                    section = "cp"; continue
                if "Construct-by-Construct" in first:
                    # skip construct-by-construct totals entirely
                    section = "skip"; continue
                if first in ("Complexity Level", "Score Range", "Construct", "Function Count", "% of Total"):
                    continue   # header row
                if first == "Total":
                    section = None; continue
                if section == "cx" and row[1] is not None:
                    cx_summary[first] = (row[1], str(row[2] or ""))
                elif section == "cp" and row[1] is not None:
                    cp_summary[first] = (row[1], str(row[2] or ""))
                elif section is None and row[0] is not None and row[1] is not None:
                    summary_rows.append((str(row[0]), str(row[1])))

        # ── Sheet 1: Function_Match_Report ────────────────────────────────────
        sheet1_data = {}
        if "Function_Match_Report" in wb.sheetnames:
            ws1 = wb["Function_Match_Report"]
            hdrs1 = [str(c).strip() if c else "" for c in next(ws1.iter_rows(min_row=1, max_row=1, values_only=True))]
            try:
                col_file   = hdrs1.index("File Name")
                col_func   = hdrs1.index("Function Name")
                col_status = hdrs1.index("Reuse/New")
                col_which  = hdrs1.index("Suggested Reference Base")
            except ValueError:
                col_file, col_func, col_status, col_which = 0, 1, len(hdrs1)-3, len(hdrs1)-2
            # LOC column — "Total LOC" is column D (index 3) in your Excel sheet
            col_loc = None
            for _loc_name in ("Total LOC", "LOC", "Lines of Code", "Length of Code", "Line Count", "Length"):
                if _loc_name in hdrs1:
                    col_loc = hdrs1.index(_loc_name)
                    break
            if col_loc is None and len(hdrs1) > 3:
                col_loc = 3   # fallback: column D (0-based index 3)
            for row in ws1.iter_rows(min_row=2, values_only=True):
                if not row or row[col_func] is None:
                    continue
                key = (str(row[col_file] or ""), str(row[col_func] or ""))
                sheet1_data[key] = {
                    "reuse_status": str(row[col_status] or ""),
                    "which_base":   str(row[col_which]  or ""),
                    "loc":          str(row[col_loc] or "—") if col_loc is not None else "—",
                }

        # ── Sheet 3: Complexity_Compatibility ─────────────────────────────────
        sheet3_data = {}
        if "Complexity_Compatibility" in wb.sheetnames:
            ws3 = wb["Complexity_Compatibility"]
            hdrs3 = [str(c).strip() if c else "" for c in next(ws3.iter_rows(min_row=1, max_row=1, values_only=True))]
            try:
                c3_file  = hdrs3.index("File Name")
                c3_func  = hdrs3.index("Function Name")
                c3_level = hdrs3.index("Complexity Level")
            except ValueError:
                c3_file, c3_func, c3_level = 0, 1, len(hdrs3) - 1
            for row in ws3.iter_rows(min_row=2, values_only=True):
                if not row or row[c3_func] is None:
                    continue
                key = (str(row[c3_file] or ""), str(row[c3_func] or ""))
                sheet3_data[key] = str(row[c3_level] or "")

        # ── Sheet 4: Compatibility_Score ──────────────────────────────────────
        sheet4_data = {}
        if "Compatibility_Score" in wb.sheetnames:
            ws4 = wb["Compatibility_Score"]
            hdrs4 = [str(c).strip() if c else "" for c in next(ws4.iter_rows(min_row=1, max_row=1, values_only=True))]
            try:
                c4_file  = hdrs4.index("File Name")
                c4_func  = hdrs4.index("Function Name")
                c4_compat= hdrs4.index("Compatibility %")
            except ValueError:
                c4_file, c4_func, c4_compat = 0, 1, len(hdrs4) - 1
            for row in ws4.iter_rows(min_row=2, values_only=True):
                if not row or row[c4_func] is None:
                    continue
                key = (str(row[c4_file] or ""), str(row[c4_func] or ""))
                sheet4_data[key] = str(row[c4_compat] or "")

        wb.close()

        # ── Build merged rows ─────────────────────────────────────────────────
        merged = []
        for (file_name, func_name), s1 in sheet1_data.items():
            key = (file_name, func_name)
            merged.append({
                "file_name":    file_name,
                "func_name":    func_name,
                "reuse_status": s1["reuse_status"],
                "which_base":   s1["which_base"],
                "loc":          s1.get("loc", "—"),
                "complexity":   sheet3_data.get(key, "—"),
                "compat_pct":   sheet4_data.get(key, "—"),
            })

        # ── HTML helpers ──────────────────────────────────────────────────────
        def _status_class(v):
            if v == "Reuse":           return "reuse"
            if v == "Reuse (Modified)": return "modified"
            if v == "New":             return "new-func"
            return ""

        def _complexity_class(v):
            mapping = {"Low": "cx-low", "Medium": "cx-medium",
                       "High": "cx-high", "Very High": "cx-veryhigh", "Complex": "cx-complex"}
            return mapping.get(v, "")

        def _compat_class(v):
            try:
                pct = float(str(v).replace("%", "").strip())
                if pct >= 75: return "cp-high"
                if pct >= 40: return "cp-mid"
                return "cp-low"
            except Exception:
                return ""

        # ── Summary HTML block ────────────────────────────────────────────────
        CX_COLORS = {
            "Low":       ("#d6e4bc", "#3a5a10"),
            "Medium":    ("#ffe699", "#7a5500"),
            "High":      ("#f4b183", "#7a3010"),
            "Very High": ("#ff7070", "#5a0000"),
            "Complex":   ("#cc0000", "#ffffff"),
        }
        CP_COLORS = {
            "0% – 24%  (Poor)":         ("#ffc7ce", "#9c0006"),
            "25% – 49%  (Low)":         ("#ffeb9c", "#9c5700"),
            "50% – 74%  (Medium)":      ("#ffeb9c", "#9c5700"),
            "75% – 89%  (Good)":        ("#c6efce", "#276221"),
            "90% – 100%  (Excellent)":  ("#a9d18e", "#1a3a00"),
        }

        summary_html = ""
        # Build summary key-value block (narrow column)
        kv_block = ""
        if summary_rows:
            rows_html = "".join(
                f'<tr><td class="sum-key">{k}</td><td class="sum-val">{v}</td></tr>'
                for k, v in summary_rows
            )
            kv_block = f"""
    <div class="sum-card sum-card-narrow">
      <h2>\U0001f4cb Summary</h2>
      <table class="summary-table">
        <tbody>{rows_html}</tbody>
      </table>
    </div>"""

        # Build complexity level card
        cx_block = ""
        if cx_summary:
            cx_rows = "".join(
                f'<tr>'
                f'<td class="sum-badge" style="background:{CX_COLORS.get(lv,("#eee","#333"))[0]};color:{CX_COLORS.get(lv,("#eee","#333"))[1]}">{lv}</td>'
                f'<td class="sum-cnt">{cnt}</td>'
                f'<td class="sum-pct">{pct}</td>'
                f'</tr>'
                for lv, (cnt, pct) in cx_summary.items()
            )
            cx_block = f"""
    <div class="sum-card">
      <h2>\U0001f4ca Complexity Level Distribution</h2>
      <table class="sum-table">
        <thead><tr><th>Level</th><th>Functions</th><th>% of Total</th></tr></thead>
        <tbody>{cx_rows}</tbody>
      </table>
    </div>"""

        # Build compatibility score card
        cp_block = ""
        if cp_summary:
            cp_rows = "".join(
                f'<tr>'
                f'<td class="sum-badge" style="background:{CP_COLORS.get(label,("#eee","#333"))[0]};color:{CP_COLORS.get(label,("#eee","#333"))[1]}">{label}</td>'
                f'<td class="sum-cnt">{cnt}</td>'
                f'<td class="sum-pct">{pct}</td>'
                f'</tr>'
                for label, (cnt, pct) in cp_summary.items()
            )
            cp_block = f"""
    <div class="sum-card">
      <h2>\U0001f517 Compatibility Score Distribution</h2>
      <table class="sum-table">
        <thead><tr><th>Score Range</th><th>Functions</th><th>% of Total</th></tr></thead>
        <tbody>{cp_rows}</tbody>
      </table>
    </div>"""

        if kv_block or cx_block or cp_block:
            summary_html = f"""
<section class="summary-block">
  <div class="summary-grid-3">
{kv_block}
{cx_block}
{cp_block}
  </div>
</section>"""

        # ── Detail table rows ─────────────────────────────────────────────────
        detail_rows_html = ""
        html_idx = 1
        for row in merged:
            if row["reuse_status"] == "Reuse":
                continue
            sc  = _status_class(row["reuse_status"])
            cxc = _complexity_class(row["complexity"])
            cpc = _compat_class(row["compat_pct"])
            detail_rows_html += (
                f'<tr>'
                f'<td class="center">{html_idx}</td>'
                f'<td>{row["file_name"]}</td>'
                f'<td>{row["func_name"]}</td>'
                f'<td class="center loc-cell">{row["loc"]}</td>'
                f'<td class="center {sc}">{row["reuse_status"]}</td>'
                f'<td>{row["which_base"]}</td>'
                f'<td class="center {cxc}">{row["complexity"]}</td>'
                f'<td class="center {cpc}">{row["compat_pct"]}</td>'
                f'</tr>\n'
            )
            html_idx += 1

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>FuncAtlas Report</title>
<style>
  *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    font-family: Arial, sans-serif; font-size: 13px;
    background: #f0f4f9; color: #1a2a3a; padding: 28px 32px;
  }}
  h1 {{ color: #1f4e78; font-size: 22px; margin-bottom: 6px; }}
  .subtitle {{ color: #5a7a9a; font-size: 12px; margin-bottom: 24px; }}
  h2 {{ color: #1f4e78; font-size: 15px; margin-bottom: 10px; }}

  /* ── Summary block ── */
  .summary-block {{ margin-bottom: 32px; }}
  .summary-grid-3 {{
    display: grid;
    grid-template-columns: minmax(220px, 0.8fr) 1fr 1fr;
    gap: 20px;
    align-items: start;
  }}
  .sum-card {{
    background: #fff; border-radius: 10px;
    box-shadow: 0 2px 10px rgba(0,0,0,.09);
    padding: 18px 20px;
  }}
  .sum-card-narrow {{
    background: #fff; border-radius: 10px;
    box-shadow: 0 2px 10px rgba(0,0,0,.09);
    padding: 18px 20px;
  }}
  .sum-card h2, .sum-card-narrow h2 {{ margin-bottom: 12px; font-size: 14px; color: #1f4e78; }}
  .sum-table {{ border-collapse: collapse; width: 100%; }}
  .sum-table th {{
    background: #1f4e78; color: #fff; padding: 7px 12px;
    font-size: 11px; text-align: center;
  }}
  .sum-table td {{ padding: 6px 12px; border-bottom: 1px solid #e0ecf8; }}
  .sum-table tr:last-child td {{ border-bottom: none; }}
  td.sum-badge {{ font-weight: bold; border-radius: 4px; padding: 4px 10px; }}
  td.sum-cnt {{ text-align: center; font-weight: bold; color: #1a3a5c; }}
  td.sum-pct {{ text-align: center; color: #5a7a9a; }}

  /* legacy summary table */
  .summary-table {{ border-collapse: collapse; width: 100%; background: transparent;
                    border-radius: 8px; overflow: hidden;
                    box-shadow: 0 2px 8px rgba(0,0,0,.08); }}
  .summary-table td {{ padding: 7px 14px; border-bottom: 1px solid #dce8f4; }}
  .summary-table tr:last-child td {{ border-bottom: none; }}
  td.sum-key {{ background: #ebf3fb; font-weight: bold; color: #1a3a5c;
                width: 200px; white-space: nowrap; }}
  td.sum-val {{ color: #2a4a6a; }}

  /* ── Legend ── */
  .legend {{ display: flex; flex-wrap: wrap; gap: 10px; margin-bottom: 14px; }}
  .legend span {{ display: inline-block; padding: 3px 14px; border-radius: 12px;
                  font-size: 11px; font-weight: bold; }}

  /* ── Detail table ── */
  .detail-table {{ border-collapse: collapse; width: 100%; background: #fff;
                   box-shadow: 0 2px 10px rgba(0,0,0,.10); border-radius: 8px;
                   overflow: hidden; }}
  .detail-table th {{
    background: #1f4e78; color: #fff; padding: 10px 12px;
    text-align: center; font-size: 12px; white-space: nowrap;
  }}
  .detail-table td {{ padding: 7px 12px; border: 1px solid #d0e4f0; vertical-align: middle; }}
  .detail-table tr:nth-child(even) td {{ background: #f5f9fd; }}
  .detail-table tr:hover td {{ background: #ddeeff; }}
  .center {{ text-align: center; }}

  /* Reuse/New */
  .reuse     {{ background: #c6efce !important; color: #276221; font-weight: bold; }}
  .modified  {{ background: #ffeb9c !important; color: #9c5700; font-weight: bold; }}
  .new-func  {{ background: #ffc7ce !important; color: #9c0006; font-weight: bold; }}

  /* Complexity Level */
  .cx-low      {{ background: #d6e4bc !important; color: #3a5a10; font-weight: bold; }}
  .cx-medium   {{ background: #ffe699 !important; color: #7a5500; font-weight: bold; }}
  .cx-high     {{ background: #f4b183 !important; color: #7a3010; font-weight: bold; }}
  .cx-veryhigh {{ background: #ff7070 !important; color: #5a0000; font-weight: bold; }}
  .cx-complex  {{ background: #cc0000 !important; color: #fff;    font-weight: bold; }}

  /* Compatibility % */
  .cp-high {{ background: #c6efce !important; color: #276221; font-weight: bold; }}
  .cp-mid  {{ background: #ffeb9c !important; color: #9c5700; font-weight: bold; }}
  .cp-low  {{ background: #ffc7ce !important; color: #9c0006; font-weight: bold; }}

  /* LOC — Lines of Code column */
  .loc-cell {{ background: #e8f0fe !important; color: #1a3a7a; font-weight: bold; }}
</style>
</head>
<body>

<h1>FuncAtlas — Function Analysis Report</h1>
<p class="subtitle">Generated from: {os.path.basename(excel_path)}</p>

{summary_html}

<section>
  <h2>Function Detail</h2>
  <div class="legend">
    <span class="modified">Reuse (Modified)</span>
    <span class="new-func">New</span>
    <span style="width:12px;"></span>
    <span class="cx-low">Low</span>
    <span class="cx-medium">Medium</span>
    <span class="cx-high">High</span>
    <span class="cx-veryhigh">Very High</span>
    <span class="cx-complex">Complex</span>
    <span style="width:12px;"></span>
    <span class="cp-high">Compat ≥75%</span>
    <span class="cp-mid">Compat ≥40%</span>
    <span class="cp-low">Compat &lt;40%</span>
  </div>
  <table class="detail-table">
    <thead>
      <tr>
        <th>S.No</th>
        <th>File Name</th>
        <th>Function Name</th>
        <th>LOC</th>
        <th>Status</th>
        <th>Suggested Base</th>
        <th>Complexity Level</th>
        <th>Compatibility %</th>
      </tr>
    </thead>
    <tbody>
{detail_rows_html}    </tbody>
  </table>
</section>

</body>
</html>"""

        if save_path:
            html_path = save_path
        else:
            import re as _re
            _stem = os.path.splitext(os.path.basename(excel_path))[0]
            _stem = _re.sub(r"_FuncAtlas_Report_\d{8}_\d{6}$", "", _stem)
            _ts   = __import__("datetime").datetime.now().strftime("%Y%m%d_%H%M%S")
            html_path = os.path.join(os.path.dirname(excel_path), f"{_stem}_FuncAtlas_Report_{_ts}.html")
        with open(html_path, "w", encoding="utf-8") as fh:
            fh.write(html)
        return html_path

    def _on_report_error_html(self, msg: str):
        self._unlock_nav()
        if msg == '__CANCELLED__':
            self._report_append_log("⛔ Generation cancelled by user.")
            self.report_phase_label.setText("⛔ Cancelled")
            self.report_generate_html_btn.setEnabled(True)
            self.report_generate_html_btn.setText("Generate HTML Report")
            self.report_cancel_btn.setVisible(False)
            if hasattr(self, "report_summary_chips"):
                self.report_summary_chips["phase"].set_value("Cancelled")
                self.report_summary_chips["result"].set_value("—")
            self._unlock_nav()
            return
        self._report_append_log(f"ERROR: {msg}")
        self.report_phase_label.setText("⚠ Error")
        self.report_generate_html_btn.setEnabled(True)
        self.report_generate_html_btn.setText("Generate HTML Report")
        self.report_cancel_btn.setVisible(False)
        if hasattr(self, "report_summary_chips"):
            self.report_summary_chips["phase"].set_value("Error")
            self.report_summary_chips["result"].set_value("Failed")
        self._unlock_nav()
        QMessageBox.critical(self, "HTML Report Error", msg)

    # ── Excel report generation ───────────────────────────────────────────────
    def _on_compare_done(self, out_file: str):
        from ui.widgets import StepStatusWidget
        self._unlock_nav()
        self._report_output_file = out_file
        self._last_report_excel  = out_file
        log_output_file(out_file, kind="Excel Report")
        _log.info("_on_compare_done: report ready at %s", out_file)
        if hasattr(self, 'complexity_report_display'):
            self.complexity_report_display.setText(out_file)
        self.report_step_widget.set_state("📊 Comparing & Writing Excel Report",
                                          StepStatusWidget.STATE_DONE, "Excel report written")
        self._set_report_progress(100, "Report ready")
        self.report_phase_label.setText("✅ Complete")
        self.report_generate_btn.setEnabled(True)
        self.report_generate_btn.setText("Generate Report")
        self.report_cancel_btn.setVisible(False)
        self.report_open_btn.setEnabled(True)
        if hasattr(self, "report_summary_chips"):
            self.report_summary_chips["phase"].set_value("Completed")
            self.report_summary_chips["result"].set_value("Excel ready")
        # ── Stop elapsed timer ────────────────────────────────────────────
        if hasattr(self, "_report_timer"):
            self._report_timer.stop()
        elapsed = int(time.time() - getattr(self, "_report_start_time", time.time()))
        elapsed_str = f"{elapsed // 60}m {elapsed % 60}s" if elapsed >= 60 else f"{elapsed}s"
        self._report_append_log(f"✓ Report: {out_file}")
        self._report_append_log(f"⏱ Total time: {elapsed_str}")
        if hasattr(self, "report_status_label"):
            self.report_status_label.setText(f"Done — {elapsed_str}")
        if hasattr(self, "report_timer_lbl"):
            self.report_timer_lbl.setText(f"⏱ {elapsed_str}")
        self._show_report_success_popup(out_file, elapsed_str)


    def _show_report_success_popup(self, out_file: str, elapsed_str: str):
        """Show a compact success dialog after report generation."""
        from PySide6.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton
        from PySide6.QtCore import Qt

        dlg = QDialog(self)
        dlg.setWindowTitle("Success")
        dlg.setFixedWidth(420)

        outer = QVBoxLayout(dlg)
        outer.setContentsMargins(18, 16, 18, 14)
        outer.setSpacing(10)

        top = QHBoxLayout()
        top.setSpacing(12)
        icon_lbl = QLabel("ℹ")
        icon_lbl.setFixedSize(38, 38)
        icon_lbl.setAlignment(Qt.AlignCenter)
        icon_lbl.setStyleSheet(
            "background:#0078D4; color:white; border-radius:19px; font-size:20px; font-weight:700;"
        )
        msg_lbl = QLabel("Report generated successfully.")
        msg_lbl.setWordWrap(True)
        msg_lbl.setStyleSheet("font-size:13px;")
        top.addWidget(icon_lbl, 0, Qt.AlignTop)
        top.addWidget(msg_lbl, 1)
        outer.addLayout(top)

        detail_lbl = QLabel(f"Updated report:\n{out_file}\n\n⏱ Time: {elapsed_str}")
        detail_lbl.setWordWrap(True)
        detail_lbl.setStyleSheet("font-size:12px; color:#444; margin-left:50px;")
        outer.addWidget(detail_lbl)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        ok_btn = QPushButton("OK")
        ok_btn.setFixedSize(80, 28)
        ok_btn.setDefault(True)
        ok_btn.clicked.connect(dlg.accept)
        btn_row.addWidget(ok_btn)
        outer.addLayout(btn_row)

        dlg.exec()

    def _on_report_error(self, msg: str):
        self._unlock_nav()
        if hasattr(self, "_report_timer"):
            self._report_timer.stop()
        if msg == '__CANCELLED__':
            self._report_append_log("⛔ Generation cancelled by user.")
            self.report_phase_label.setText("⛔ Cancelled")
            self.report_status_label.setText("Cancelled by user.")
            self.report_generate_btn.setEnabled(True)
            self.report_generate_btn.setText("Generate Report")
            self.report_cancel_btn.setVisible(False)
            if hasattr(self, "report_summary_chips"):
                self.report_summary_chips["phase"].set_value("Cancelled")
                self.report_summary_chips["result"].set_value("—")
            self._unlock_nav()
            return
        self._report_append_log(f"ERROR: {msg}")
        self.report_phase_label.setText("⚠ Error")
        self.report_status_label.setText("Error — see log below.")
        self.report_generate_btn.setEnabled(True)
        self.report_generate_btn.setText("Generate Report")
        self.report_cancel_btn.setVisible(False)
        if hasattr(self, "report_summary_chips"):
            self.report_summary_chips["phase"].set_value("Error")
            self.report_summary_chips["result"].set_value("Failed")
        self._unlock_nav()
        QMessageBox.critical(self, "Report Error", msg)

    def _on_report_cancel(self):
        """Cancel any in-progress report generation."""
        log_user_action("click", "Cancel Report button", page="report")
        self.report_cancel_btn.setEnabled(False)
        self.report_cancel_btn.setText("Cancelling …")
        # Signal extraction worker to stop
        if hasattr(self, '_report_ext_worker') and self._report_ext_worker:
            self._report_ext_worker._cancel_requested = True
        # Signal compare worker to stop
        if hasattr(self, '_report_cmp_worker') and self._report_cmp_worker:
            self._report_cmp_worker._cancel_requested = True

    def _report_append_log(self, line: str):
        self.report_log_box.append(line)
        if hasattr(self, "report_log_panel"):
            self.report_log_panel.set_expanded(True)
        sb = self.report_log_box.verticalScrollBar()
        sb.setValue(sb.maximum())

    def _update_report_elapsed(self):
        """Called every second to update elapsed time display while report runs."""
        elapsed = int(time.time() - getattr(self, "_report_start_time", time.time()))
        mins, secs = divmod(elapsed, 60)
        btn_label   = f"Running … {mins}m {secs:02d}s" if mins else f"Running … {secs}s"
        timer_label = f"⏱ {mins}m {secs:02d}s" if mins else f"⏱ {secs}s"
        if hasattr(self, "report_generate_btn") and not self.report_generate_btn.isEnabled():
            self.report_generate_btn.setText(btn_label)
        if hasattr(self, "report_timer_lbl"):
            self.report_timer_lbl.setText(timer_label)

    def _on_report_clear(self):
        log_user_action("click", "Clear Report button", page="report")
        self.report_output_field.clear_selection()
        self._set_report_progress(0, "Idle — ready to generate.")
        self.report_log_box.clear()
        self.report_phase_label.setText("Waiting to start")
        if hasattr(self, "report_timer_lbl"):
            self.report_timer_lbl.setText("")
        self.report_open_btn.setEnabled(False)
        self._report_output_file = ""
        self.report_step_widget.clear_steps()
        self._report_total_steps = 0
        self._update_report_step_chip()
        if hasattr(self, "report_summary_chips"):
            self.report_summary_chips["phase"].set_value("Idle")
            self.report_summary_chips["result"].set_value("Not generated")
        if hasattr(self, "report_log_panel"):
            self.report_log_panel.set_expanded(False)

    def _on_report_open(self):
        """Open the Excel report generated on the Report page directly."""
        from PySide6.QtCore import QUrl

        excel_path = getattr(self, "_last_report_excel", "") or ""
        log_user_action("click", "Open Report button", page="report",
                        extra=f"excel_path={excel_path!r}")

        if excel_path and os.path.exists(excel_path):
            QDesktopServices.openUrl(QUrl.fromLocalFile(excel_path))
            return

        path, _ = QFileDialog.getOpenFileName(
            self, "Open Report", "",
            "Excel Files (*.xlsx *.xls);;All Files (*.*)"
        )
        if path:
            QDesktopServices.openUrl(QUrl.fromLocalFile(path))

    # ── report help menus ─────────────────────────────────────────────────────
    def _show_report_help_menu(self):
        menu = QMenu(self)
        menu.addAction('📖  How to Use',     self._show_report_how_to_use)
        menu.addAction('✅  Prerequisites',  self._show_report_prerequisites)
        menu.addAction('📋  View Log File',  self._show_report_log_dialog)
        menu.exec(self.report_help_btn.mapToGlobal(self.report_help_btn.rect().bottomLeft()))

    def _show_report_how_to_use(self):
        sections = [
            {"title": "Select Output Folder",        "body": "Choose the folder where extracted functions and the final Excel report will be saved."},
            {"title": "Load Reference Data First",    "body": "Go to Input → Reference Bases and submit the target base, reference bases, and function list before running the report."},
            {"title": "Click Generate Report",        "body": "The tool first extracts functions from the target and every reference base."},
            {"title": "Wait for Step Completion",     "body": "Each base runs one by one. The step list shows Waiting, Running, and Complete states clearly."},
            {"title": "Comparison Starts Automatically", "body": "After extraction, FuncAtlas compares target functions against the extracted reference functions and writes the Excel report."},
            {"title": "Open Report",                  "body": "Once processing finishes, use Open Report to open the generated Excel file directly."},
            {"title": "Check Detailed Log",           "body": "Use View Log File from the Help menu if you need to inspect progress details or troubleshoot failures."},
        ]
        HelpOverlayDialog(self, "Help — How to Use", sections,
            footer_text="⚠ Prerequisites: Windows OS · Microsoft Excel installed · Input data must be submitted before generating the report.",
            tip_text="💡 Tip: Use a local drive path for output. Network folders and locked Excel files slow this workflow down or break it."
        ).exec()

    def _show_report_prerequisites(self):
        sections = [
            {"title": "Reference data must be submitted",
             "body": "The Report and Diff pages both depend on the target base and reference bases already loaded from Input → Reference Bases. Without this, report generation and diff comparison will fail or produce empty output."},
            {"title": "Reference data required for Diff comparison",
             "body": "The Diff page compares extracted function sets between your target and reference bases. You must submit at least one target and one reference base from the Input page before running any diff."},
            {"title": "Output folder must be writable",
             "body": "Pick a folder where Excel files and extracted text files can be created. Read-only or network folders may fail silently."},
            {"title": "Close locked Excel files",
             "body": "If the report file is already open in Excel, writing can fail or force a renamed output. Always close the file before regenerating."},
            {"title": "Large inputs take real time",
             "body": "This build is threaded so the UI stays responsive, but extraction and comparison time still depends on file count and size. Monitor progress using the step indicators."},
        ]
        HelpOverlayDialog(self, "Prerequisites", sections,
            footer_text="⚠ Do not run Report or Diff before loading target and reference sources. That is the main reason users get empty or failed output.",
            tip_text="💡 Tip: Keep output on SSD/local disk and avoid opening the same report while generation is still running."
        ).exec()

    def _show_report_log_dialog(self):
        dlg = QDialog(self)
        dlg.setWindowTitle('Report Log')
        dlg.setModal(True)
        dlg.resize(860, 520)
        outer = QVBoxLayout(dlg)
        outer.setContentsMargins(0, 0, 0, 0)
        shell = QFrame()
        shell.setObjectName('aboutDialogShell')
        shell_layout = QVBoxLayout(shell)
        shell_layout.setContentsMargins(24, 24, 24, 24)
        head = QHBoxLayout()
        title_lbl = QLabel('Report Log File')
        title_lbl.setObjectName('aboutDialogTitle')
        close_btn = QPushButton('×')
        close_btn.setObjectName('aboutDialogCloseButton')
        close_btn.setFixedSize(42, 42)
        close_btn.clicked.connect(dlg.accept)
        head.addWidget(title_lbl); head.addStretch(); head.addWidget(close_btn)
        shell_layout.addLayout(head)
        log_box = QTextEdit()
        log_box.setReadOnly(True)
        log_box.setPlainText(self.report_log_box.toPlainText() or 'No log entries yet.')
        shell_layout.addWidget(log_box, 1)
        outer.addWidget(shell)
        dlg.exec()

    # ── About dialog ──────────────────────────────────────────────────────────
    def show_about_dialog(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("About FuncAtlas")
        dlg.setModal(True)
        dlg.setFixedSize(790, 620)
        outer = QVBoxLayout(dlg)
        outer.setContentsMargins(0, 0, 0, 0)
        shell = QFrame(); shell.setObjectName("aboutDialogShell")
        shell_layout = QVBoxLayout(shell)
        shell_layout.setContentsMargins(34, 26, 34, 28)
        shell_layout.setSpacing(20)

        head = QHBoxLayout()
        title_lbl = QLabel("About — FuncAtlas"); title_lbl.setObjectName("aboutDialogTitle")
        close_btn = QPushButton("×"); close_btn.setObjectName("aboutDialogCloseButton")
        close_btn.setFixedSize(42, 42); close_btn.clicked.connect(dlg.accept)
        head.addWidget(title_lbl); head.addStretch(); head.addWidget(close_btn)
        shell_layout.addLayout(head)

        hero = QFrame(); hero.setObjectName("aboutDialogHero")
        hero_layout = QVBoxLayout(hero)
        hero_layout.setContentsMargins(22, 20, 22, 18); hero_layout.setSpacing(6)
        h1 = QLabel("FuncAtlas — Automated Function Analysis & Reporting Engine")
        h1.setObjectName("aboutDialogHeroTitle")
        h2 = QLabel("Version 1.0.0 · Reuse Analysis Workspace")
        h2.setObjectName("aboutDialogHeroSubtitle")
        hero_layout.addWidget(h1); hero_layout.addWidget(h2)
        shell_layout.addWidget(hero)

        bullets_frame = QFrame(); bullets_frame.setObjectName("aboutDialogBody")
        bullets_layout = QVBoxLayout(bullets_frame)
        bullets_layout.setContentsMargins(2, 8, 2, 2); bullets_layout.setSpacing(18)
        items = [
            ("🔎", "Compare target and reference source files from selected folders"),
            ("📄", "Read function lists from TXT and XLSX files without changing the source data"),
            ("📊", "Generate Excel summary reports for matched functions and reuse status"),
            ("🧭", "Preview detected files and function bodies inside the explorer workspace"),
            ("🔐", "Background processing with progress tracking and safer output handling"),
            ("🧹", "Avoid locked-file overwrite failures by writing a new report name when needed"),
            ("📝", "Detailed runtime logging in the report panel for easier troubleshooting"),
        ]
        for icon_text, line in items:
            row = QHBoxLayout(); row.setSpacing(14)
            icon = QLabel(icon_text); icon.setObjectName("aboutDialogBulletIcon")
            icon.setFixedWidth(28); icon.setAlignment(Qt.AlignTop | Qt.AlignHCenter)
            txt = QLabel(line); txt.setObjectName("aboutDialogBulletText"); txt.setWordWrap(True)
            row.addWidget(icon); row.addWidget(txt, 1)
            bullets_layout.addLayout(row)
        shell_layout.addWidget(bullets_frame)
        outer.addWidget(shell)

        dlg.setStyleSheet(f"""
            QDialog {{ background: transparent; }}
            QFrame#aboutDialogShell {{ background: {self.theme['bg_header']}; border: 1px solid {self.theme['border_strong']}; border-radius: 26px; }}
            QLabel#aboutDialogTitle {{ color: {self.accent_color.name()}; font-size: 17px; font-weight: 900; background: transparent; }}
            QFrame#aboutDialogHero {{ background: {self.theme['bg_card']}; border: 1px solid {self.theme['border']}; border-radius: 18px; }}
            QLabel#aboutDialogHeroTitle {{ color: {self.theme['text_primary']}; font-size: 15px; font-weight: 900; background: transparent; }}
            QLabel#aboutDialogHeroSubtitle {{ color: {self.theme['text_muted']}; font-size: 11px; background: transparent; }}
            QLabel#aboutDialogBulletIcon {{ color: {self.accent_color.name()}; font-size: 17px; background: transparent; }}
            QLabel#aboutDialogBulletText {{ color: {self.theme['text_secondary']}; font-size: 12px; font-weight: 600; background: transparent; }}
            QPushButton#aboutDialogCloseButton {{ background: transparent; color: {self.theme['text_muted']}; border: none; font-size: 26px; font-weight: 500; }}
            QPushButton#aboutDialogCloseButton:hover {{ color: {self.theme['text_primary']}; }}
        """)
        dlg.exec()

    # ── resize handler ────────────────────────────────────────────────────────
    def resizeEvent(self, event):
        super().resizeEvent(event)
        try:
            self.refresh_home_hero_image()
        except Exception:
            pass

    def closeEvent(self, event):
        """Stop and wait for every background thread before closing."""
        _log.info("Application closing — stopping background threads")
        log_user_action("close", "Application window")
        for thread in list(self._active_threads):
            if thread and thread.isRunning():
                thread.quit()
                thread.wait(3000)

        for field_attr in ("con_func_col_field", "con_base_col_field"):
            field = getattr(self, field_attr, None)
            if field is None:
                continue
            for thread in list(getattr(field, "_threads", [])):
                if thread and thread.isRunning():
                    thread.quit()
                    thread.wait(2000)

        for t_attr in ("_report_ext_thread", "_report_cmp_thread", "con_thread", "_upfront_thread"):
            t = getattr(self, t_attr, None)
            if t and t.isRunning():
                # Signal worker to cancel before quitting thread
                w_attr = t_attr.replace("_thread", "_worker")
                w = getattr(self, w_attr, None)
                if w and hasattr(w, "cancel"):
                    w.cancel()
                t.quit()
                t.wait(3000)

        super().closeEvent(event)