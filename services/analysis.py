"""
services/analysis.py  –  scan/match logic + QObject workers 


"""
import os, re, json as _json
from collections import OrderedDict
from concurrent.futures import ThreadPoolExecutor, as_completed

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from PySide6.QtCore import QObject, Signal

from core.utils import (
    normalize_path, normalize_name, clean_text, extract_function_body,
    is_probable_function_name, iter_source_files, detect_functions_in_file,
    extract_excel_column_letters, extract_excel_row_number,
)
from core.logger import get_logger, log_function_extraction, log_output_file, log_file_upload
from core.function_cache import FUNCTION_CACHE

_log = get_logger(__name__)

# ── Column-header scoring ─────────────────────────────────────────────────────
def score_header_for_function(text):
    t = normalize_name(text)
    if not t: return 0
    score = 0
    if t in {"function","function name","function_name","func","func name","func_name"}: score += 100
    if "function" in t: score += 40
    if "func" in t: score += 35
    if "name" in t: score += 10
    # strong penalties — these are never function-name columns
    if "base" in t: score -= 25
    if "file" in t: score -= 40
    if "path" in t: score -= 40
    if "source" in t: score -= 30
    if t in {"batch","no","s.no","sno","sn","sr","serial","number","svn","sl"}: score -= 50
    if "s.no" in t or "serial" in t or t in {"sn","sno"}: score -= 30
    return score

def score_header_for_base(text):
    t = normalize_name(text)
    if not t: return 0
    score = 0
    # Exact matches for base/file-path column headers
    if t in {"base","base name","base_name","bases"}: score += 100
    if t in {"file name","file_name","filename","file path","file_path",
             "filepath","source file","source_file","sourcefile",
             "svn path","svn_path","path","source path","source_path"}: score += 100
    if "base" in t: score += 50
    # file / path / source are strong signals for this column
    if "file" in t: score += 45
    if "path" in t: score += 40
    if "source" in t: score += 30
    if "name" in t: score += 10
    # penalise clearly wrong columns
    if "function" in t or "func" in t: score -= 25
    if t in {"batch","no","s.no","sno","sn","sr","serial","number","svn","sl"}: score -= 50
    if "s.no" in t or "serial" in t or t in {"sn","sno"}: score -= 30
    return score

def detect_best_column_in_workbook(excel_path, kind):
    if not os.path.isfile(excel_path): return None
    scorer = score_header_for_function if kind == "function" else score_header_for_base
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    best = None
    try:
        for sn in wb.sheetnames:
            ws = wb[sn]
            for r in range(1, min(ws.max_row or 1, 25)+1):
                for c in range(1, min(ws.max_column or 1, 60)+1):
                    value = clean_text(ws.cell(r,c).value)
                    if not value: continue
                    score = scorer(value)
                    if score <= 0: continue
                    candidate = {"sheet":sn,"row":r,"col_index":c,
                                 "col_letter":get_column_letter(c),
                                 "ref":f"{get_column_letter(c)}{r}",
                                 "header":value,"score":score}
                    if best is None or candidate["score"] > best["score"]: best = candidate
    finally: wb.close()
    return best

# ── Source scanning ───────────────────────────────────────────────────────────
def scan_source_for_all_functions(root_folder):
    file_entries = list(iter_source_files(root_folder))
    if not file_entries:
        _log.info("scan_source_for_all_functions: no source files found in %s", root_folder)
        return OrderedDict()
    _log.info("scan_source_for_all_functions: scanning %d files in %s",
              len(file_entries), root_folder)
    max_workers = min(8, max(2, (os.cpu_count() or 4)))

    def _scan_one(entry):
        full_path, file_name = entry
        try:
            fns = detect_functions_in_file(full_path)
        except Exception:
            fns = []
        return full_path, {"display_name": file_name, "functions": fns}

    results = {}
    # Per-file timeout (seconds) — skip files that hang on regex
    FILE_TIMEOUT = 30

    if len(file_entries) >= 24:
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {executor.submit(_scan_one, entry): entry[0] for entry in file_entries}
            for future in as_completed(futures):
                fp = futures[future]
                try:
                    full_path, info = future.result(timeout=FILE_TIMEOUT)
                    results[full_path] = info
                    log_function_extraction(full_path, info["functions"])
                except Exception:
                    # Timed out or errored — skip this file and continue
                    results[fp] = {"display_name": os.path.basename(fp), "functions": []}
                    _log.warning("scan_source: skipped (timeout/error) %s", os.path.basename(fp))
    else:
        for entry in file_entries:
            try:
                full_path, info = _scan_one(entry)
                results[full_path] = info
                log_function_extraction(full_path, info["functions"])
            except Exception:
                results[entry[0]] = {"display_name": entry[1], "functions": []}
                _log.warning("scan_source: error scanning %s", entry[1])

    total_fns = sum(len(v["functions"]) for v in results.values())
    _log.info("scan_source_for_all_functions: done — %d files, %d functions total",
              len(results), total_fns)

    ordered = OrderedDict()
    for full_path, _ in sorted(file_entries, key=lambda x: x[0].lower()):
        if full_path in results: ordered[full_path] = results[full_path]
    return ordered

def find_function_in_folder(folder, function_name):
    matches = []
    for root, _, files in os.walk(folder):
        for file in files:
            if not file.endswith((".c",".cpp",".h",".txt")): continue
            full_path = os.path.join(root, file)
            try:
                content = __import__("core.utils", fromlist=["read_source_file"]).read_source_file(full_path)
            except: continue
            if re.compile(rf"\b{function_name}\s*\([^;]*\)\s*\{{").search(content):
                matches.append(full_path)
    return matches

# ── Function list parsing ─────────────────────────────────────────────────────
_EXCEL_FALSE = {
    "if","else","for","while","do","switch","case","default","break","continue",
    "return","goto","sizeof","typeof","typedef","struct","union","enum","class",
    "static","extern","inline","const","volatile","register","auto","unsigned",
    "signed","long","short","void","int","char","float","double","bool","true",
    "false","null","none","nan","function","name","func","method","procedure",
    "routine","s.no","sno","sr","no","sl","serial","number","base","module",
    "file","filename","source","target","yes","ok","na","n/a",
}

def _valid_excel_fn(raw):
    if raw is None: return ""
    s = str(raw).strip()
    if not s or re.fullmatch(r'[\d\.\-\+eE]+',s) or ' ' in s: return ""
    if re.search(r'[^A-Za-z0-9_]',s) or not re.match(r'^[A-Za-z_]',s): return ""
    if len(s) > 128 or len(s) < 2 or s.lower() in _EXCEL_FALSE: return ""
    return s if is_probable_function_name(s) else ""

def _extract_fn_name(raw):
    s = clean_text(raw)
    if not s: return ""
    if " - " in s: s = s.rsplit(" - ",1)[-1].strip()
    s = s.replace("\\","/").strip().strip('"').strip("'")
    m = re.search(r"\.c_([A-Za-z_][A-Za-z0-9_]*)$",s,re.IGNORECASE)
    if m:
        fn = m.group(1).strip(); return fn if is_probable_function_name(fn) else ""
    m = re.search(r"\.c(?:\s*[-:#]\s*|\s+)([A-Za-z_][A-Za-z0-9_]*)$",s,re.IGNORECASE)
    if m:
        fn = m.group(1).strip(); return fn if is_probable_function_name(fn) else ""
    for token in reversed(re.findall(r"[A-Za-z_][A-Za-z0-9_]*",s)):
        if is_probable_function_name(token): return token.strip()
    return ""

def parse_function_list_file(path):
    names = []
    ext = os.path.splitext(path)[1].lower()
    _log.debug("parse_function_list_file: reading %s (ext=%s)", os.path.basename(path), ext)
    try:
        if ext == ".txt":
            with open(path,"r",encoding="utf-8",errors="ignore") as f:
                for line in f:
                    fn = _extract_fn_name(line)
                    if fn: names.append(fn)
        elif ext == ".xlsx":
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                for sn in wb.sheetnames:
                    for row in wb[sn].iter_rows(values_only=True):
                        for cell in row:
                            fn = _valid_excel_fn(cell)
                            if fn: names.append(fn)
            finally: wb.close()
    except: return []
    out, seen = [], set()
    for name in names:
        key = normalize_name(name)
        if key and key not in seen: seen.add(key); out.append(clean_text(name))
    _log.info("parse_function_list_file: %s → %d unique function names",
              os.path.basename(path), len(out))
    return out

def parse_function_list_files(file_paths):
    out, seen = [], set()
    for fp in file_paths:
        for name in parse_function_list_file(fp):
            key = normalize_name(name)
            if key and key not in seen: seen.add(key); out.append(name)
    return out

# ── Match helpers ─────────────────────────────────────────────────────────────
def match_target_with_function_list(target_root, function_list_paths):
    requested = {normalize_name(x) for x in parse_function_list_files(function_list_paths)}
    if not requested: return OrderedDict()
    target_records = scan_source_for_all_functions(target_root)
    matched = OrderedDict()
    for file_path, info in target_records.items():
        fns = [fn for fn in info["functions"] if normalize_name(fn) in requested]
        if fns: matched[file_path] = {"display_name": info["display_name"], "functions": fns}
    return matched

def match_target_with_reference_bases(target_root, reference_folders):
    if not reference_folders: return OrderedDict()
    target_records = scan_source_for_all_functions(target_root)
    ref_index = {}
    for folder in reference_folders:
        for _, info in scan_source_for_all_functions(folder).items():
            ref_index.setdefault(info["display_name"], set()).update(info["functions"])
    matched = OrderedDict()
    for file_path, info in target_records.items():
        common = [fn for fn in info["functions"] if fn in ref_index.get(info["display_name"],set())]
        if common: matched[file_path] = {"display_name": info["display_name"], "functions": common}
    return matched

def merge_record_sets(*record_sets):
    merged = OrderedDict()
    for rs in record_sets:
        for file_path, info in rs.items():
            if file_path not in merged:
                merged[file_path] = {"display_name": info["display_name"], "functions": []}
            for fn in info["functions"]:
                if fn not in merged[file_path]["functions"]: merged[file_path]["functions"].append(fn)
    return merged

# ── Consolidated DB helpers ───────────────────────────────────────────────────
def read_consolidated_matching_rows(excel_path, func_ref, preferred_sheet=None,
                                    base_col_ref=None):
    """
    Read every data row from the consolidated DB excel.
    Returns (headers, rows, func_col_idx) where:
      - headers: list of header cell values (row at header_row)
      - rows: list of full row tuples (values_only) below the header row
      - func_col_idx: 0-based index of the function-name column

    func_ref     - cell ref for the Function Name column (may come from
                   function-list file, e.g. "C1").
    base_col_ref - cell ref for the Base/File Name column detected FROM the
                   consolidated DB (e.g. "E11").  When provided, the header
                   row is taken from whichever ref has the larger row number,
                   so both columns land on the real header row in the DB.
    """
    if not os.path.isfile(excel_path):
        raise FileNotFoundError("Consolidated DB Excel not found.")

    # Determine header row: use the larger row number between the two refs.
    # base_col_ref is always detected from the consolidated DB itself, so its
    # row is the reliable anchor when func_ref came from a different file.
    func_row   = extract_excel_row_number(func_ref)
    base_row   = extract_excel_row_number(base_col_ref) if base_col_ref else 0
    header_row = max(func_row, base_row)
    data_start = header_row + 1

    # func_col_idx: when func_ref was sourced from a different file (its row <
    # header_row), its column letter no longer maps to the right column in the
    # consolidated DB.  Re-derive the index from base_col_ref's column letter
    # when that ref is the authoritative one, OR scan the header row to find
    # whichever cell best matches a "function name" header.
    if base_row >= func_row and base_col_ref:
        # base_col_ref came directly from the DB — use its column as the
        # function-name column (user pointed "Function Name Column" to the DB).
        # If the user set func_ref from the DB too (func_row == base_row), use
        # func_ref's column letter; otherwise fall back to base_col_ref column.
        if func_row == base_row:
            func_col_idx = column_index_from_string(extract_excel_column_letters(func_ref)) - 1
        else:
            # func_ref came from function-list file — its column letter is
            # meaningless in the DB.  Derive from base_col_ref column for now;
            # we will refine by scanning the header row below.
            func_col_idx = column_index_from_string(extract_excel_column_letters(base_col_ref)) - 1
    else:
        func_col_idx = column_index_from_string(extract_excel_column_letters(func_ref)) - 1

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        sheet_names = wb.sheetnames
        if preferred_sheet and preferred_sheet in sheet_names:
            sheet_names = [preferred_sheet] + [s for s in sheet_names if s != preferred_sheet]

        # Use the first sheet that has data at header_row
        ws = None
        for sn in sheet_names:
            candidate = wb[sn]
            # check that the header row exists
            rows_so_far = list(candidate.iter_rows(
                min_row=header_row, max_row=header_row, values_only=True))
            if rows_so_far and any(c is not None for c in rows_so_far[0]):
                ws = candidate
                chosen_sheet = sn
                break
        if ws is None:
            raise ValueError("Could not find a valid header row in the workbook.")

        # Read headers
        hdr_rows = list(ws.iter_rows(
            min_row=header_row, max_row=header_row, values_only=True))
        headers = [clean_text(c) for c in hdr_rows[0]] if hdr_rows else []

        # Refine func_col_idx by scanning the actual DB header row.
        # When func_ref came from a different file its column letter is
        # unreliable, so find the best "function name" match in the headers.
        if base_row > func_row:
            best_fn_score = -1
            for i, hdr_text in enumerate(headers):
                s = score_header_for_function(hdr_text)
                if s > best_fn_score:
                    best_fn_score = s
                    func_col_idx = i

        # Read all data rows
        all_rows = []
        for row in ws.iter_rows(min_row=data_start, values_only=True):
            all_rows.append(list(row))
    finally:
        wb.close()

    return headers, all_rows, func_col_idx


def create_output_filtered_rows_excel(source_excel_path, headers, matched_rows,
                                       unmatched_names,
                                       parsed_functions=None, matched_functions=None):
    """
    Write a single-sheet Excel with header row + every matched DB row.
    No summary sheet, no unmatched sheet — just the data.
    """
    src_dir  = os.path.dirname(source_excel_path)
    out_dir  = os.path.join(src_dir, "FuncAtlas_Output")
    os.makedirs(out_dir, exist_ok=True)
    out_file = os.path.join(out_dir, "FuncAtlas_Consolidated_Output.xlsx")
    if os.path.exists(out_file):
        try:
            os.remove(out_file)
        except PermissionError:
            raise PermissionError(
                f"Cannot overwrite output file because it is open:\n{out_file}\n\n"
                "Close Excel and try again.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Output"

    hdr_fill  = PatternFill("solid", fgColor="1F4E78")
    hdr_font  = Font(color="FFFFFF", bold=True, name="Arial", size=10)
    thin      = Side(style="thin", color="C5D8EC")
    thick     = Side(style="medium", color="1A3A5C")
    hdr_bdr   = Border(left=thick, right=thick, top=thick, bottom=thick)
    data_bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr       = Alignment(horizontal="center", vertical="center")
    left_al   = Alignment(horizontal="left",   vertical="center")

    # Strip fully-empty leading/trailing header columns
    # so the output does not have blank columns on the left
    first_col = 0
    for i, h in enumerate(headers):
        if h:
            first_col = i
            break
    trimmed_headers = headers[first_col:]

    # Write header row
    for c, h in enumerate(trimmed_headers, 1):
        cell = ws.cell(1, c, h if h else "")
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.border    = hdr_bdr
        cell.alignment = ctr
    ws.row_dimensions[1].height = 28

    # Write data rows (trim same leading columns)
    data_font = Font(name="Arial", size=10)
    for r, row_data in enumerate(matched_rows, 2):
        trimmed = list(row_data)[first_col:]
        for c, val in enumerate(trimmed, 1):
            cell = ws.cell(r, c, val if val is not None else "")
            cell.font      = data_font
            cell.border    = data_bdr
            cell.alignment = left_al
        ws.row_dimensions[r].height = 18

    # Auto-fit column widths (cap at 60)
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    ws.freeze_panes = "A2"
    wb.save(out_file)
    wb.close()
    return out_file


# keep old names as aliases so nothing else breaks
def read_consolidated_matches(excel_path, func_ref, base_ref, selected_functions, preferred_sheet=None):
    """Legacy alias — not used by ConsolidatedWorker any more."""
    return {}

def make_unique_base_assignment(function_order, matched_map):
    return {fn: matched_map.get(fn, []) for fn in function_order}

def create_output_matrix_excel(source_excel_path, unique_map,
                                parsed_functions=None, matched_functions=None, unmatched_functions=None):
    """Legacy alias — not used by ConsolidatedWorker any more."""
    return source_excel_path


# ── Folder → temp Excel extractor ────────────────────────────────────────────
def extract_functions_from_folder_to_excel(root_folder: str) -> str:
    """Walk *root_folder* recursively, detect every function in every source
    file, and write a temp Excel with columns:
        A: File Path | B: File Name | C: Function Name

    Returns the path of the generated temp .xlsx file.
    The caller is responsible for auto-detecting the function-name column
    (always column C, row 1 header = "Function Name").
    """
    import tempfile

    _log.info("extract_functions_from_folder_to_excel: scanning folder %s", root_folder)
    log_file_upload("folder", root_folder, field="Folder Extraction Source")

    records = []
    scan_result = scan_source_for_all_functions(root_folder)
    for full_path, info in scan_result.items():
        rel_path = os.path.relpath(full_path, root_folder)
        file_name = info["display_name"]
        for fn in info["functions"]:
            records.append((rel_path, file_name, fn))

    _log.info("extract_functions_from_folder_to_excel: %d function records from %d files",
              len(records), len(scan_result))

    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Functions"

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1565C0")
    headers = ["File Path", "File Name", "Function Name"]
    for col_idx, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, (rel_path, file_name, fn_name) in enumerate(records, start=2):
        ws.cell(row=row_idx, column=1, value=rel_path)
        ws.cell(row=row_idx, column=2, value=file_name)
        ws.cell(row=row_idx, column=3, value=fn_name)

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    tmp_dir = tempfile.gettempdir()
    out_path = os.path.join(tmp_dir, "funcatlas_extracted_functions.xlsx")
    wb.save(out_path)
    log_output_file(out_path, kind="Extracted Functions Excel (temp)")
    _log.info("extract_functions_from_folder_to_excel: saved %d records → %s",
              len(records), out_path)
    return out_path, len(records)


# ── Workers ───────────────────────────────────────────────────────────────────
class ConsolidatedWorker(QObject):
    finished = Signal(dict)
    error    = Signal(str)

    def __init__(self, function_list_files, consolidated_excel,
                 func_col_ref, base_col_ref, preferred_sheet=None):
        super().__init__()
        self.function_list_files = function_list_files
        self.consolidated_excel  = consolidated_excel
        self.func_col_ref        = func_col_ref
        self.base_col_ref        = base_col_ref          # kept for API compat
        self.preferred_sheet     = preferred_sheet

    def run(self):
        try:
            # 1. Parse function names from all function-list excel files
            parsed = parse_function_list_files(self.function_list_files)
            if not parsed:
                self.error.emit("No valid function names found in Function List files.")
                return

            wanted = {normalize_name(fn): fn for fn in parsed}

            # 2. Read full rows from consolidated DB, keep only matching ones
            headers, all_rows, func_col_idx = read_consolidated_matching_rows(
                self.consolidated_excel, self.func_col_ref, self.preferred_sheet,
                base_col_ref=self.base_col_ref)

            matched_rows     = []
            matched_fn_names = set()
            for row in all_rows:
                if func_col_idx >= len(row):
                    continue
                fn_val = clean_text(row[func_col_idx])
                if not fn_val:
                    continue
                norm = normalize_name(fn_val)
                if norm in wanted:
                    matched_rows.append(row)
                    matched_fn_names.add(norm)

            matched_functions  = [fn for fn in parsed if normalize_name(fn) in matched_fn_names]
            unmatched_names    = [fn for fn in parsed if normalize_name(fn) not in matched_fn_names]

            # 3. Write output
            out_file = create_output_filtered_rows_excel(
                self.consolidated_excel, headers, matched_rows, unmatched_names,
                parsed_functions=parsed, matched_functions=matched_functions)

            self.finished.emit({
                "output_file":          out_file,
                "function_list_count":  len(self.function_list_files),
                "functions_read":       len(parsed),
                "matched_count":        len(matched_functions),
                "unmatched_count":      len(unmatched_names),
            })
        except Exception as e:
            self.error.emit(str(e))


class BuiltinExtractionWorker(QObject):
    base_started  = Signal(str)
    base_progress = Signal(str, int, str)
    step_done     = Signal(str, int)
    log           = Signal(str)
    finished      = Signal(dict)
    error         = Signal(str)

    def __init__(self, bases, output_root, function_filter=None):
        """
        bases           – list of {label, src_path, is_target (bool)}
        output_root     – folder to write FuncAtlas_Extracted/ into
        function_filter – set/list of function names from the function list file.
                          When provided, the TARGET base is filtered to only
                          extract those functions. Reference bases always extract
                          all functions (so matches can be found).
        """
        super().__init__()
        self.bases           = bases
        self.output_root     = output_root
        self._cancel_requested = False
        # Normalised lowercase set for fast lookup; empty = no filter
        self.function_filter = (
            {normalize_name(fn) for fn in function_filter if fn}
            if function_filter else set()
        )

    def _safe_name(self, label):
        return re.sub(r'[^A-Za-z0-9_. -]+','_',clean_text(label)).strip().replace(' ','_') or 'base'

    def run(self):
        try:
            if not self.bases: self.error.emit('No bases provided.'); return
            os.makedirs(self.output_root, exist_ok=True)
            import tempfile as _tempfile
            root_out = os.path.join(_tempfile.gettempdir(), 'FuncAtlas_Extracted')
            os.makedirs(root_out, exist_ok=True)
            self.log.emit(f'Extracted .txt files → {root_out}')
            all_results = {}
            for base_idx, base in enumerate(self.bases, 1):
                if self._cancel_requested:
                    self.error.emit('__CANCELLED__')
                    return
                label    = clean_text(base.get('label'))
                src_path = normalize_path(base.get('src_path',''))
                is_target = base.get('is_target', False)

                if not os.path.isdir(src_path):
                    self.error.emit(f'Source folder not found for {label}:\n{src_path}'); return
                self.base_started.emit(label)
                self.log.emit(f'[{base_idx}/{len(self.bases)}] Starting: {label}')

                # ── Use upfront cache when available (set at Submit time) ────
                role = 'target' if is_target else 'reference'
                cached_meta = FUNCTION_CACHE.get_meta(src_path, role)
                if cached_meta is not None:
                    records = OrderedDict(cached_meta)
                    self.log.emit(f'  → Using pre-cached scan for {label}')
                else:
                    records = scan_source_for_all_functions(src_path)
                total_files = max(1, len(records))
                base_out    = os.path.join(root_out, self._safe_name(label))
                if os.path.isdir(base_out):
                    for name in os.listdir(base_out):
                        try: os.remove(os.path.join(base_out, name))
                        except: pass
                os.makedirs(base_out, exist_ok=True)

                # Apply function filter ONLY for the target base when a list was uploaded
                apply_filter = is_target and bool(self.function_filter)
                if apply_filter:
                    self.log.emit(
                        f'  → Function filter active: extracting only '
                        f'{len(self.function_filter)} listed functions from target.'
                    )

                extracted, seen_pairs = 0, set()
                index_data = {}

                for file_idx, (file_path, info) in enumerate(records.items(), 1):
                    if self._cancel_requested:
                        self.error.emit('__CANCELLED__')
                        return
                    pct = int((file_idx / total_files) * 100)
                    self.base_progress.emit(label, pct, f'Scanning {file_idx}/{total_files} files')

                    for fn in info.get('functions', []):
                        key = normalize_name(fn)
                        if not key:
                            continue
                        # Deduplicate by (function_name, file_path) — same name in
                        # different files are distinct functions and must both appear
                        pair_key = (key, normalize_path(file_path))
                        if pair_key in seen_pairs:
                            continue
                        # Filter: skip functions NOT in the list (target only)
                        if apply_filter and key not in self.function_filter:
                            continue
                        body = (FUNCTION_CACHE.get_body(src_path, role, file_path, fn)
                                or extract_function_body(file_path, fn))
                        # .txt filename: encode full filepath + function name so the
                        # filename itself is human-readable and unique per (fn, file).
                        # e.g. a/b.c -> add  =>  a__b.c__add.txt
                        #      e/b.c -> add  =>  e__b.c__add.txt
                        norm_fp   = normalize_path(file_path)
                        # Replace path separators with __ so the name is a flat file
                        safe_path = re.sub(r'[/\\]', '__', norm_fp).strip('_')
                        # Strip characters that are illegal in filenames
                        safe_path = re.sub(r'[<>:"|?*]', '_', safe_path)
                        txt_name  = f'{safe_path}__{fn}.txt'
                        # index key must be unique per (fn, file) pair
                        index_key = f'{fn.lower()}|{norm_fp}'
                        try:
                            with open(os.path.join(base_out, txt_name), 'w',
                                      encoding='utf-8', errors='ignore') as fh:
                                fh.write(body)
                            # Record in index ONLY when .txt is successfully written
                            index_data[index_key] = {"display_name": fn, "source_file": file_path,
                                                      "txt_name": txt_name}
                            seen_pairs.add(pair_key)
                            extracted += 1
                        except Exception as e:
                            self.log.emit(f'Could not write {txt_name}: {e}')


                try:
                    with open(os.path.join(base_out, '_index.json'), 'w', encoding='utf-8') as fh:
                        _json.dump(index_data, fh, indent=2)
                except Exception as e:
                    self.log.emit(f'Warning: index write failed: {e}')

                self.log.emit(f'[{base_idx}/{len(self.bases)}] Done: {label} -> {extracted} functions')
                self.step_done.emit(label, extracted)
                all_results[label] = base_out

            self.finished.emit(all_results)
        except Exception as e:
            self.error.emit(str(e))