"""
services/report_worker.py
─────────────────────────
ReportCompareWorker – compares extracted function .txt files and writes the
Excel match report (FuncAtlas_Report .xlsx).

Sheets produced
───────────────
  Sheet 1  "Function_Match_Report"   — per-function match % per reference base
                                        + Which Reference Base + Reference base file path
  Sheet 2  "Summary"                 — aggregate counts (Reuse / New / Modified)
  Sheet 3  "Complexity_Compatibility"— per-function construct counts, weighted
                                        score and complexity level (uses the
                                        same weights/bands the user configured
                                        in the Complexity & Compatibility page)
"""
import os
import re
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
import json as _json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from PySide6.QtCore import QObject, Signal

from core.utils import read_source_file,\
     normalize_path, normalize_name

# ── Complexity helpers (shared with complexity_worker) ────────────────────────
CONSTRUCTS = [
    ("If...Else",           r'\bif\s*\('),
    ("If...Else if...Else", r'\belse\s+if\s*\('),
    ("Nested If",           r'\bif\s*\([^)]*\)\s*\{[^}]*\bif\s*\('),
    ("Switch",              r'\bswitch\s*\('),
    ("For",                 r'\bfor\s*\('),
    ("While",               r'\bwhile\s*\('),
    ("Do...While",          r'\bdo\s*\{'),
    ("Return",              r'\breturn\b'),
    ("Function Call",       r'\b[A-Za-z_]\w*\s*\('),
    ("Pointers",            r'\*[A-Za-z_]\w*|\b[A-Za-z_]\w*\s*\*'),
    ("Struct",              r'\bstruct\s+\w+'),
    ("Assign",              r'(?<![=!<>])=(?!=)'),
]

DEFAULT_WEIGHTS = {n: 1 for n, _ in CONSTRUCTS}
DEFAULT_BANDS = [
    ("Low",       1,   5),
    ("Medium",    6,  12),
    ("High",     13,  25),
    ("Very High", 26,  40),
    ("Complex",  41, 999),
]


def _strip_comments(text):
    text = re.sub(r'/\*.*?\*/', ' ', text, flags=re.DOTALL)
    text = re.sub(r'//[^\n]*', ' ', text)
    return text


def _count_constructs(body):
    clean = _strip_comments(body)
    return {name: len(re.findall(pat, clean)) for name, pat in CONSTRUCTS}


def _complexity_level(score, bands):
    for label, start, end in bands:
        if start <= score <= end:
            return label
    return bands[-1][0] if bands else "Unknown"


class ReportCompareWorker(QObject):
    progress = Signal(int, str)
    log      = Signal(str)
    finished = Signal(str)
    error    = Signal(str)

    def __init__(self, target_label, target_folder, ref_labels, ref_folders, output_root,
                 target_src_path="", ref_src_paths=None,
                 weights=None, bands=None):
        super().__init__()
        self.target_label    = target_label
        self.target_folder   = normalize_path(target_folder)
        self.ref_labels      = ref_labels
        self.ref_folders     = [normalize_path(x) for x in ref_folders]
        self.output_root     = normalize_path(output_root)
        self.target_src_path = target_src_path
        self.ref_src_paths   = ref_src_paths or []
        # Complexity settings from Complexity & Compatibility page
        self.weights = {n: w for n, w in (weights or [])} if weights else dict(DEFAULT_WEIGHTS)
        self.bands   = bands or list(DEFAULT_BANDS)

    # ── helpers ───────────────────────────────────────────────────────────────
    def _read_text(self, path):
        from core.utils import read_source_file
        return read_source_file(path)

    def _match_percent(self, a, b):
        def _clean(text):
            out = []
            for raw in text.splitlines():
                s = raw.strip()
                if not s or s.startswith('//') or s.startswith('/*') or s.startswith('*'):
                    continue
                out.append(s)
            return out
        lines_a = _clean(a); lines_b = _clean(b)
        if not lines_a or not lines_b:
            return 0
        cnt_a = Counter(lines_a); cnt_b = Counter(lines_b)
        matched = sum(min(cnt_a[ln], cnt_b[ln]) for ln in cnt_a)
        total   = max(len(lines_a), len(lines_b))
        return int(round((matched / total) * 100))

    def _classify_reuse_status(self, ref_data):
        valid = [p for (p, *_) in ref_data if p is not None]
        if not valid:
            return 'New'
        best = max(valid)
        if best < 60:
            return 'New'
        if best == 100:
            return 'Reuse'
        return 'Reuse (Modified)'   # 60–99

    def _best_ref(self, ref_data):
        """Return (which_ref_label, ref_file_path) for the highest-match reference.
        Returns em-dash only when there is genuinely no match (all pct are None).
        Even a 1% match should show the reference name/path.
        """
        best_pct   = -1
        best_label = u'\u2014'   # em dash
        best_path  = u'\u2014'
        for lbl, (pct, _rfname, r_file_path) in zip(self.ref_labels, ref_data):
            if pct is not None and pct > best_pct:
                best_pct   = pct
                best_label = lbl
                best_path  = r_file_path if r_file_path and r_file_path != u'\u2014' else u'\u2014'
        if best_pct < 0:
            # No match found at all — match % column also shows —
            return u'\u2014', u'\u2014'
        return best_label, best_path

    def _trim_path(self, full_path: str) -> str:
        """Trim an absolute path to start from the target folder name.
        e.g. D:\Projects\GUI2C\...\1BB00650_Target\src\BIOS\... 
          →  1BB00650_Target\src\BIOS\...
        Falls back to the full path if no known anchor is found.
        """
        if not full_path or full_path == '\u2014':
            return full_path
        # Normalise separators
        norm = os.path.normpath(full_path)
        parts = norm.split(os.sep)
        # Find the target label segment — try self.target_label first
        tgt = (self.target_label or "").strip()
        for i, p in enumerate(parts):
            if p == tgt:
                return os.path.join(*parts[i:])
        # Fallback: find a part that contains common anchor keywords
        for i, p in enumerate(parts):
            if p.lower().startswith(("1bb", "target", "ref", "base")):
                return os.path.join(*parts[i:])
        return full_path

    def _count_loc(self, body_text: str) -> int:
        """Count non-empty, non-comment lines in a function body."""
        if not body_text:
            return 0
        lines = body_text.splitlines()
        count = 0
        in_block = False
        for line in lines:
            s = line.strip()
            if in_block:
                if '*/' in s:
                    in_block = False
                continue
            if s.startswith('/*'):
                if '*/' not in s[2:]:
                    in_block = True
                continue
            if s.startswith('//') or not s:
                continue
            count += 1
        return count

    # ── Excel writer ──────────────────────────────────────────────────────────
    def _write_excel(self, rows):
        """
        rows: list of (func_name, t_file_name, t_file_path, ref_data, body_text)
          ref_data : [(pct_or_None, r_file_name, r_file_path), ...]  one per ref
          body_text: raw extracted function body text (used for Sheet 3)
        """
        os.makedirs(self.output_root, exist_ok=True)
        out_path = os.path.join(self.output_root, 'FuncAtlas_Report.xlsx')
        wb = Workbook()

        # ── shared style helpers ──────────────────────────────────────────────
        thin  = Side(style='thin',   color='C5D8EC')
        thick = Side(style='medium', color='1A3A5C')

        def bdr(left=None, right=None, top=None, bottom=None):
            return Border(left=left or thin, right=right or thin,
                          top=top or thin,   bottom=bottom or thin)

        hdr_fill   = PatternFill('solid', fgColor='1F4E78')
        hdr_font   = Font(color='FFFFFF', bold=True, name='Arial', size=10)
        hdr_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
        pct_align  = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left',   vertical='center')
        sno_fill   = PatternFill('solid', fgColor='EBF3FB')

        green_fill  = PatternFill('solid', fgColor='C6EFCE')
        green_font  = lambda: Font(color='276221', name='Arial', size=10, bold=True)
        yellow_fill = PatternFill('solid', fgColor='FFEB9C')
        yellow_font = lambda: Font(color='9C5700', name='Arial', size=10, bold=True)
        red_fill    = PatternFill('solid', fgColor='FFC7CE')
        red_font    = lambda: Font(color='9C0006', name='Arial', size=10, bold=True)
        blue_fill   = PatternFill('solid', fgColor='BDD7EE')
        data_font   = lambda: Font(name='Arial', size=10)

        # ═══════════════════════════════════════════════════════════════════
        # SHEET 1 — Function_Match_Report
        # S.No | Function Name | Target File Name | Target File Path
        # | Reference Base - {lbl} Match %  (one col per ref)
        # | Reuse/New | Which Reference Base | Reference base file path
        # ═══════════════════════════════════════════════════════════════════
        ws = wb.active
        ws.title = 'Function_Match_Report'

        headers = ['File Name', 'Function Name', 'Target File Path', 'Total LOC']
        for lbl in self.ref_labels:
            headers.append(f'{lbl}\nMatch %')
        headers += ['Reuse/New', 'Suggested Reference Base', 'Reference base file path']

        last_col = len(headers)
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = hdr_align
            cell.border = bdr(left=thick if c == 1 else thin,
                              right=thick if c == last_col else thin,
                              top=thick, bottom=thick)
        ws.row_dimensions[1].height = 44

        for r, row_data in enumerate(rows, 2):
            func_name, t_file_name, t_file_path, ref_data, _body = row_data
            col = 1

            # File Name (first column)
            c = ws.cell(r, col, t_file_name); col += 1
            c.fill = sno_fill; c.font = Font(name='Arial', size=10, color='2C5F8A')
            c.alignment = left_align; c.border = bdr(left=thick)

            # Function Name
            c = ws.cell(r, col, func_name); col += 1
            c.font = Font(bold=True, name='Arial', size=10)
            c.alignment = left_align; c.border = bdr()

            # Target File Path (trimmed)
            trimmed_path = self._trim_path(t_file_path)
            c = ws.cell(r, col, trimmed_path); col += 1
            c.font = data_font(); c.alignment = left_align; c.border = bdr()

            # Total LOC (column D)
            loc = self._count_loc(_body)
            c = ws.cell(r, col, loc); col += 1
            c.font = Font(name='Arial', size=10, bold=True)
            c.alignment = pct_align; c.border = bdr()
            if loc > 0:
                c.fill = PatternFill('solid', fgColor='EBF3FB')

            # One Match % column per reference
            for pct, _rfname, _rfpath in ref_data:
                display = u'\u2014' if pct is None else f'{pct}%'
                c = ws.cell(r, col, display); col += 1
                c.alignment = pct_align; c.border = bdr()
                if pct is not None:
                    if pct >= 80:  c.fill = green_fill;  c.font = green_font()
                    elif pct > 0:  c.fill = yellow_fill; c.font = yellow_font()
                    else:          c.fill = red_fill;    c.font = red_font()

            # Reuse/New
            status = self._classify_reuse_status(ref_data)
            c = ws.cell(r, col, status); col += 1
            c.alignment = pct_align; c.border = bdr()
            if   status == 'Reuse':
                c.fill = green_fill;  c.font = green_font()
            elif status == 'Reuse (Modified)':
                c.fill = yellow_fill; c.font = yellow_font()
            elif status == 'New':
                c.fill = blue_fill;   c.font = Font(color='1F1F1F', name='Arial', size=10, bold=True)

            # Which Reference Base
            which_ref, ref_base_path = self._best_ref(ref_data)
            c = ws.cell(r, col, which_ref); col += 1
            c.font = Font(bold=True, name='Arial', size=10)
            c.alignment = pct_align; c.border = bdr()
            if which_ref != u'\u2014':
                c.fill = PatternFill('solid', fgColor='DEEAF1')

            # Reference base file path
            c = ws.cell(r, col, ref_base_path); col += 1
            c.font = data_font(); c.alignment = left_align
            c.border = bdr(right=thick)

        # Column widths — Sheet 1
        ws.column_dimensions['A'].width = 28   # File Name
        ws.column_dimensions['B'].width = 38   # Function Name
        ws.column_dimensions['C'].width = 50   # Target File Path
        ws.column_dimensions['D'].width = 12   # Total LOC
        ci = 5
        for _ in self.ref_labels:
            ws.column_dimensions[get_column_letter(ci)].width = 24; ci += 1
        ws.column_dimensions[get_column_letter(ci)].width = 18; ci += 1   # Reuse/New
        ws.column_dimensions[get_column_letter(ci)].width = 32; ci += 1   # Which Ref Base
        ws.column_dimensions[get_column_letter(ci)].width = 55            # Ref base file path
        ws.freeze_panes = 'A2'

        # ═══════════════════════════════════════════════════════════════════
        # SHEET 2 — Summary
        # ═══════════════════════════════════════════════════════════════════
        ss = wb.create_sheet('Summary')
        ss['A1'] = 'Metric'; ss['B1'] = 'Value'
        hdr_fill2 = PatternFill('solid', fgColor='1F4E78')
        hdr_font2 = Font(color='FFFFFF', bold=True, name='Arial', size=10)
        for cell in ss[1]:
            cell.fill = hdr_fill2; cell.font = hdr_font2
        ss_data = [
            ('Target Base',       self.target_label),
            # ('Target Source',     self.target_src_path),
            ('Reference Bases',   ', '.join(self.ref_labels) if self.ref_labels else 'None'),
            # ('Reference Sources', ', '.join(self.ref_src_paths) if self.ref_src_paths else 'None'),
            ('Total Functions',   len(rows)),
            ('New',               sum(1 for rd in rows if self._classify_reuse_status(rd[3]) == 'New')),
            ('Reuse',             sum(1 for rd in rows if self._classify_reuse_status(rd[3]) == 'Reuse')),
            ('Reuse (Modified)',  sum(1 for rd in rows if self._classify_reuse_status(rd[3]) == 'Reuse (Modified)')),
        ]
        for i, (k, v) in enumerate(ss_data, 2):
            ss.cell(i, 1, k).font = Font(bold=True, name='Arial')
            ss.cell(i, 2, str(v))
        ss.column_dimensions['A'].width = 22
        ss.column_dimensions['B'].width = 60

        # ═══════════════════════════════════════════════════════════════════
        # ── save ─────────────────────────────────────────────────────────────
        try:
            wb.save(out_path)
            final_path = out_path
        except PermissionError:
            from datetime import datetime
            alt = os.path.join(self.output_root,
                               f"FuncAtlas_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            wb.save(alt)
            final_path = alt
            self.log.emit(f"Default file was locked, saved as:\n{alt}")
        finally:
            wb.close()
        return final_path

    # ── run ───────────────────────────────────────────────────────────────────
    def run(self):
        try:
            if not os.path.isdir(self.target_folder):
                self.error.emit(f'Target extraction folder not found:\n{self.target_folder}')
                return

            def _load_index(folder):
                try:
                    with open(os.path.join(folder, '_index.json'), 'r', encoding='utf-8') as fh:
                        return _json.load(fh)
                except Exception:
                    return {}

            target_index = _load_index(self.target_folder)
            ref_indexes  = [_load_index(rf) for rf in self.ref_folders]

            # Build target_files from index (supports fn__hash.txt naming)
            # index key format: 'fn_name|/full/file/path' or legacy 'fn_name'
            target_files = {}
            for idx_key, idx_val in target_index.items():
                txt_name = idx_val.get('txt_name') or (idx_val.get('display_name', '') + '.txt')
                txt_path = os.path.join(self.target_folder, txt_name)
                if os.path.isfile(txt_path):
                    target_files[idx_key] = (idx_val.get('display_name', idx_key), txt_path)
            if not target_files:
                self.error.emit(f'No extracted .txt functions found in:\n{self.target_folder}')
                return

            self.progress.emit(0, 'Preparing comparison …')
            self.log.emit(
                f'Comparing {len(target_files)} target functions against '
                f'{len(self.ref_folders)} reference bases'
            )

            # Build ref_file_maps from each ref index (supports fn__hash.txt naming)
            ref_file_maps = []
            for ri, ref_folder in enumerate(self.ref_folders):
                ref_map = {}  # idx_key -> txt_path
                for idx_key, idx_val in ref_indexes[ri].items():
                    txt_name = idx_val.get('txt_name') or (idx_val.get('display_name', '') + '.txt')
                    txt_path = os.path.join(ref_folder, txt_name)
                    if os.path.isfile(txt_path):
                        ref_map[idx_key] = txt_path
                ref_file_maps.append(ref_map)

            items       = sorted(target_files.items())
            rows        = [None] * len(items)
            total       = len(items)
            max_workers = min(8, max(2, (os.cpu_count() or 4)))

            def _compare_one(payload):
                idx, item = payload
                idx_key, (func_display, target_path) = item
                target_text = self._read_text(target_path)
                # idx_key is 'fn_name|/source/file/path' (or legacy 'fn_name')
                t_info      = target_index.get(idx_key, {})
                t_src_full  = t_info.get('source_file', '') or ''
                t_file_name = os.path.basename(t_src_full) or u'\u2014'
                t_file_path = t_src_full or u'\u2014'

                # For ref matching: look up by idx_key first (exact fn+file match).
                # If not found, scan ALL ref-index entries whose display_name matches
                # the function name (handles same-named funcs in different files) and
                # pick the candidate with the highest similarity score.
                fn_lower = func_display.lower()
                ref_data = []
                for ri, ref_map in enumerate(ref_file_maps):
                    # 1. Exact composite-key lookup
                    ref_path = ref_map.get(idx_key)
                    r_info   = ref_indexes[ri].get(idx_key)

                    # 2. Legacy name-only key (old index format)
                    if ref_path is None:
                        ref_path = ref_map.get(fn_lower)
                        r_info   = ref_indexes[ri].get(fn_lower)

                    # 3. Scan all entries for matching display_name (new-format
                    #    refs where same fn exists in multiple files) — best score wins
                    if ref_path is None:
                        best_pct_scan  = -1
                        best_path_scan = None
                        best_info_scan = None
                        for rk, rv in ref_indexes[ri].items():
                            if normalize_name(rv.get('display_name', '')) == normalize_name(func_display):
                                candidate_path = ref_map.get(rk)
                                if candidate_path and os.path.isfile(candidate_path):
                                    candidate_text = self._read_text(candidate_path)
                                    candidate_pct  = self._match_percent(target_text, candidate_text)
                                    if candidate_pct > best_pct_scan:
                                        best_pct_scan  = candidate_pct
                                        best_path_scan = candidate_path
                                        best_info_scan = rv
                        if best_path_scan is not None:
                            ref_path = best_path_scan
                            r_info   = best_info_scan

                    r_info      = r_info or {}
                    r_src_full  = r_info.get('source_file', '') or ''
                    r_file_name = os.path.basename(r_src_full) or u'\u2014'
                    r_file_path = r_src_full or u'\u2014'
                    if ref_path and os.path.isfile(ref_path):
                        ref_text = self._read_text(ref_path)
                        pct      = self._match_percent(target_text, ref_text)
                        ref_data.append((pct, r_file_name, r_file_path))
                    else:
                        ref_data.append((None,
                                         r_file_name if r_src_full else u'\u2014',
                                         r_file_path if r_src_full else u'\u2014'))

                # body_text passed through for Sheet 3 complexity analysis
                return idx, (func_display, t_file_name, t_file_path, ref_data, target_text)

            if total >= 24:
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = [executor.submit(_compare_one, (i, item))
                               for i, item in enumerate(items)]
                    done = 0
                    for future in as_completed(futures):
                        idx, row = future.result(); rows[idx] = row; done += 1
                        self.progress.emit(int((done / total) * 92), f'Comparing … {done}/{total}')
            else:
                for idx, item in enumerate(items):
                    _, row = _compare_one((idx, item)); rows[idx] = row
                    self.progress.emit(int(((idx + 1) / total) * 92),
                                       f'Comparing … {idx + 1}/{total}')

            rows = [r for r in rows if r is not None]
            if not rows:
                self.error.emit('No rows produced for report.')
                return

            self.progress.emit(96, 'Writing Excel report …')
            out_file = self._write_excel(rows)
            self.log.emit(f'Excel report written: {out_file}')
            self.progress.emit(100, 'Report ready')
            self.finished.emit(out_file)

        except Exception as e:
            import traceback
            self.error.emit(f"{e}\n{traceback.format_exc()}")