"""
services/complexity_worker.py
──────────────────────────────
ComplexityAnalysisWorker  – scans source files, saves function bodies as .txt,
builds an Excel report with:
  Sheet 1 "Function_Complexity" — per-function construct counts + weighted
           complexity score + level (Low / Medium / High / Very High / Complex)
  Sheet 2 "Construct_Summary"   — total count of every construct across all
           functions
"""

import os
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PySide6.QtCore import QObject, Signal

from core.logger import get_logger

from core.utils import (
    normalize_path, normalize_name, iter_source_files,
    detect_functions_in_file, extract_function_body,
)


# ── Construct patterns (C language) ──────────────────────────────────────────
CONSTRUCTS = [
    # ── Control Flow ──────────────────────────────────────────────────────────
    ("If Statement",        r'\bif\s*\('),
    ("Else Branch",         r'\belse\b(?!\s*if)'),
    ("Else-if Chain",       r'\belse\s+if\s*\('),
    ("Switch Statement",    r'\bswitch\s*\('),
    ("Case Label",          r'\bcase\s+'),
    ("Default Label",       r'\bdefault\s*:'),
    ("Break Statement",     r'\bbreak\s*;'),
    ("For Loop",            r'\bfor\s*\('),
    ("Return Statement",    r'\breturn\b'),
    ("Continue Statement",  r'\bcontinue\s*;'),
    ("While Loop",          r'\bwhile\s*\('),
    ("Do...While",          r'\bdo\s*\{'),
    # ── Function ──────────────────────────────────────────────────────────────
    ("Function Call",       r'\b[A-Za-z_]\w*\s*\('),
    ("Function Definition", r'\b[A-Za-z_]\w*\s+[A-Za-z_]\w*\s*\([^)]*\)\s*\{'),
    ("Function Declaration",r'\b[A-Za-z_]\w*\s+[A-Za-z_]\w*\s*\([^)]*\)\s*;'),
    # ── Pointer ───────────────────────────────────────────────────────────────
    ("Address-of Operator", r'&[A-Za-z_]\w*'),
    ("Pointer Declaration", r'\b[A-Za-z_]\w*\s*\*+\s*[A-Za-z_]\w*'),
    ("Pointer Dereference", r'\*[A-Za-z_]\w*'),
    ("Arrow Operator",      r'->\s*[A-Za-z_]\w*'),
    ("Void Pointer",        r'\bvoid\s*\*'),
    ("Pointer Arithmetic",  r'\b[A-Za-z_]\w*\s*[\+\-]\s*\d+|\b[A-Za-z_]\w*\s*\+\+|\+\+\s*[A-Za-z_]\w*'),
    # ── Array ─────────────────────────────────────────────────────────────────
    ("Array Access",        r'[A-Za-z_]\w*\s*\['),
    ("Array Declaration",   r'[A-Za-z_]\w*\s+[A-Za-z_]\w*\s*\[\d*\]'),
    ("String Literal",      r'"[^"]*"'),
    ("Char Array",          r'\bchar\s+[A-Za-z_]\w*\s*\['),
    # ── Type ──────────────────────────────────────────────────────────────────
    ("Cast Operation",      r'\(\s*[A-Za-z_]\w*\s*\*?\s*\)\s*[A-Za-z_\(]'),
    ("Typedef",             r'\btypedef\b'),
    ("Enum Definition",     r'\benum\s+\w*\s*\{'),
    ("Sizeof Operator",     r'\bsizeof\s*\('),
    ("Const Declaration",   r'\bconst\b'),
    ("Signed/Unsigned",     r'\b(?:signed|unsigned)\b'),
    ("Short/Long",          r'\b(?:short|long)\b'),
    # ── Storage ───────────────────────────────────────────────────────────────
    ("Static Keyword",      r'\bstatic\b'),
    # ── Preprocessor ─────────────────────────────────────────────────────────
    ("Macro Definition",    r'#\s*define\b'),
    ("Ifdef Directive",     r'#\s*ifdef\b'),
    ("If Directive",        r'#\s*if\b(?!def)'),
    ("Elif Directive",      r'#\s*elif\b'),
    ("Else Directive",      r'#\s*else\b'),
    ("Endif Directive",     r'#\s*endif\b'),
    ("Pragma Directive",    r'#\s*pragma\b'),
    # ── Operator ──────────────────────────────────────────────────────────────
    ("Compound Assignment", r'[A-Za-z_]\w*\s*(?:\+=|-=|\*=|/=|%=|&=|\|=|\^=|<<=|>>=)'),
    ("Increment",           r'\+\+'),
    ("Decrement",           r'--'),
    ("Bitwise AND",         r'(?<![&])&(?![&])'),
    ("Bitwise OR",          r'(?<!\|)\|(?!\|)'),
    ("Bitwise XOR",         r'\^'),
    ("Bitwise NOT",         r'~'),
    ("Left Shift",          r'<<'),
    ("Right Shift",         r'>>'),
    ("Logical AND",         r'&&'),
    ("Logical OR",          r'\|\|'),
    ("Logical NOT",         r'!(?!=)'),
    # ── Safety ────────────────────────────────────────────────────────────────
    ("NULL Check",          r'\bNULL\b.*(?:==|!=)|(?:==|!=).*\bNULL\b'),
    ("NULL Assignment",     r'=\s*NULL\b'),
    # ── Memory ────────────────────────────────────────────────────────────────
    ("memset",              r'\bmemset\s*\('),
    ("memcpy",              r'\bmemcpy\s*\('),
    # ── Math ──────────────────────────────────────────────────────────────────
    ("fabs",                r'\bfabs\s*\('),
    # ── Misc ──────────────────────────────────────────────────────────────────
    ("Designated Initializer", r'\.\s*[A-Za-z_]\w*\s*='),
    ("Compound Literal",    r'\(\s*[A-Za-z_]\w+\s*\)\s*\{'),
    ("Bit Field",           r':\s*\d+\s*;'),
]


# ── Default weights & bands (single source of truth) ─────────────────────────
DEFAULT_WEIGHTS = {
    "If Statement":           1, "Else Branch":            1,
    "Else-if Chain":          2, "Switch Statement":       2,
    "Case Label":             1, "Default Label":          1,
    "Break Statement":        1, "For Loop":               2,
    "Return Statement":       1, "Continue Statement":     1,
    "While Loop":             2, "Do...While":             2,
    "Function Call":          4, "Function Definition":    3,
    "Function Declaration":   2, "Address-of Operator":    3,
    "Pointer Declaration":    4, "Pointer Dereference":    4,
    "Arrow Operator":         3, "Void Pointer":           5,
    "Pointer Arithmetic":     5, "Array Access":           2,
    "Array Declaration":      2, "String Literal":         1,
    "Char Array":             2, "Cast Operation":         3,
    "Typedef":                2, "Enum Definition":        2,
    "Sizeof Operator":        2, "Const Declaration":      1,
    "Signed/Unsigned":        1, "Short/Long":             1,
    "Static Keyword":         2, "Compound Assignment":    1,
    "Increment":              1, "Decrement":              1,
    "Bitwise AND":            2, "Bitwise OR":             2,
    "Bitwise XOR":            2, "Bitwise NOT":            2,
    "Left Shift":             2, "Right Shift":            2,
    "Logical AND":            1, "Logical OR":             1,
    "Logical NOT":            1, "Designated Initializer": 2,
    "Compound Literal":       2, "Bit Field":              3,
    "Macro Definition":       2, "Ifdef Directive":        1,
    "If Directive":           1, "Elif Directive":         1,
    "Else Directive":         1, "Endif Directive":        1,
    "Pragma Directive":       1, "NULL Check":             3,
    "NULL Assignment":        2, "memset":                 3,
    "memcpy":                 3, "fabs":                   2,
}

DEFAULT_BANDS = [
    ("Low",       0,   5),
    ("Medium",    6,  12),
    ("High",      13,  25),
    ("Very High", 26,  40),
    ("Complex",   41, 999),
]


def _strip_comments(text: str) -> str:
    text = re.sub(r'/\*.*?\*/', ' ', text, flags=re.DOTALL)
    text = re.sub(r'//[^\n]*', ' ', text)
    return text


def count_constructs(body: str) -> dict:
    """Return {construct_name: count} for a function body string."""
    clean = _strip_comments(body)
    counts = {}
    for name, pattern in CONSTRUCTS:
        counts[name] = len(re.findall(pattern, clean))
    return counts


def complexity_level(score: float, bands: list) -> str:
    """Given a numeric score and band list [(label, start, end), ...], return level label."""
    for label, start, end in bands:
        if start <= score <= end:
            return label
    # above all bands
    if bands:
        return bands[-1][0]
    return "Unknown"


# ── Worker ────────────────────────────────────────────────────────────────────
class ComplexityAnalysisWorker(QObject):
    progress = Signal(int, str)   # percent, message
    log      = Signal(str)
    finished = Signal(str)        # path to generated Excel
    error    = Signal(str)

    def __init__(self, source_folder: str, output_root: str,
                 weights: list = None, bands: list = None):
        """
        source_folder : root folder to scan
        output_root   : where to create function_body/ and the Excel
        weights       : [(name, weight), ...]  — from ComplexitySettingsDialog
        bands         : [(label, start, end), ...] — from ComplexitySettingsDialog
        """
        super().__init__()
        self.source_folder = normalize_path(source_folder)
        self.output_root   = normalize_path(output_root)
        self.weights       = {n: w for n, w in (weights or [])} or DEFAULT_WEIGHTS
        self.bands         = bands or DEFAULT_BANDS

    # ── main entry ────────────────────────────────────────────────────────────
    def run(self):
        try:
            self._run()
        except Exception as exc:
            import traceback
            self.error.emit(f"{exc}\n{traceback.format_exc()}")

    def _run(self):
        body_dir = os.path.join(self.output_root, "function_body")
        os.makedirs(body_dir, exist_ok=True)
        self.log.emit(f"📁 Output folder: {self.output_root}")

        # Scan all source files
        file_entries = list(iter_source_files(self.source_folder))
        total_files  = len(file_entries)
        if total_files == 0:
            self.error.emit("No source files found in the selected folder.")
            return

        self.log.emit(f"🔍 Found {total_files} source files — scanning …")
        all_records = []   # list of dicts for Excel

        for file_idx, (full_path, file_name) in enumerate(file_entries):
            pct = int((file_idx / total_files) * 85)
            self.progress.emit(pct, f"Scanning {file_name} …")

            functions = detect_functions_in_file(full_path)
            if not functions:
                continue

            rel_path = os.path.join(os.path.basename(self.source_folder), os.path.relpath(full_path, self.source_folder))

            for fn_name in functions:
                body = extract_function_body(full_path, fn_name)

                # Save .txt
                safe_name = re.sub(r'[\\/:*?"<>|]', '_', f"{file_name}__{fn_name}")
                txt_path  = os.path.join(body_dir, f"{safe_name}.txt")
                try:
                    with open(txt_path, "w", encoding="utf-8") as fh:
                        fh.write(f"// File     : {rel_path}\n")
                        fh.write(f"// Function : {fn_name}\n")
                        fh.write("// " + "-" * 60 + "\n\n")
                        fh.write(body)
                except Exception as e:
                    self.log.emit(f"  ⚠ Could not save {safe_name}.txt: {e}")

                counts = count_constructs(body)

                # Weighted score
                score = sum(counts.get(cn, 0) * self.weights.get(cn, 1)
                            for cn, _ in CONSTRUCTS)
                level = complexity_level(score, self.bands)

                all_records.append({
                    "file_path":   rel_path,
                    "file_name":   file_name,
                    "function":    fn_name,
                    "counts":      counts,
                    "score":       score,
                    "level":       level,
                })

        self.log.emit(f"✅ Extracted {len(all_records)} functions — building Excel …")
        self.progress.emit(90, "Writing Excel report …")

        out_path = self._write_excel(all_records)
        self.progress.emit(100, "Done")
        self.log.emit(f"📊 Report saved: {out_path}")
        self.finished.emit(out_path)

    # ── Excel writer ──────────────────────────────────────────────────────────
    def _write_excel(self, records: list) -> str:
        wb = Workbook()

        # ── Sheet 1: Function Complexity ──────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Function_Complexity"

        construct_names = [c[0] for c in CONSTRUCTS]
        headers = ["File Name", "File Path", "Function Name"] + \
                  construct_names + ["Complexity Score", "Complexity Level"]

        thin    = Side(style="thin",   color="C5D8EC")
        thick   = Side(style="medium", color="1A3A5C")
        def bdr(): return Border(left=thin, right=thin, top=thin, bottom=thin)

        hdr_fill = PatternFill("solid", fgColor="1F4E78")
        hdr_font = Font(color="FFFFFF", bold=True, name="Arial", size=10)

        level_colors = {
            "Low":       "D6E4BC",
            "Medium":    "FFE699",
            "High":      "F4B183",
            "Very High": "FF7070",
            "Complex":   "CC0000",
        }
        level_font_dark = {"Complex"}

        for col_idx, hdr in enumerate(headers, start=1):
            cell = ws1.cell(row=1, column=col_idx, value=hdr)
            cell.font  = hdr_font
            cell.fill  = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = bdr()

        for row_idx, rec in enumerate(records, start=2):
            row_vals = [
                rec["file_path"],
                rec["file_name"],
                rec["function"],
            ] + [rec["counts"].get(cn, 0) for cn in construct_names] + [
                rec["score"],
                rec["level"],
            ]
            for col_idx, val in enumerate(row_vals, start=1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = bdr()
                # Colour the level cell
                if col_idx == len(headers):
                    lvl = rec["level"]
                    fg  = level_colors.get(lvl, "FFFFFF")
                    cell.fill = PatternFill("solid", fgColor=fg)
                    cell.font = Font(
                        bold=True,
                        color="FFFFFF" if lvl in level_font_dark else "1A1A1A",
                        name="Arial", size=10
                    )
                # Left-align text columns
                if col_idx <= 3:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto column widths
        for col in ws1.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws1.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)
        ws1.row_dimensions[1].height = 30
        ws1.freeze_panes = "A2"

        # ── Sheet 2: Construct Summary + Level/Compatibility Summary ─────────
        ws2 = wb.create_sheet("Construct_Summary")

        thin2  = Side(style="thin",   color="C5D8EC")
        thick2 = Side(style="medium", color="1A3A5C")
        def bdr2(left=None, right=None, top=None, bottom=None):
            return Border(left=left or thin2, right=right or thin2,
                          top=top or thin2,   bottom=bottom or thin2)

        hdr_fill2 = PatternFill("solid", fgColor="1F4E78")
        hdr_font2 = Font(color="FFFFFF", bold=True, name="Arial", size=10)
        sub_fill  = PatternFill("solid", fgColor="2E75B6")
        sub_font  = Font(color="FFFFFF", bold=True, name="Arial", size=11)

        # ── Summary block: Complexity Level Distribution ──────────────────────
        level_order  = ["Low", "Medium", "High", "Very High", "Complex"]
        level_counts = {lv: 0 for lv in level_order}
        for rec in records:
            lv = rec.get("level", "Unknown")
            if lv in level_counts:
                level_counts[lv] += 1
        total_fns = len(records)

        level_colors2 = {
            "Low":       "D6E4BC", "Medium":    "FFE699",
            "High":      "F4B183", "Very High": "FF7070", "Complex": "CC0000",
        }

        # Title row
        ws2.merge_cells("A1:C1")
        t = ws2.cell(1, 1, "📊 Complexity Level Summary")
        t.fill = sub_fill; t.font = sub_font
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 26

        # Header
        for ci, hdr in enumerate(["Complexity Level", "Function Count", "% of Total"], 1):
            c = ws2.cell(2, ci, hdr)
            c.fill = hdr_fill2; c.font = hdr_font2
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = bdr2(top=thick2, bottom=thick2,
                            left=thick2 if ci == 1 else thin2,
                            right=thick2 if ci == 3 else thin2)

        for ri, lv in enumerate(level_order, 3):
            cnt = level_counts[lv]
            pct = f"{cnt / total_fns * 100:.1f}%" if total_fns else "0.0%"
            fg  = level_colors2.get(lv, "FFFFFF")
            for ci, val in enumerate([lv, cnt, pct], 1):
                c = ws2.cell(ri, ci, val)
                c.fill = PatternFill("solid", fgColor=fg)
                c.font = Font(bold=True,
                              color="FFFFFF" if lv == "Complex" else "1A1A1A",
                              name="Arial", size=10)
                c.alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                        vertical="center")
                c.border = bdr2(left=thick2 if ci == 1 else thin2,
                                right=thick2 if ci == 3 else thin2,
                                bottom=thick2 if ri == 3 + len(level_order) - 1 else thin2)

        # Total row
        tot_row = 3 + len(level_order)
        for ci, val in enumerate(["Total", total_fns, "100%"], 1):
            c = ws2.cell(tot_row, ci, val)
            c.fill = PatternFill("solid", fgColor="1F4E78")
            c.font = Font(color="FFFFFF", bold=True, name="Arial", size=10)
            c.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
            c.border = bdr2(left=thick2 if ci == 1 else thin2,
                            right=thick2 if ci == 3 else thin2,
                            top=thick2, bottom=thick2)

        # ── Spacer ────────────────────────────────────────────────────────────
        spacer_row = tot_row + 2

        # ── Summary block: Compatibility Score Distribution ───────────────────
        # Define ranges
        compat_ranges = [
            ("0% – 24%  (Poor)",      0,  24,  "FFC7CE", "9C0006"),
            ("25% – 49%  (Low)",     25,  49,  "FFEB9C", "9C5700"),
            ("50% – 74%  (Medium)",  50,  74,  "FFEB9C", "9C5700"),
            ("75% – 89%  (Good)",    75,  89,  "C6EFCE", "276221"),
            ("90% – 100%  (Excellent)", 90, 100, "A9D18E", "1A3A00"),
        ]

        # Recompute compat scores per function
        construct_names2 = [c[0] for c in CONSTRUCTS]
        compat_range_counts = {r[0]: 0 for r in compat_ranges}
        for rec in records:
            available = [cn for cn in construct_names2 if rec["counts"].get(cn, 0) > 0]
            total_avail = sum(rec["counts"].get(cn, 0) for cn in available)
            # For standalone report, all constructs are "handled" by default
            compat_pct = 100.0 if total_avail == 0 else round(
                sum(rec["counts"].get(cn, 0) for cn in available) / total_avail * 100, 2
            )
            for label, lo, hi, _, _ in compat_ranges:
                if lo <= compat_pct <= hi:
                    compat_range_counts[label] += 1
                    break

        ws2.merge_cells(f"A{spacer_row}:C{spacer_row}")
        t2 = ws2.cell(spacer_row, 1, "🔗 Compatibility Score Distribution")
        t2.fill = sub_fill; t2.font = sub_font
        t2.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[spacer_row].height = 26

        hdr_row2 = spacer_row + 1
        for ci, hdr in enumerate(["Score Range", "Function Count", "% of Total"], 1):
            c = ws2.cell(hdr_row2, ci, hdr)
            c.fill = hdr_fill2; c.font = hdr_font2
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = bdr2(top=thick2, bottom=thick2,
                            left=thick2 if ci == 1 else thin2,
                            right=thick2 if ci == 3 else thin2)

        for ri2, (label, lo, hi, bg, fc) in enumerate(compat_ranges, hdr_row2 + 1):
            cnt = compat_range_counts[label]
            pct = f"{cnt / total_fns * 100:.1f}%" if total_fns else "0.0%"
            for ci, val in enumerate([label, cnt, pct], 1):
                c = ws2.cell(ri2, ci, val)
                c.fill = PatternFill("solid", fgColor=bg)
                c.font = Font(color=fc, bold=True, name="Arial", size=10)
                c.alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                        vertical="center")
                c.border = bdr2(left=thick2 if ci == 1 else thin2,
                                right=thick2 if ci == 3 else thin2,
                                bottom=thick2 if ri2 == hdr_row2 + len(compat_ranges) else thin2)

        tot_row2 = hdr_row2 + 1 + len(compat_ranges)
        for ci, val in enumerate(["Total", total_fns, "100%"], 1):
            c = ws2.cell(tot_row2, ci, val)
            c.fill = PatternFill("solid", fgColor="1F4E78")
            c.font = Font(color="FFFFFF", bold=True, name="Arial", size=10)
            c.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
            c.border = bdr2(left=thick2 if ci == 1 else thin2,
                            right=thick2 if ci == 3 else thin2,
                            top=thick2, bottom=thick2)

        # ── Spacer before construct table ─────────────────────────────────────
        construct_start = tot_row2 + 2

        # Title for construct table
        ws2.merge_cells(f"A{construct_start}:B{construct_start}")
        t3 = ws2.cell(construct_start, 1, "🔩 Construct-by-Construct Totals")
        t3.fill = sub_fill; t3.font = sub_font
        t3.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[construct_start].height = 24

        # Construct table headers
        ch_row = construct_start + 1
        for ci, hdr in enumerate(["Construct", "Total Count"], 1):
            c = ws2.cell(ch_row, ci, hdr)
            c.fill = hdr_fill2; c.font = hdr_font2
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = bdr2(top=thick2, bottom=thick2,
                            left=thick2 if ci == 1 else thin2,
                            right=thick2 if ci == 2 else thin2)

        totals = {cn: sum(r["counts"].get(cn, 0) for r in records) for cn in construct_names2}
        for ri3, cn in enumerate(construct_names2, ch_row + 1):
            c1 = ws2.cell(ri3, 1, cn)
            c1.alignment = Alignment(horizontal="left", vertical="center")
            c1.border = bdr2(left=thick2)
            c1.font = Font(name="Arial", size=10)
            c2 = ws2.cell(ri3, 2, totals[cn])
            c2.alignment = Alignment(horizontal="center", vertical="center")
            c2.border = bdr2(right=thick2)
            c2.font = Font(name="Arial", size=10)

        ws2.column_dimensions["A"].width = 34
        ws2.column_dimensions["B"].width = 18
        ws2.column_dimensions["C"].width = 16

        from datetime import datetime as _dt
        _ts = _dt.now().strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join(self.output_root, f"FuncAtlas_Complexity_Report_{_ts}.xlsx")
        wb.save(out_path)
        return out_path

# ── Append Worker (adds Sheet 3 to existing FuncAtlas_Report.xlsx) ───────────
class ComplexityAppendWorker(QObject):
    """
    Scans source files, computes complexity, then appends
    'Complexity_Compatibility' as Sheet 3 to an existing FuncAtlas_Report.xlsx.
    """
    progress = Signal(int, str)
    log      = Signal(str)
    finished = Signal(str)   # path to updated Excel
    error    = Signal(str)

    def __init__(self, report_path: str, source_folder: str,
                 weights: list = None, bands: list = None,
                 handled_scenarios=None):
        super().__init__()
        self.report_path   = report_path
        self.source_folder = normalize_path(source_folder)
        self.weights       = {n: w for n, w in (weights or [])} or DEFAULT_WEIGHTS
        self.bands         = bands or DEFAULT_BANDS
        self.handled_scenarios = set(handled_scenarios) if handled_scenarios else set()
        self._cancel_requested = False

    def run(self):
        try:
            self._run()
        except Exception as exc:
            import traceback
            self.error.emit(f"{exc}\n{traceback.format_exc()}")

    def _run(self):
        from openpyxl import load_workbook

        if not os.path.isfile(self.report_path):
            self.error.emit(f"Report file not found:\n{self.report_path}")
            return

        # ── Read Sheet 1 to collect ALL New / Reuse (Modified) rows ────────────
        # We drive Sheet 3 directly from Sheet 1 rows so the row count and
        # file/path values are identical.  Key: (fn_name_lower, file_path_lower)
        # so same-named functions in different source files stay distinct.
        wb_check = load_workbook(self.report_path, read_only=True, data_only=True)
        # sheet1_rows: list of (fn_display, file_name, file_path) for New/RM rows
        sheet1_rows = []
        # include_set: set of (fn_lower, filepath_lower) — for source lookup
        include_set = set()
        if "Function_Match_Report" in wb_check.sheetnames:
            ws_check = wb_check["Function_Match_Report"]
            for row in ws_check.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                fn_name  = str(row[1] or "").strip()   # col B = Function Name
                fname    = str(row[0] or "").strip()   # col A = File Name
                fpath    = str(row[2] or "").strip()   # col C = Target File Path
                status   = None
                for cell_val in reversed(row):
                    if cell_val in ("Reuse", "New", "Reuse (Modified)"):
                        status = cell_val
                        break
                if status in ("New", "Reuse (Modified)") and fn_name:
                    sheet1_rows.append((fn_name, fname, fpath))
                    include_set.add((normalize_name(fn_name), fpath.lower()))
        wb_check.close()

        if not sheet1_rows:
            self.error.emit(
                "No 'New' or 'Reuse (Modified)' functions found in the report.\n"
                "All functions may be 100% Reuse — nothing to analyse."
            )
            return

        self.log.emit(f"📋 {len(sheet1_rows)} functions to analyse "
                      f"(New + Reuse Modified only) …")

        # ── Scan source files and build a lookup: (fn_lower, filepath_lower) -> body
        # We need to match each Sheet 1 row to the correct source file body.
        file_entries = list(iter_source_files(self.source_folder))
        if not file_entries:
            self.error.emit("No source files found in the target source folder.")
            return

        self.log.emit(f"🔍 Scanning {len(file_entries)} source files …")

        # body_lookup: (fn_lower, full_path_lower) -> body_text
        # Also keep a name-only fallback: fn_lower -> [(full_path, body)]
        body_lookup   = {}   # (fn_lower, full_path_lower) -> body
        body_fallback = {}   # fn_lower -> [(full_path, body)]

        for idx, (full_path, file_name) in enumerate(file_entries):
            if self._cancel_requested:
                self.error.emit('__CANCELLED__')
                return
            pct = int((idx / len(file_entries)) * 75)
            self.progress.emit(pct, f"Scanning {file_name} …")
            functions = detect_functions_in_file(full_path)
            if not functions:
                continue
            for fn_name in functions:
                fn_key = normalize_name(fn_name)
                body   = extract_function_body(full_path, fn_name)
                body_lookup[(fn_key, full_path.lower())] = body
                body_fallback.setdefault(fn_key, []).append((full_path, body))

        # ── Build records list in Sheet 1 order ─────────────────────────────
        # For each Sheet 1 row, find the matching body via composite key first,
        # then fall back to name-only (picks the first matching source file).
        records = []
        for fn_display, s1_fname, s1_fpath in sheet1_rows:
            fn_key = normalize_name(fn_display)
            body   = None

            # Try composite match: function name + any source file whose path
            # ends with the same basename as the Sheet 1 file path
            s1_basename = os.path.basename(s1_fpath).lower()
            for (bk_fn, bk_path_lower), bk_body in body_lookup.items():
                if bk_fn == fn_key and os.path.basename(bk_path_lower) == s1_basename:
                    body = bk_body
                    break

            # Fallback: name-only (first occurrence in source)
            if body is None:
                candidates = body_fallback.get(fn_key, [])
                if candidates:
                    body = candidates[0][1]

            if body is None:
                body = ""

            counts = count_constructs(body)
            score  = sum(counts.get(cn, 0) * self.weights.get(cn, 1)
                         for cn, _ in CONSTRUCTS)
            level  = complexity_level(score, self.bands)
            records.append({
                "function":  fn_display,
                "file_name": s1_fname,
                "file_path": s1_fpath,
                "counts":    counts,
                "score":     score,
                "level":     level,
            })

        self.log.emit(f"✅ {len(records)} functions processed — appending Sheet 3 …")
        self.progress.emit(85, "Appending Complexity_Compatibility sheet …")

        # ── Load existing workbook and append Sheet 3 ────────────────────────
        wb = load_workbook(self.report_path)

        # Remove old Sheet 3 if it already exists
        if "Complexity_Compatibility" in wb.sheetnames:
            del wb["Complexity_Compatibility"]

        ws3 = wb.create_sheet("Complexity_Compatibility")

        thin  = Side(style="thin",   color="C5D8EC")
        thick = Side(style="medium", color="1A3A5C")
        def bdr(left=None, right=None, top=None, bottom=None):
            return Border(left=left or thin, right=right or thin,
                          top=top or thin,   bottom=bottom or thin)

        hdr_fill  = PatternFill("solid", fgColor="1F4E78")
        hdr_font  = Font(color="FFFFFF", bold=True, name="Arial", size=10)
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        pct_align = Alignment(horizontal="center", vertical="center")
        left_align= Alignment(horizontal="left",   vertical="center")
        sno_fill  = PatternFill("solid", fgColor="EBF3FB")
        data_font = lambda: Font(name="Arial", size=10)

        green_fill  = PatternFill("solid", fgColor="C6EFCE")
        green_font  = lambda: Font(color="276221", name="Arial", size=10, bold=True)
        yellow_fill = PatternFill("solid", fgColor="FFEB9C")
        yellow_font = lambda: Font(color="9C5700", name="Arial", size=10, bold=True)
        red_fill    = PatternFill("solid", fgColor="FFC7CE")
        red_font    = lambda: Font(color="9C0006", name="Arial", size=10, bold=True)

        level_colors = {
            "Low":       "D6E4BC",
            "Medium":    "FFE699",
            "High":      "F4B183",
            "Very High": "FF7070",
            "Complex":   "CC0000",
        }

        construct_names = [c[0] for c in CONSTRUCTS]
        headers = (
            ["File Name", "Function Name", "Target File Path"]
            + construct_names
            + ["Complexity Score", "Complexity Level"]
        )
        last_col = len(headers)

        for c_idx, h in enumerate(headers, 1):
            cell = ws3.cell(1, c_idx, h)
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = hdr_align
            cell.border    = bdr(
                left=thick  if c_idx == 1       else thin,
                right=thick if c_idx == last_col else thin,
                top=thick, bottom=thick
            )
        ws3.row_dimensions[1].height = 44

        for r, rec in enumerate(records, 2):
            col = 1

            # File Name (first column)
            c = ws3.cell(r, col, rec["file_name"]); col += 1
            c.fill = sno_fill; c.font = Font(name="Arial", size=10, color="2C5F8A")
            c.alignment = left_align; c.border = bdr(left=thick)

            # Function Name
            c = ws3.cell(r, col, rec["function"]); col += 1
            c.font = Font(bold=True, name="Arial", size=10)
            c.alignment = left_align; c.border = bdr()

            # Target File Path
            c = ws3.cell(r, col, rec["file_path"]); col += 1
            c.font = data_font(); c.alignment = left_align; c.border = bdr()

            # Construct counts
            for cn in construct_names:
                cnt = rec["counts"].get(cn, 0)
                c = ws3.cell(r, col, cnt); col += 1
                c.alignment = pct_align; c.border = bdr(); c.font = data_font()
                if cnt > 0:
                    c.fill = PatternFill("solid", fgColor="EBF3FB")

            # Complexity Score
            c = ws3.cell(r, col, rec["score"]); col += 1
            c.alignment = pct_align; c.border = bdr()
            c.font = Font(bold=True, name="Arial", size=10)

            # Complexity Level
            c = ws3.cell(r, col, rec["level"]); col += 1
            c.alignment = pct_align
            c.border = bdr(right=thick)
            c.fill = PatternFill("solid", fgColor=level_colors.get(rec["level"], "FFFFFF"))
            c.font = Font(bold=True,
                          color="FFFFFF" if rec["level"] == "Complex" else "1A1A1A",
                          name="Arial", size=10)

        # Column widths
        ws3.column_dimensions["A"].width = 28  # File Name
        ws3.column_dimensions["B"].width = 36  # Function Name
        ws3.column_dimensions["C"].width = 50  # Target File Path
        for ci in range(4, 4 + len(construct_names)):
            ws3.column_dimensions[get_column_letter(ci)].width = 20
        score_col = 4 + len(construct_names)
        ws3.column_dimensions[get_column_letter(score_col)].width = 18
        ws3.column_dimensions[get_column_letter(score_col + 1)].width = 18
        ws3.freeze_panes = "A2"

        # ── Sheet 4: Compatibility Score ───────────────────────────────────────────────
        if "Compatibility_Score" in wb.sheetnames:
            del wb["Compatibility_Score"]

        ws4 = wb.create_sheet("Compatibility_Score")

        compat_hdrs = [
            "File Name", "Function Name", "File Path",
            "Available Scenarios", "Handled Scenarios",
            "Unhandled Scenarios", "Compatibility %"
        ]
        for c_idx, h in enumerate(compat_hdrs, 1):
            cell = ws4.cell(1, c_idx, h)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = hdr_align
            cell.border = bdr(left=thick if c_idx == 1 else thin,
                               right=thick if c_idx == len(compat_hdrs) else thin,
                               top=thick, bottom=thick)
        ws4.row_dimensions[1].height = 32

        handled = self.handled_scenarios
        for r2, rec in enumerate(records, 2):
            # available = constructs detected in this function (count > 0)
            available = [cn for cn in construct_names if rec["counts"].get(cn, 0) > 0]
            # handled_in_fn = intersection of available and user-marked handled
            handled_in_fn = [cn for cn in available if cn in handled]
            unhandled_in_fn = [cn for cn in available if cn not in handled]
            # Score = total appearances of handled scenarios /
            #         total appearances of available scenarios * 100
            total_avail_appearances   = sum(rec["counts"].get(cn, 0) for cn in available)
            total_handled_appearances = sum(rec["counts"].get(cn, 0) for cn in handled_in_fn)
            compat_pct = round(
                (total_handled_appearances / total_avail_appearances * 100), 2
            ) if total_avail_appearances > 0 else 0.0

            row_data = [
                rec["file_name"],
                rec["function"],
                rec["file_path"],
                ", ".join(available) if available else "None",
                ", ".join(handled_in_fn) if handled_in_fn else "None",
                ", ".join(unhandled_in_fn) if unhandled_in_fn else "None",
                compat_pct,
            ]
            for c_idx2, val in enumerate(row_data, 1):
                cell2 = ws4.cell(r2, c_idx2, val)
                cell2.border = bdr(left=thick if c_idx2 == 1 else thin,
                                   right=thick if c_idx2 == len(compat_hdrs) else thin)
                cell2.alignment = left_align if c_idx2 <= 3 else pct_align
                cell2.font = data_font()
                # Colour compatibility percentage cell
                if c_idx2 == len(compat_hdrs):
                    if compat_pct >= 75:
                        cell2.fill = green_fill; cell2.font = green_font()
                    elif compat_pct >= 40:
                        cell2.fill = yellow_fill; cell2.font = yellow_font()
                    else:
                        cell2.fill = red_fill; cell2.font = red_font()
                    cell2.value = f"{compat_pct:.1f}%"

        ws4.column_dimensions["A"].width = 26
        ws4.column_dimensions["B"].width = 36
        ws4.column_dimensions["C"].width = 50
        ws4.column_dimensions["D"].width = 50
        ws4.column_dimensions["E"].width = 50
        ws4.column_dimensions["F"].width = 50
        ws4.column_dimensions["G"].width = 20
        ws4.freeze_panes = "A2"

        # ── Rebuild / update Sheet 2 summary ────────────────────────────────
        SUM_SHEET = "Construct_Summary"
        if SUM_SHEET in wb.sheetnames:
            del wb[SUM_SHEET]
        ws2 = wb.create_sheet(SUM_SHEET, 1)   # insert as 2nd sheet

        thin2  = Side(style="thin",   color="C5D8EC")
        thick2 = Side(style="medium", color="1A3A5C")
        def bdr2(left=None, right=None, top=None, bottom=None):
            return Border(left=left or thin2, right=right or thin2,
                          top=top or thin2,   bottom=bottom or thin2)

        hdr_fill2 = PatternFill("solid", fgColor="1F4E78")
        hdr_font2 = Font(color="FFFFFF", bold=True, name="Arial", size=10)
        sub_fill2 = PatternFill("solid", fgColor="2E75B6")
        sub_font2 = Font(color="FFFFFF", bold=True, name="Arial", size=11)

        level_order2 = ["Low", "Medium", "High", "Very High", "Complex"]
        level_counts2 = {lv: 0 for lv in level_order2}
        for rec in records:
            lv = rec.get("level", "Unknown")
            if lv in level_counts2:
                level_counts2[lv] += 1
        total_fns2 = len(records)

        level_colors2 = {
            "Low": "D6E4BC", "Medium": "FFE699",
            "High": "F4B183", "Very High": "FF7070", "Complex": "CC0000",
        }

        # Complexity summary title
        ws2.merge_cells("A1:C1")
        t = ws2.cell(1, 1, "📊 Complexity Level Summary")
        t.fill = sub_fill2; t.font = sub_font2
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 26

        for ci, h in enumerate(["Complexity Level", "Function Count", "% of Total"], 1):
            c = ws2.cell(2, ci, h)
            c.fill = hdr_fill2; c.font = hdr_font2
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = bdr2(top=thick2, bottom=thick2,
                            left=thick2 if ci == 1 else thin2,
                            right=thick2 if ci == 3 else thin2)

        for ri, lv in enumerate(level_order2, 3):
            cnt = level_counts2[lv]
            pct = f"{cnt / total_fns2 * 100:.1f}%" if total_fns2 else "0.0%"
            fg  = level_colors2.get(lv, "FFFFFF")
            for ci, val in enumerate([lv, cnt, pct], 1):
                c = ws2.cell(ri, ci, val)
                c.fill = PatternFill("solid", fgColor=fg)
                c.font = Font(bold=True, color="FFFFFF" if lv == "Complex" else "1A1A1A",
                              name="Arial", size=10)
                c.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
                c.border = bdr2(left=thick2 if ci == 1 else thin2,
                                right=thick2 if ci == 3 else thin2)

        tot_row_s2 = 3 + len(level_order2)
        for ci, val in enumerate(["Total", total_fns2, "100%"], 1):
            c = ws2.cell(tot_row_s2, ci, val)
            c.fill = PatternFill("solid", fgColor="1F4E78")
            c.font = Font(color="FFFFFF", bold=True, name="Arial", size=10)
            c.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
            c.border = bdr2(left=thick2 if ci == 1 else thin2,
                            right=thick2 if ci == 3 else thin2, top=thick2, bottom=thick2)

        # Compatibility summary
        compat_ranges2 = [
            ("0% – 24%  (Poor)",         0,  24,  "FFC7CE", "9C0006"),
            ("25% – 49%  (Low)",        25,  49,  "FFEB9C", "9C5700"),
            ("50% – 74%  (Medium)",     50,  74,  "FFEB9C", "9C5700"),
            ("75% – 89%  (Good)",       75,  89,  "C6EFCE", "276221"),
            ("90% – 100%  (Excellent)", 90, 100,  "A9D18E", "1A3A00"),
        ]
        compat_range_counts2 = {r[0]: 0 for r in compat_ranges2}

        # Recompute per-function compat from records using handled_scenarios
        for rec in records:
            available2 = [cn for cn in construct_names if rec["counts"].get(cn, 0) > 0]
            handled_in_fn2   = [cn for cn in available2 if cn in handled]
            total_avail2     = sum(rec["counts"].get(cn, 0) for cn in available2)
            total_handled2   = sum(rec["counts"].get(cn, 0) for cn in handled_in_fn2)
            cp = round(total_handled2 / total_avail2 * 100, 2) if total_avail2 > 0 else 0.0
            for label2, lo2, hi2, _, _ in compat_ranges2:
                if lo2 <= cp <= hi2:
                    compat_range_counts2[label2] += 1
                    break

        spacer2 = tot_row_s2 + 2
        ws2.merge_cells(f"A{spacer2}:C{spacer2}")
        t2 = ws2.cell(spacer2, 1, "🔗 Compatibility Score Distribution")
        t2.fill = sub_fill2; t2.font = sub_font2
        t2.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[spacer2].height = 26

        hdr_r2 = spacer2 + 1
        for ci, h in enumerate(["Score Range", "Function Count", "% of Total"], 1):
            c = ws2.cell(hdr_r2, ci, h)
            c.fill = hdr_fill2; c.font = hdr_font2
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = bdr2(top=thick2, bottom=thick2,
                            left=thick2 if ci == 1 else thin2,
                            right=thick2 if ci == 3 else thin2)

        for ri2, (label2, lo2, hi2, bg2, fc2) in enumerate(compat_ranges2, hdr_r2 + 1):
            cnt2 = compat_range_counts2[label2]
            pct2 = f"{cnt2 / total_fns2 * 100:.1f}%" if total_fns2 else "0.0%"
            for ci, val in enumerate([label2, cnt2, pct2], 1):
                c = ws2.cell(ri2, ci, val)
                c.fill = PatternFill("solid", fgColor=bg2)
                c.font = Font(color=fc2, bold=True, name="Arial", size=10)
                c.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
                c.border = bdr2(left=thick2 if ci == 1 else thin2,
                                right=thick2 if ci == 3 else thin2)

        tot_row_compat2 = hdr_r2 + 1 + len(compat_ranges2)
        for ci, val in enumerate(["Total", total_fns2, "100%"], 1):
            c = ws2.cell(tot_row_compat2, ci, val)
            c.fill = PatternFill("solid", fgColor="1F4E78")
            c.font = Font(color="FFFFFF", bold=True, name="Arial", size=10)
            c.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
            c.border = bdr2(left=thick2 if ci == 1 else thin2,
                            right=thick2 if ci == 3 else thin2, top=thick2, bottom=thick2)

        # Construct totals
        construct_start2 = tot_row_compat2 + 2
        ws2.merge_cells(f"A{construct_start2}:B{construct_start2}")
        t3 = ws2.cell(construct_start2, 1, "🔩 Construct-by-Construct Totals")
        t3.fill = sub_fill2; t3.font = sub_font2
        t3.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[construct_start2].height = 24

        ch_r2 = construct_start2 + 1
        for ci, h in enumerate(["Construct", "Total Count"], 1):
            c = ws2.cell(ch_r2, ci, h)
            c.fill = hdr_fill2; c.font = hdr_font2
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = bdr2(top=thick2, bottom=thick2,
                            left=thick2 if ci == 1 else thin2,
                            right=thick2 if ci == 2 else thin2)

        totals2 = {cn: sum(r["counts"].get(cn, 0) for r in records) for cn in construct_names}
        for ri3, cn in enumerate(construct_names, ch_r2 + 1):
            c1 = ws2.cell(ri3, 1, cn)
            c1.alignment = Alignment(horizontal="left", vertical="center")
            c1.border = bdr2(left=thick2)
            c1.font = Font(name="Arial", size=10)
            c2 = ws2.cell(ri3, 2, totals2[cn])
            c2.alignment = Alignment(horizontal="center", vertical="center")
            c2.border = bdr2(right=thick2)
            c2.font = Font(name="Arial", size=10)

        ws2.column_dimensions["A"].width = 34
        ws2.column_dimensions["B"].width = 18
        ws2.column_dimensions["C"].width = 16

        wb.save(self.report_path)
        wb.close()

        self.progress.emit(100, "Done")
        self.log.emit(f"📊 Sheets 3 & 4 appended: {self.report_path}")
        self.finished.emit(self.report_path)